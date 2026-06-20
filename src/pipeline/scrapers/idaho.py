"""
scrapers/idaho.py — Download Idaho campaign finance data from THREE distinct
sources/eras. Pure HTTP/requests for all three — no Playwright needed.

Source 1 — current portal, 2020+ (sunshine.voteidaho.gov):
    Unauthenticated REST API at api-sunshine.voteidaho.gov, discovered by
    capturing the portal's own network calls.
    relation_type: portal_candidates, portal_committees
        — full pull, year="all" (one CSV row per candidate/committee per
          election cycle; the source itself doesn't split by year).
    relation_type: portal_contributions (TCON), portal_expenditures (TEXP)
        — year-split, 2020..current_year+1. TCON includes contributions,
          in-kind, loans received/forgiven, and outstanding loans (so
          loans_debts.csv is derived from this file by the parser).
        — 2020-2022 return near-empty files (header only); real data starts
          2023. Pulled anyway for completeness/manifest consistency.

Source 2 — electionstats platform, ~2008-2024 (canvass.sos.idaho.gov):
    Unauthenticated POST to https://canvass.sos.idaho.gov/eng/finances/get_activity.json
    (NOT the sunshine.sos.idaho.gov host previously tried). Enriched
    activity-log records (activity_type: file/donate/spend) with inline
    from_/to_ entity name/type/address/office/district/party/treasurer/status.
    relation_type: electionstats_activity — full pull, year="all".
        Pagination by `page` (1..N), limit hard-capped at 1000/page.
        382,124 total rows => 383 pages (last page has 124 rows). RE-VERIFIED
        2026-06-13: plain {"page": N, "limit": 1000} requests against
        canvass.sos.idaho.gov return DISTINCT rows per page (confirmed page 1
        vs page 2 return different `id`s, and page 383 returns exactly 124
        rows = 382124 - 382*1000) — the earlier "static blob, filters
        ignored" finding was against a different host and does not apply
        here. No date chunking needed; plain page-based pagination works.
    Per Henry's decision (2026-06-11): pull this IN FULL even though it
    overlaps with Source 1's 2023 data and Source 3's pre-2018 data. Dedup is
    the parser's job.
    This full pull also covers the donate->PAC / spend->Candidate /
    spend->PAC activity for 2020-2022 that was previously only available via
    the third-party id.electionstats.com static export (see below) —
    confirmed via the sunshine.sos.idaho.gov UI, which calls this same
    backend and shows $20.16M donated to 257 PACs + $19.86M spent by 266 PACs
    for 2020-2022.

Source 2 LEGACY (REMOVED 2026-06-13): a separate `--legacy` mode used to
    download a static cached CSV from the third-party site
    id.electionstats.com (donate->Candidate activity, 2020-2022, 100,936
    rows). Per Henry's ".gov-only data sources" policy and now that the
    canvass.sos.idaho.gov full pull above covers the same 2020-2022 activity
    (and more, across all activity types), this third-party fetch has been
    removed entirely. No non-.gov sources remain in this scraper.

Source 3 — static archive, 2000-2018 biennium files (archive.sos.idaho.gov):
    Direct Excel downloads, no API. Confirmed inventory (2026-06-11):
        2000, 2002            : cand_contributions.xls / comm_contributions.xls /
                                 cand_expenditures.xls / comm_expenditures.xls
        2004, 2006, 2008, 2010: candcont.xls / commcont.xls / candexp.xls / commexp.xls
        2012                  : 2012_cand_cont.xls / 2012_comm_cont.xls /
                                 2012_cand_exp.xls / 2012_comm_exp.xls
        2014, 2016            : {y}_cand_cont.xlsx / {y}_comm_cont.xlsx /
                                 {y}_cand_exp.xlsx / {y}_comm_exp.xlsx
        2018                  : 2018_cand_cont.xlsx / 2018_cand_expend.xlsx ONLY
                                 (no committee-level bulk files exist for 2018 —
                                 documented gap, see idaho_pending memory).
    relation_type: archive_cand_cont, archive_comm_cont, archive_cand_exp,
                   archive_comm_exp — year = biennium (2000,2002,...,2018).

Source 4 — legacy 2019 per-filer PDF reports (archive.sos.idaho.gov):
    Idaho's archive program has no 2020 biennium file (Source 3 stops at
    2018), and the 2020-2022 electionstats feed (Source 2) starts
    2020-01-01 — leaving 2019 almost entirely uncovered. However,
    archive.sos.idaho.gov/ELECT/Finance/2020/index.html has a "2019 Reports"
    section with 5 index pages listing one C-2 "Campaign Financial
    Disclosure Report" PDF per filer (First Annual / Mid-Year, some
    _amended/_terminated variants). These PDFs are text-extractable
    (not scans) and contain itemized Schedule A (contributions), Schedule B
    (expenditures), Schedule C (in-kind), and Schedule D (loans).
    relation_type: legacy_2019_pdfs — one-time full pull, year="2019".
        461 unique PDF links across the 5 index pages:
        last_year_mid_year_report.html, last_year_statewide_judicial_and_undeclared.html,
        last_year_state_legislature.html, last_year_party_committees.html,
        last_year_measure_and_miscellaneous_committees.html
        (all under https://archive.sos.idaho.gov/ELECT/Finance/2020/).
        Downloaded into data/Idaho/raw/legacy_2019_pdfs/, one file per link,
        named by flattening the relative path (e.g. "First%20Annual/10817.pdf"
        -> "First_Annual_10817.pdf"). Resumable: the link list is cached to
        legacy_2019_pdfs/_links.json on first call; subsequent calls skip
        any output file that already exists on disk.

Raw files (data/Idaho/raw/):
    portal_candidates.csv, portal_committees.csv
    portal_contributions_{year}.csv, portal_expenditures_{year}.csv
    electionstats_activity.csv
    archive_{year}_{cand_cont|comm_cont|cand_exp|comm_exp}.{xls|xlsx}
    legacy_2019_pdfs/{First_Annual|Mid-Year}_{filer_id}[_amended|_terminated...].pdf
"""

import csv
import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import unquote

import requests

PROJECT_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(PROJECT_ROOT))
from src.reporting.logger import get_logger

# =============================== paths ================================
RAW_DIR  = PROJECT_ROOT / "data" / "Idaho" / "raw"
MANIFEST = PROJECT_ROOT / "data" / "Idaho" / "manifest.csv"
RAW_DIR.mkdir(parents=True, exist_ok=True)

MANIFEST_COLS = ["relation_type", "year", "filename", "row_count"]

EARLIEST_PORTAL_YEAR = 2020


# ========================== Manifest helpers ==========================
def load_manifest() -> set[tuple[str, str]]:
    """Return set of (relation_type, year) already recorded in the manifest."""
    done: set[tuple[str, str]] = set()
    if not MANIFEST.exists():
        return done
    with open(MANIFEST, newline="") as f:
        for row in csv.DictReader(f):
            done.add((row["relation_type"], row["year"]))
    return done


def strip_manifest(keep_fn) -> None:
    """Rewrite the manifest keeping only rows where keep_fn(row) is True."""
    if not MANIFEST.exists():
        return
    with open(MANIFEST, newline="") as f:
        rows = list(csv.DictReader(f))
    with open(MANIFEST, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=MANIFEST_COLS)
        writer.writeheader()
        writer.writerows(r for r in rows if keep_fn(r))


def upsert_manifest(record: dict) -> None:
    """Add or overwrite the manifest entry matching (relation_type, year)."""
    existing = []
    if MANIFEST.exists():
        with open(MANIFEST, newline="") as f:
            existing = [
                r for r in csv.DictReader(f)
                if not (r["relation_type"] == record["relation_type"]
                        and r["year"] == record["year"])
            ]
    with open(MANIFEST, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=MANIFEST_COLS)
        writer.writeheader()
        writer.writerows(existing)
        writer.writerow(record)


# =================== Source 1: current portal (REST API) ===================
PORTAL_BASE   = "https://api-sunshine.voteidaho.gov/api"
PORTAL_ORIGIN = "https://sunshine.voteidaho.gov"

CAND_FILTER = {
    "pageNumber": 1, "pageSize": 50, "filerTypeCode": "CAN",
    "filerName": None, "filingEntityId": None, "politicalPartyCode": None,
    "OfficeSought": None, "totalRaisedMax": None, "totalRaisedMin": None,
    "totalSpentMax": None, "totalSpentMin": None, "balanceFundsMax": None,
    "balanceFundsMin": None, "accountStatus": None, "electionYear": None,
}

COMM_FILTER = {
    "pageNumber": 1, "pageSize": 50, "filerTypeCode": "COM",
    "filerName": None, "filingEntityId": None, "politicalPartyCode": None,
    "committeeType": None, "filingYear": None, "chairPersonName": None,
    "totalRaisedMax": None, "totalRaisedMin": None,
    "totalSpentMax": None, "totalSpentMin": None, "balanceFundsMax": None,
    "balanceFundsMin": None, "accountStatus": None, "electionYear": None,
}


def _portal_post(path: str, body: dict, referer_suffix: str,
                  timeout: int = 120, retries: int = 4) -> requests.Response:
    headers = {
        "Content-Type": "application/json",
        "Origin":       PORTAL_ORIGIN,
        "Referer":      f"{PORTAL_ORIGIN}/public/cf/{referer_suffix}",
    }
    last_err = None
    for attempt in range(retries):
        try:
            resp = requests.post(f"{PORTAL_BASE}{path}", json=body,
                                  headers=headers, timeout=timeout)
            resp.raise_for_status()
            return resp
        except Exception as e:
            last_err = e
            time.sleep(2 * (attempt + 1))
    raise last_err


def fetch_portal_candidates(out_path: Path) -> int:
    body = {"publicGridName": "CandidatePublicGrid",
            "candidateCommitteeSearchFilter": CAND_FILTER}
    resp = _portal_post("/ExportData/DownloadPublicGridData", body, "publiccandidate")
    text = resp.text
    out_path.write_text(text, encoding="utf-8")
    # Row 1 = "Candidate Download as of ..." metadata, row 2 = real header
    return max(0, text.count("\n") - 2)


def fetch_portal_committees(out_path: Path) -> int:
    body = {"publicGridName": "CommitteePublicGrid",
            "candidateCommitteeSearchFilter": COMM_FILTER}
    resp = _portal_post("/ExportData/DownloadPublicGridData", body, "publiccommitte")
    text = resp.text
    out_path.write_text(text, encoding="utf-8")
    return max(0, text.count("\n") - 2)


def fetch_portal_transactions(txn_type: str, year: int, out_path: Path) -> int:
    """txn_type: 'TCON' (contributions/loans) or 'TEXP' (expenditures)."""
    body = {"transactionTypeCode": txn_type, "type": "CSV",
            "filingYear": str(year), "openInNewTab": False}
    resp = _portal_post("/ExportData/GetExportPublicDownloadData", body,
                         "publiccandidate", timeout=180)
    content = resp.content
    out_path.write_bytes(content)
    text = content.decode("latin-1", errors="replace")
    return max(0, text.count("\n") - 1)


# ================= Source 2: electionstats activity feed =================
# canvass.sos.idaho.gov (NOT sunshine.sos.idaho.gov — that host's
# get_activity.json returns a static cached blob regardless of filters/page).
# Re-verified 2026-06-13: this host correctly paginates and respects `dates`.
ES_URL       = "https://canvass.sos.idaho.gov/eng/finances/get_activity.json"
ES_PAGE_SIZE = 1000

ES_FIELDS = [
    "id", "activity_type", "pk_id", "date", "datetime", "amount",
    "from_pk_id", "from_display_name", "from_entity_type", "from_address",
    "from_city", "from_state", "from_zip", "from_is_dupe",
    "from_office_id", "from_office_name", "from_district_name",
    "from_party_code", "from_treasurer_id", "from_treasurer_name",
    "from_reg_district", "from_status",
    "to_pk_id", "to_display_name", "to_entity_type", "to_address",
    "to_city", "to_state", "to_zip", "to_is_dupe",
    "to_office_id", "to_office_name", "to_district_name",
    "to_party_code", "to_treasurer_id", "to_treasurer_name",
    "to_reg_district", "to_status",
    "elec_year", "elec_stage", "donate_type", "spend_type",
    "donate_count", "spend_count", "spend_code",
    "report_id", "report_code", "report_name", "report_status",
    "report_submit_date", "report_due_date", "r_cash", "r_loan", "r_debt",
    "report2_id", "report2_code", "report2_name", "report2_status",
    "report2_submit_date", "report2_due_date",
    "o_amount_desc", "o_amount_asc", "o_date_desc", "o_date_asc",
]


def _es_request(page: int, retries: int = 4) -> list[dict]:
    last_err = None
    for attempt in range(retries):
        try:
            resp = requests.post(ES_URL, json={"page": page, "limit": ES_PAGE_SIZE},
                                  headers={"Content-Type": "application/json"},
                                  timeout=60)
            resp.raise_for_status()
            data = resp.json()
            return data.get("output") or []
        except Exception as e:
            last_err = e
            time.sleep(2 * (attempt + 1))
    raise last_err


ES_TIME_BUDGET = 38.0  # seconds per call — leaves headroom under the 45s bash tool limit


def fetch_electionstats_activity(out_path: Path, log=None, time_budget: float = ES_TIME_BUDGET) -> dict:
    """Page through the electionstats activity feed. ~383 pages of 1000.

    RESUMABLE: progress (the last fully-written page number) is tracked in a
    sidecar file `<out_path>.progress`. Each call works for up to
    `time_budget` seconds — appending rows to out_path and updating the
    sidecar after each page — then returns. Call repeatedly until the
    returned dict has done=True.

    This exists because the full pull (~383 pages) takes several minutes,
    far longer than a single foreground invocation should run. On done=True
    the sidecar file is removed.

    Returns {"done": bool, "next_page": int | None, "rows_written": int}
    rows_written counts only rows written during THIS call.
    """
    progress_path = out_path.with_name(out_path.name + ".progress")

    start_page = 1
    if progress_path.exists() and out_path.exists():
        start_page = int(progress_path.read_text().strip()) + 1

    write_header = start_page == 1
    mode = "w" if write_header else "a"

    rows_written = 0
    done = False
    page = start_page
    t0 = time.monotonic()

    with open(out_path, mode, newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=ES_FIELDS, extrasaction="ignore", restval="")
        if write_header:
            writer.writeheader()

        while True:
            if time.monotonic() - t0 > time_budget:
                break
            rows = _es_request(page)
            if not rows:
                done = True
                break
            writer.writerows(rows)
            f.flush()
            rows_written += len(rows)
            progress_path.write_text(str(page))
            if log and page % 20 == 0:
                log.info(f"    electionstats page {page} done ({rows_written:,} rows this call)")
            if len(rows) < ES_PAGE_SIZE:
                done = True
                break
            page += 1
            time.sleep(0.15)

    if done:
        progress_path.unlink(missing_ok=True)

    return {"done": done, "next_page": None if done else page, "rows_written": rows_written}


# =================== Source 3: static archive (Excel files) ===================
ARCHIVE_BASE = "https://archive.sos.idaho.gov/ELECT/Finance"

# year -> {relation_type: filename_on_server}
ARCHIVE_FILES: dict[int, dict[str, str]] = {
    2000: {"archive_cand_cont": "cand_contributions.xls",
           "archive_comm_cont": "comm_contributions.xls",
           "archive_cand_exp":  "cand_expenditures.xls",
           "archive_comm_exp":  "comm_expenditures.xls"},
    2002: {"archive_cand_cont": "cand_contributions.xls",
           "archive_comm_cont": "comm_contributions.xls",
           "archive_cand_exp":  "cand_expenditures.xls",
           "archive_comm_exp":  "comm_expenditures.xls"},
    2004: {"archive_cand_cont": "candcont.xls", "archive_comm_cont": "commcont.xls",
           "archive_cand_exp": "candexp.xls",   "archive_comm_exp": "commexp.xls"},
    2006: {"archive_cand_cont": "candcont.xls", "archive_comm_cont": "commcont.xls",
           "archive_cand_exp": "candexp.xls",   "archive_comm_exp": "commexp.xls"},
    2008: {"archive_cand_cont": "candcont.xls", "archive_comm_cont": "commcont.xls",
           "archive_cand_exp": "candexp.xls",   "archive_comm_exp": "commexp.xls"},
    2010: {"archive_cand_cont": "candcont.xls", "archive_comm_cont": "commcont.xls",
           "archive_cand_exp": "candexp.xls",   "archive_comm_exp": "commexp.xls"},
    2012: {"archive_cand_cont": "2012_cand_cont.xls", "archive_comm_cont": "2012_comm_cont.xls",
           "archive_cand_exp": "2012_cand_exp.xls",   "archive_comm_exp": "2012_comm_exp.xls"},
    2014: {"archive_cand_cont": "2014_cand_cont.xlsx", "archive_comm_cont": "2014_comm_cont.xlsx",
           "archive_cand_exp": "2014_cand_exp.xlsx",   "archive_comm_exp": "2014_comm_exp.xlsx"},
    2016: {"archive_cand_cont": "2016_cand_cont.xlsx", "archive_comm_cont": "2016_comm_cont.xlsx",
           "archive_cand_exp": "2016_cand_expend.xlsx", "archive_comm_exp": "2016_comm_expend.xlsx"},
    # 2018: as of 2026-06, the live archive page publishes these under a
    # different naming convention than 2014/2016 (no "2018_" prefix), and
    # DOES include committee-level bulk files (previously believed absent).
    2018: {"archive_cand_cont": "candidate_contributions.xlsx",
           "archive_comm_cont": "committee_contributions.xlsx",
           "archive_cand_exp":  "candidate_expenditures.xlsx",
           "archive_comm_exp":  "committee_expenditures.xlsx"},
}

ARCHIVE_RELATIONS = ["archive_cand_cont", "archive_comm_cont", "archive_cand_exp", "archive_comm_exp"]


def _excel_row_count(path: Path) -> int:
    """Number of data rows (excluding header) in a downloaded .xls/.xlsx file."""
    try:
        if path.suffix.lower() == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            ws = wb.active
            n = (ws.max_row or 1) - 1
            wb.close()
            return max(0, n)
        else:
            import xlrd
            wb = xlrd.open_workbook(str(path))
            ws = wb.sheet_by_index(0)
            return max(0, ws.nrows - 1)
    except Exception:
        return 0


def fetch_archive_file(year: int, filename: str, out_path: Path, retries: int = 4) -> int:
    url = f"{ARCHIVE_BASE}/{year}/{filename}"
    last_err = None
    for attempt in range(retries):
        try:
            resp = requests.get(url, timeout=60)
            resp.raise_for_status()
            out_path.write_bytes(resp.content)
            return _excel_row_count(out_path)
        except Exception as e:
            last_err = e
            time.sleep(2 * (attempt + 1))
    raise last_err


# =================== Source 4: legacy 2019 per-filer PDF reports ===================
LEGACY_2019_BASE = "https://archive.sos.idaho.gov/ELECT/Finance/2020"
LEGACY_2019_INDEX_PAGES = [
    "last_year_mid_year_report.html",
    "last_year_statewide_judicial_and_undeclared.html",
    "last_year_state_legislature.html",
    "last_year_party_committees.html",
    "last_year_measure_and_miscellaneous_committees.html",
]
LEGACY_2019_DIR      = RAW_DIR / "legacy_2019_pdfs"
LEGACY_2019_RELATION = "legacy_2019_pdfs"
LEGACY_2019_TIME_BUDGET = 200.0  # seconds per call


def _legacy_2019_links() -> list[str]:
    """Fetch the 5 index pages and return a sorted, deduped list of relative
    PDF paths, e.g. 'First%20Annual/10817.pdf', 'Mid-Year/10856.pdf'."""
    links: set[str] = set()
    for page in LEGACY_2019_INDEX_PAGES:
        resp = requests.get(f"{LEGACY_2019_BASE}/{page}", timeout=60)
        resp.raise_for_status()
        links.update(re.findall(r'href="([^"]+\.pdf)"', resp.text))
    return sorted(links)


def _legacy_2019_outname(rel: str) -> str:
    """'First%20Annual/10817.pdf' -> 'First_Annual_10817.pdf'"""
    return unquote(rel).replace("/", "_").replace(" ", "_")


def fetch_legacy_2019_reports(out_dir: Path, log=None,
                               time_budget: float = LEGACY_2019_TIME_BUDGET) -> dict:
    """Download all per-filer 2019 C-2 PDF reports.

    RESUMABLE: the full link list (~461 files) is cached to
    `<out_dir>/_links.json` on first call (avoids refetching the 5 index
    pages every call). Progress is tracked implicitly — any output file that
    already exists with size > 0 is skipped, so repeated calls just pick up
    where the previous one left off. Each call works for up to `time_budget`
    seconds, then returns.

    Returns {"done": bool, "downloaded": int, "errors": int, "total": int}
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    links_cache = out_dir / "_links.json"
    if links_cache.exists():
        links = json.loads(links_cache.read_text())
    else:
        links = _legacy_2019_links()
        links_cache.write_text(json.dumps(links))

    downloaded = errors = 0
    t0 = time.monotonic()

    for rel in links:
        out_path = out_dir / _legacy_2019_outname(rel)
        if out_path.exists() and out_path.stat().st_size > 0:
            continue
        if time.monotonic() - t0 > time_budget:
            break
        url = f"{LEGACY_2019_BASE}/{rel}"
        try:
            resp = requests.get(url, timeout=60)
            resp.raise_for_status()
            out_path.write_bytes(resp.content)
            downloaded += 1
            if log and downloaded % 25 == 0:
                log.info(f"    legacy_2019_pdfs: {downloaded} downloaded this call")
        except Exception as e:
            errors += 1
            if log:
                log.warning(f"    legacy_2019_pdfs: failed {rel}: {e}")
        time.sleep(0.1)

    still_missing = sum(
        1 for rel in links
        if not (out_dir / _legacy_2019_outname(rel)).exists()
        or (out_dir / _legacy_2019_outname(rel)).stat().st_size == 0
    )

    return {"done": still_missing == 0, "downloaded": downloaded,
            "errors": errors, "total": len(links)}


# Scope groupings
ENTITY_RELATIONS       = ["portal_candidates", "portal_committees"]
PORTAL_TXN_RELATIONS   = ["portal_contributions", "portal_expenditures"]
ELECTIONSTATS_RELATION = "electionstats_activity"

# RE-ENABLED 2026-06-13: the 2026-06-11 "DISABLED" note (static cached blob
# regardless of filters) was tested against sunshine.sos.idaho.gov and does
# NOT apply to canvass.sos.idaho.gov/eng/finances/get_activity.json (see
# ES_URL above and the Source 2 docstring) — plain page-based pagination
# against that host returns correct, distinct pages. electionstats_activity
# is now part of the normal pull and covers the previously-documented
# 2020-2022 donate->PAC/spend gap.
CONTRIBUTION_LIKE = {"portal_contributions", ELECTIONSTATS_RELATION, LEGACY_2019_RELATION,
                      "archive_cand_cont", "archive_comm_cont"}
EXPENDITURE_LIKE  = {"portal_expenditures", ELECTIONSTATS_RELATION, LEGACY_2019_RELATION,
                      "archive_cand_exp", "archive_comm_exp"}

ALL_TXN_RELATIONS = (PORTAL_TXN_RELATIONS + [ELECTIONSTATS_RELATION, LEGACY_2019_RELATION]
                     + ARCHIVE_RELATIONS)


# ============================ orchestrator ============================
def run(
    force: bool = False,
    entities: bool = False,
    transactions: bool = False,
    start_year: int | None = None,
    end_year: int | None = None,
    contributions: bool = False,
    expenditures: bool = False,
    candidates: bool = False,
    committees: bool = False,
):
    """Orchestrate download of Idaho campaign finance data across all 3 sources.

    Vertical scope (mutually exclusive):
        force=True            — re-download everything in scope, wipe relevant manifest entries
        start_year / end_year — restrict year-based downloads (portal txns 2020+,
                                 archive biennium files 2000-2018). Does not affect
                                 the "all"-scoped relations (portal_candidates,
                                 portal_committees, electionstats_activity).

    Horizontal scope:
        No flags               — download everything
        transactions           — all transaction relations (portal TCON/TEXP,
                                  electionstats activity, archive cont/exp files)
        entities / candidates / committees — portal_candidates + portal_committees
                                  (the only source with separate entity registries)
        contributions          — "money in": portal_contributions, electionstats_activity,
                                  archive_cand_cont, archive_comm_cont
        expenditures           — "money out": portal_expenditures, electionstats_activity,
                                  archive_cand_exp, archive_comm_exp
    """
    log = get_logger("idaho", "scrape")
    t0  = time.perf_counter()
    log.info("Starting Idaho scraper")
    log._emit("scrape_started", force=force, entities=entities, transactions=transactions,
              start_year=start_year, end_year=end_year,
              contributions=contributions, expenditures=expenditures,
              candidates=candidates, committees=committees)

    files_ok = files_err = 0
    current_year = datetime.today().year

    try:
        # ── Resolve horizontal scope ───────────────────────────────────
        no_horizontal = not (entities or transactions or contributions or
                              expenditures or candidates or committees)

        if no_horizontal or transactions:
            txn_relations = list(ALL_TXN_RELATIONS)
        else:
            txn_relations = []
            if contributions:
                txn_relations += [r for r in ALL_TXN_RELATIONS if r in CONTRIBUTION_LIKE]
            if expenditures:
                txn_relations += [r for r in ALL_TXN_RELATIONS if r in EXPENDITURE_LIKE]
            # de-dupe while preserving order (electionstats_activity can appear in both)
            seen = set()
            txn_relations = [r for r in txn_relations if not (r in seen or seen.add(r))]

        do_entities = no_horizontal or entities or candidates or committees
        entity_relations = list(ENTITY_RELATIONS) if do_entities else []

        # ── Scoped manifest clearing ───────────────────────────────────
        if force:
            relations_to_clear = set(txn_relations) | set(entity_relations)
            strip_manifest(lambda r: r["relation_type"] not in relations_to_clear)
        elif start_year is not None or end_year is not None:
            # Year-range filtering only applies to relations with numeric years
            # (portal_contributions/expenditures, archive_*). "all"-scoped
            # relations (portal_candidates/committees, electionstats_activity)
            # are untouched.
            ranged_set = set(txn_relations) & (set(PORTAL_TXN_RELATIONS) | set(ARCHIVE_RELATIONS))

            def _outside_range(r: dict) -> bool:
                if r["relation_type"] not in ranged_set:
                    return True
                try:
                    yr = int(r["year"])
                except ValueError:
                    return True
                if start_year is not None and yr < start_year:
                    return True
                if end_year is not None and yr > end_year:
                    return True
                return False

            strip_manifest(_outside_range)

        done = load_manifest()

        # ── Entity registries (Source 1), full pull ────────────────────
        for relation in entity_relations:
            stem = "portal_candidates" if relation == "portal_candidates" else "portal_committees"
            expected_stem = f"{stem}.csv"
            expected_file = RAW_DIR / expected_stem
            key = (relation, "all")

            already_done = key in done or (
                expected_file.exists() and expected_file.stat().st_size > 0
            )
            if already_done and not force:
                log.file_download_skip(filename=expected_stem)
                continue

            log.file_download_start(filename=expected_stem)
            t_file = time.perf_counter()
            try:
                if relation == "portal_candidates":
                    row_count = fetch_portal_candidates(expected_file)
                else:
                    row_count = fetch_portal_committees(expected_file)
            except Exception as e:
                log.file_download_error(filename=expected_stem, error=str(e))
                files_err += 1
                continue

            size = expected_file.stat().st_size
            log.file_download_ok(filename=expected_stem, bytes=size, rows=row_count,
                                  duration_s=time.perf_counter() - t_file)
            files_ok += 1
            upsert_manifest({"relation_type": relation, "year": "all",
                              "filename": expected_stem, "row_count": row_count})
            done.add(key)

        # ── Source 1 transactions (TCON/TEXP), year-split ──────────────
        # current_year only — next year's filingYear 404s (no data yet) and
        # wastes ~20s in retries via _portal_post before falling through.
        max_portal_year = current_year
        for relation in [r for r in txn_relations if r in PORTAL_TXN_RELATIONS]:
            txn_type = "TCON" if relation == "portal_contributions" else "TEXP"
            stem     = relation  # "portal_contributions" / "portal_expenditures"

            for year in range(EARLIEST_PORTAL_YEAR, max_portal_year + 1):
                if start_year is not None and year < start_year:
                    continue
                if end_year is not None and year > end_year:
                    continue

                key = (relation, str(year))
                expected_stem = f"{stem}_{year}.csv"
                expected_file = RAW_DIR / expected_stem

                year_range_active = start_year is not None or end_year is not None
                already_done = key in done or (
                    not year_range_active
                    and expected_file.exists()
                    and expected_file.stat().st_size > 0
                )
                if already_done and year != current_year and not force:
                    log.file_download_skip(filename=expected_stem)
                    continue

                log.file_download_start(filename=expected_stem)
                t_file = time.perf_counter()
                try:
                    row_count = fetch_portal_transactions(txn_type, year, expected_file)
                except Exception as e:
                    log.file_download_error(filename=expected_stem, error=str(e))
                    files_err += 1
                    continue

                size = expected_file.stat().st_size
                log.file_download_ok(filename=expected_stem, bytes=size, rows=row_count,
                                      duration_s=time.perf_counter() - t_file)
                files_ok += 1
                upsert_manifest({"relation_type": relation, "year": str(year),
                                  "filename": expected_stem, "row_count": row_count})
                done.add(key)
                time.sleep(0.3)

        # ── Source 2: electionstats activity feed, full pull ───────────
        if ELECTIONSTATS_RELATION in txn_relations:
            expected_stem = "electionstats_activity.csv"
            expected_file = RAW_DIR / expected_stem
            progress_file = RAW_DIR / (expected_stem + ".progress")
            key = (ELECTIONSTATS_RELATION, "all")

            already_done = key in done or (
                expected_file.exists() and expected_file.stat().st_size > 0
                and not progress_file.exists()
            )
            if already_done and not force:
                log.file_download_skip(filename=expected_stem)
            else:
                if force:
                    expected_file.unlink(missing_ok=True)
                    progress_file.unlink(missing_ok=True)

                log.file_download_start(filename=expected_stem)
                t_file = time.perf_counter()
                try:
                    result = fetch_electionstats_activity(expected_file, log=log)
                except Exception as e:
                    log.file_download_error(filename=expected_stem, error=str(e))
                    files_err += 1
                    result = None

                if result is not None:
                    if result["done"]:
                        with open(expected_file, newline="", encoding="utf-8") as f:
                            row_count = sum(1 for _ in f) - 1  # exclude header
                        size = expected_file.stat().st_size
                        log.file_download_ok(filename=expected_stem, bytes=size, rows=row_count,
                                              duration_s=time.perf_counter() - t_file)
                        files_ok += 1
                        upsert_manifest({"relation_type": ELECTIONSTATS_RELATION, "year": "all",
                                          "filename": expected_stem, "row_count": row_count})
                        done.add(key)
                    else:
                        log.info(f"    electionstats_activity: +{result['rows_written']:,} rows "
                                 f"this call, resume at page {result['next_page']} on next run")

        # ── Source 3: static archive files, biennium years ─────────────
        for relation in [r for r in txn_relations if r in ARCHIVE_RELATIONS]:
            for year, file_map in ARCHIVE_FILES.items():
                if relation not in file_map:
                    continue  # e.g. 2018 has no committee-level files
                if start_year is not None and year < start_year:
                    continue
                if end_year is not None and year > end_year:
                    continue

                filename = file_map[relation]
                ext = Path(filename).suffix
                suffix = relation.replace("archive_", "")  # cand_cont / comm_cont / cand_exp / comm_exp
                expected_stem = f"archive_{year}_{suffix}{ext}"
                expected_file = RAW_DIR / expected_stem
                key = (relation, str(year))

                already_done = key in done or (
                    expected_file.exists() and expected_file.stat().st_size > 0
                )
                if already_done and not force:
                    log.file_download_skip(filename=expected_stem)
                    continue

                log.file_download_start(filename=expected_stem)
                t_file = time.perf_counter()
                try:
                    row_count = fetch_archive_file(year, filename, expected_file)
                except Exception as e:
                    log.file_download_error(filename=expected_stem, error=str(e))
                    files_err += 1
                    continue

                size = expected_file.stat().st_size
                log.file_download_ok(filename=expected_stem, bytes=size, rows=row_count,
                                      duration_s=time.perf_counter() - t_file)
                files_ok += 1
                upsert_manifest({"relation_type": relation, "year": str(year),
                                  "filename": expected_stem, "row_count": row_count})
                done.add(key)
                time.sleep(0.2)

        # ── Source 4: legacy 2019 per-filer PDF reports, full pull ──────
        if LEGACY_2019_RELATION in txn_relations:
            key = (LEGACY_2019_RELATION, "2019")
            already_done = key in done

            if already_done and not force:
                log.file_download_skip(filename="legacy_2019_pdfs/*")
            else:
                if force:
                    strip_manifest(lambda r: not (r["relation_type"] == LEGACY_2019_RELATION
                                                    and r["year"] == "2019"))
                    done.discard(key)

                log.file_download_start(filename="legacy_2019_pdfs/*")
                t_file = time.perf_counter()
                try:
                    result = fetch_legacy_2019_reports(LEGACY_2019_DIR, log=log)
                except Exception as e:
                    log.file_download_error(filename="legacy_2019_pdfs/*", error=str(e))
                    files_err += 1
                    result = None

                if result is not None:
                    if result["done"]:
                        log.file_download_ok(filename="legacy_2019_pdfs/*", bytes=0,
                                              rows=result["total"],
                                              duration_s=time.perf_counter() - t_file)
                        files_ok += 1
                        upsert_manifest({"relation_type": LEGACY_2019_RELATION, "year": "2019",
                                          "filename": "legacy_2019_pdfs/", "row_count": result["total"]})
                        done.add(key)
                    else:
                        log.info(f"    legacy_2019_pdfs: +{result['downloaded']} downloaded this call "
                                 f"({result['errors']} errors), {result['total']} total — resume on next run")

        duration = round(time.perf_counter() - t0, 1)
        log.info(f"Done in {duration}s")
        log._emit("scrape_completed", status="completed", duration_s=duration,
                  files_ok=files_ok, files_err=files_err)

    except KeyboardInterrupt:
        log.warning("Interrupted")
        log._emit("scrape_completed", status="interrupted",
                  duration_s=round(time.perf_counter() - t0, 1),
                  files_ok=files_ok, files_err=files_err)
        raise

    except Exception as e:
        log._emit("scrape_completed", status="error",
                  duration_s=round(time.perf_counter() - t0, 1),
                  files_ok=files_ok, files_err=files_err,
                  error_type=type(e).__name__, error=str(e))
        raise


# ====== CLI ==================================
if __name__ == "__main__":
    # Vertical scope (mutually exclusive):
    #   (no flag)                    incremental — fill manifest gaps, refresh current year
    #   --start-year / --end-year    year range only (portal TCON/TEXP, archive files)
    #   --force                      re-download everything in scope, wipe manifest entries
    #
    # Horizontal scope:
    #   (no flag)         all sources
    #   --transactions    all transaction relations (Source 1 TCON/TEXP, Source 2 activity,
    #                      Source 3 cont/exp archives)
    #   --entities        portal_candidates + portal_committees (Source 1 only)
    #   --contributions   "money in": portal_contributions, electionstats_activity,
    #                      archive_cand_cont, archive_comm_cont
    #   --expenditures    "money out": portal_expenditures, electionstats_activity,
    #                      archive_cand_exp, archive_comm_exp
    #   --candidates      portal_candidates + portal_committees
    #   --committees      portal_candidates + portal_committees
    import argparse
    ap = argparse.ArgumentParser(
        description="Download Idaho campaign finance data (3 sources: current portal API, "
                     "electionstats activity feed, 2000-2018 static archive)."
    )

    vert = ap.add_mutually_exclusive_group()
    vert.add_argument("--force",      action="store_true",
                      help="re-download everything in scope, wipe relevant manifest entries")
    vert.add_argument("--start-year", type=int, metavar="YYYY",
                      help="earliest year to download (inclusive; portal TCON/TEXP + archive files)")
    ap.add_argument("--end-year", type=int, metavar="YYYY",
                    help="latest year to download (inclusive); use with or without --start-year")

    ap.add_argument("--transactions", action="store_true",
                    help="all transaction relations across all 3 sources")
    ap.add_argument("--entities",     action="store_true",
                    help="portal_candidates + portal_committees only")

    ap.add_argument("--contributions", action="store_true",
                    help="'money in' relations only (portal TCON, electionstats activity, archive cont)")
    ap.add_argument("--expenditures",  action="store_true",
                    help="'money out' relations only (portal TEXP, electionstats activity, archive exp)")
    ap.add_argument("--candidates",    action="store_true",
                    help="portal_candidates + portal_committees")
    ap.add_argument("--committees",    action="store_true",
                    help="portal_candidates + portal_committees")

    args, _ = ap.parse_known_args()

    cy = datetime.today().year
    if args.end_year:
        if args.end_year > cy + 2:
            ap.error(f"--end-year cannot exceed {cy + 2}")
        if args.start_year and args.start_year > args.end_year:
            ap.error("--start-year cannot be greater than --end-year")
    if args.force and args.end_year:
        ap.error("--force cannot be combined with --end-year")

    try:
        run(
            force=args.force,
            entities=args.entities,
            transactions=args.transactions,
            start_year=args.start_year,
            end_year=args.end_year,
            contributions=args.contributions,
            expenditures=args.expenditures,
            candidates=args.candidates,
            committees=args.committees,
        )
    except KeyboardInterrupt:
        sys.exit(130)
    except Exception:
        sys.exit(1)
