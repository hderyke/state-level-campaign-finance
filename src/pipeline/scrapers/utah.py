"""
scrapers/utah.py — Download Utah campaign finance data from the Lieutenant
Governor's Financial Disclosures site, https://disclosures.utah.gov/.

## The three endpoints this uses

Everything Utah publishes in bulk hangs off one ASP.NET MVC controller,
`/Search/AdvancedSearch`. No login, no ViewState, no form tokens.

  1. **Bulk transaction export** (the whole point of this scraper)

         GET /Search/AdvancedSearch/GenerateReport?ReportYear=<YYYY>&EntityType=<CODE>

     Returns one CSV containing *every* itemized contribution and
     expenditure filed by *every* entity of that type for that report
     year. Eight entity types x 1998-present = ~230 combinations, of
     which roughly 100 actually contain data (~160 MB in total). This is
     the same request path the Investigative Reporting Workshop's
     Accountability Project uses for its Utah dataset, which is where
     the `EntityType`-without-an-id form was confirmed — the site's own
     UI only ever links the per-entity form (2. below), so this is not
     discoverable from the HTML alone.

     A combination with no data returns an **HTML page** rather than a
     CSV, with a 200 status. The only reliable discriminator is the
     response `Content-Type` (`application/csv` when it's real data) —
     see `_is_csv_response()`. Do not trust the status code.

  2. **Per-entity export** — linked from the search results table as
     "Download Data by Year":

         GET /Search/AdvancedSearch/GenerateReport/<entity_id>?ReportYear=<YYYY>

     Same columns as 1. minus the leading `FILED` column, scoped to a
     single filer. Not used for bulk collection (1. supersedes it), but
     it is the right tool for spot-checking one committee — e.g.
     `GenerateReport/484?ReportYear=2014`. There is also a
     per-*filing* variant, `/Search/PublicSearch/CSVDownload/<filing_id>?year=<YYYY>`,
     reachable from a folder's detail page; it is strictly narrower
     again and unused here.

  3. **Entity roster** — the search results grid, fetched as a partial:

         POST /Search/AdvancedSearch/GetEntityReportList
         Search=&EntityType=<CODE>&ReportYear=<YYYY>
         &HideContributions=false&HideExpenditures=false&PageNumber=<N>
         &X-Requested-With=XMLHttpRequest

     Returns an HTML table fragment, **25 rows per page**, one row per
     filer active in that report year: entity name (linked to
     `/Search/PublicSearch/FolderDetails/<entity_id>`), entity type
     label, per-report filed/not-filed marks, ending balance, and a
     "Download Data by Year" cell listing every year that entity has
     data for. `X-Requested-With` is sent BOTH as a header and as a form
     field, because the site's own JS does exactly that.

     This roster is the only place `entity_id` appears — the bulk CSVs
     from 1. carry filer *names* and nothing else. Without it there is
     no `state_filer_id`, and no way to use the exact-ID party join
     described under `--party` below. So it is swept, not skipped.

## Why curl_cffi and not requests

disclosures.utah.gov sits behind F5 BIG-IP ASM: responses set `TSPD_101`,
`TS<hex>` and `TSPD_101_DID` cookies, which is that product's
fingerprint-and-challenge cookie set. Plain `requests` presents a Python
TLS ClientHello no amount of header spoofing can disguise, so this uses
`curl_cffi`'s Chrome impersonation — the same approach already taken by
scrapers/ohio.py and scrapers/new_hampshire.py for the same class of
problem. This is NOT a browser: no JS is executed, and Playwright is
deliberately not used.

If a run starts coming back with HTML challenge pages on every request,
bump `IMPERSONATE` to a newer Chrome build before reaching for anything
heavier.

## Raw CSV format (see parsers/utah.py for the full mapping)

Header of the bulk (EntityType) export:

    FILED,<TYPE>,REPORT,TRAN_ID,TRAN_TYPE,TRAN_DATE,TRAN_AMT,INKIND,
    LOAN,AMENDS,NAME,PURPOSE,ADDRESS1,ADDRESS2,CITY,STATE,ZIP,
    INKIND_COMMENTS

Column 2's *name* changes with the entity type (`PAC`, `PIC`,
`CORPORATION`, ...) but always holds the filing entity's name. The
per-entity export (2.) omits `FILED`, so column 1 is the entity name
instead. The parser resolves both shapes positionally off `REPORT`
rather than hardcoding either.

Files are written to disk byte-for-byte as received. Utah's export does
not escape double quotes that appear *inside* a quoted field, so the
files are not strictly valid CSV; that is repaired at parse time (see
`_repair_line` in parsers/utah.py), not here, so raw stays raw.

## What the roster is really for

Beyond `entity_id`, the roster is the only place Utah states a
candidate's office, district or election cycle: PCC folders are named
"Last, First (YYYY Office-District)" — "Abbott, Nelson (2022 House-57)",
"Aalders, Tim (2012 Lieutenant Governor)". parsers/utah.py reads all
three off that annotation, and it is also what makes a folder
per-*candidacy* rather than per-person. The itemized export contains none
of it.

## --party: party backfill

Utah's disclosure data contains no party at all. Two external sources
fill that in, fetched by a separate `--party` mode (see `run_party()`):

  * **Open States** nightly bulk CSV,
    https://data.openstates.org/people/current/ut.csv — CC0, no API key.
    Its `links`/`sources` columns contain each legislator's own
    `disclosures.utah.gov/Search/PublicSearch/FolderDetails/<id>` URL,
    which makes an **exact entity-id join** possible rather than a name
    guess. Covers only currently-serving legislators (83 rows when the
    live file was last read; Utah seats 104, so not even every sitting
    member is present).
    Open States' retired legislators exist only as YAML in the
    openstates/people GitHub repo, not in the nightly CSV export;
    pulling those would mean walking a third-party repo tree, and is
    deliberately not done here.

  * **Utah historical election canvasses**,
    https://vote.utah.gov/historical-election-results/ — the Excel
    canvasses (2000, 2008-2020) give candidate + party + race for
    statewide, legislative and federal seats. Links are discovered from
    that page rather than hardcoded.

    !! UNVERIFIED LAYOUT !! The canvass workbooks' internal sheet
    layout could not be inspected while this was written (no egress to
    vote.utah.gov from the authoring environment). `_extract_canvass()`
    therefore scans cells for a recognizable Utah party label and takes
    the adjacent cell as the candidate name, tracking the nearest
    single-cell row above as the race title. It is written to degrade to
    zero rows and warn rather than emit garbage. **On the first real
    run, check the `canvass_rows` count per file in the log and eyeball
    `UT_ElectionResults.csv` before trusting it.** See
    docs/states/utah.md for exactly what to look at.

Usage:

    python3 src/pipeline/scrapers/utah.py                     # bulk + roster
    python3 src/pipeline/scrapers/utah.py --party             # both party sources
    python3 src/pipeline/scrapers/utah.py --party --openstates
    python3 src/pipeline/scrapers/utah.py --party --canvass

`--party` is not part of the orchestrator's flag set and is never
forwarded by main.py — it is a manual, out-of-band refresh, the same
arrangement scrapers/texas.py and scrapers/new_york.py use.
"""

import csv
import hashlib
import re
import sys
import time
from pathlib import Path
from urllib.parse import urljoin

try:
    # disclosures.utah.gov is fronted by F5 BIG-IP ASM (the TSPD_101 /
    # TS<hex> cookie set is that product's signature). ASM fingerprints the
    # TLS/HTTP2 handshake, not just headers, so a plain `requests` session
    # is identifiable as non-browser regardless of User-Agent. curl_cffi
    # wraps a patched libcurl that reproduces a real Chrome ClientHello
    # under the same requests-style API — not a browser, no JS, no
    # Playwright. Same reasoning as scrapers/ohio.py.
    from curl_cffi import requests
    from curl_cffi.requests.exceptions import SSLError as _CurlSSLError
except ImportError as e:                                   # pragma: no cover
    raise ImportError(
        "scrapers/utah.py requires curl_cffi (pip install curl_cffi). "
        "disclosures.utah.gov is behind F5 BIG-IP ASM, which fingerprints "
        "the TLS handshake — plain `requests` gets challenge pages instead "
        "of data. See the top of this file."
    ) from e
from bs4 import BeautifulSoup

PROJECT_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(PROJECT_ROOT))
import config
from src.reporting.logger import get_logger

# =============================== paths ================================
RAW_DIR     = PROJECT_ROOT / "data" / "Utah" / "raw"
CANVASS_DIR = RAW_DIR / "canvass"
MANIFEST    = PROJECT_ROOT / "data" / "Utah" / "manifest.csv"

RAW_DIR.mkdir(parents=True, exist_ok=True)

MANIFEST_COLS = ["relation_type", "entity_type", "year", "filename",
                 "bytes", "rows", "downloaded_at"]

# ========================= state-specific constants ===================

BASE     = "https://disclosures.utah.gov"
ADV      = f"{BASE}/Search/AdvancedSearch"
GENERATE = f"{ADV}/GenerateReport"
ROSTER   = f"{ADV}/GetEntityReportList"

# Earliest year offered by the site's own ReportYear dropdown. Confirmed
# against a saved copy of /Search/AdvancedSearch: options run 1998-2026.
# Itemized data is sparse before ~2008 (the LG's office only required
# electronic filing later), but the empty years cost one cheap request
# each and are skipped automatically when they come back as HTML.
EARLIEST_YEAR = 1998

# (code, display label as it appears in the roster's Type column, filename slug)
# Codes come from the EntityType <select> on /Search/AdvancedSearch. The two
# transaction-scoped options in that dropdown ("Contributions",
# "Expenditures") are search filters, not entity types, and are excluded —
# GenerateReport with those returns nothing useful.
ENTITY_TYPES = [
    ("PCC",    "Candidates & Office Holders", "pcc"),
    ("PAC",    "Political Action Committee",  "pac"),
    ("PIC",    "Political Issues Committee",  "pic"),
    ("PARTY",  "Political Party",             "party"),
    ("CORP",   "Corporation",                 "corp"),
    ("LABOR",  "Labor Organizations",         "labor"),
    ("ELECT",  "Electioneering",              "elect"),
    ("INDEXP", "Independent Expenditures",    "indexp"),
]

# The roster grid is server-paginated at a fixed 25 rows — counted off a
# saved results page. A short page therefore means "last page".
ROSTER_PAGE_SIZE = 25
# Hard stop so a site change that makes every page look full can't spin
# forever. 400 pages = 10,000 filers in a single type/year, an order of
# magnitude above anything Utah has.
ROSTER_MAX_PAGES = 400

REQUEST_PAUSE = 0.25          # polite delay between roster page requests

# Bulk CSVs are multi-megabyte and server-GENERATED per request (GenerateReport
# runs a query and builds the file), so these are far heavier than a static
# download. 0.5s between them proved too aggressive on a real run: fifteen
# files (~25 MB) went through cleanly and then every subsequent request came
# back curl (56) "Recv failure: Connection was reset" — the server accepting
# the request, starting to respond, then killing the connection. Overridable
# with --pause; see _RESET_HELP for what to try if it recurs.
DOWNLOAD_PAUSE = 1.5

# A reset is transient in principle, so retry rather than writing the file off.
# Each retry rebuilds the session (see the loop in run()), which is the part
# that actually matters if the cause is a poisoned connection or an expired
# ASM cookie rather than rate limiting.
DOWNLOAD_RETRIES = 4
DOWNLOAD_BACKOFF = 15.0       # seconds, multiplied by the attempt number

# Consecutive files that exhaust their retries before the run gives up. One
# reset is noise; this many in a row is the source refusing to talk to us, and
# grinding through the remaining ~200 combinations just deepens whatever
# threshold was tripped.
MAX_CONSECUTIVE_ERRORS = 4

# Keep in step with config.USER_AGENT's Chrome major version if that moves.
IMPERSONATE = "chrome136"

SESSION_HEADERS = {
    # No User-Agent: curl_cffi's impersonate= sets one that matches the
    # TLS fingerprint it replicates. Overriding it would desync the header
    # from the handshake, which is itself a bot-detection signal.
    "Accept": "*/*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": ADV,
    "Origin": BASE,
}

_FOLDER_ID_RE   = re.compile(r"/FolderDetails/(\d+)")
_REPORT_YEAR_RE = re.compile(r"/GenerateReport/(\d+)\?ReportYear=(\d{4})")

# Roster rows for closed committees carry a marker in the name itself, e.g.
# "2015 Transportation Alliance [CLOSED]". Stripped from committee_name so
# it joins cleanly against the transaction files (which predate the
# closure and carry the bare name), and recorded as active=0 instead.
_CLOSED_RE = re.compile(r"\s*\[\s*CLOSED\s*\]\s*$", re.I)


# ========================= manifest helpers ===========================

def load_manifest() -> dict[str, dict]:
    """filename -> manifest row."""
    if not MANIFEST.exists():
        return {}
    with open(MANIFEST, newline="", encoding="utf-8") as f:
        return {row["filename"]: row for row in csv.DictReader(f)}


def strip_manifest(keep_fn) -> None:
    """Rewrite the manifest keeping only rows for which keep_fn(row) is True."""
    if not MANIFEST.exists():
        return
    with open(MANIFEST, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    with open(MANIFEST, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=MANIFEST_COLS,
                           extrasaction="ignore", restval="")
        w.writeheader()
        w.writerows(r for r in rows if keep_fn(r))


def upsert_manifest(record: dict, done: dict) -> None:
    rows = []
    if MANIFEST.exists():
        with open(MANIFEST, newline="", encoding="utf-8") as f:
            rows = [r for r in csv.DictReader(f)
                    if r.get("filename") != record["filename"]]
    rows.append(record)
    with open(MANIFEST, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=MANIFEST_COLS,
                           extrasaction="ignore", restval="")
        w.writeheader()
        w.writerows(rows)
    done[record["filename"]] = record


# ============================ http helpers ============================

# Guidance printed when TLS verification fails. A certificate error is a
# property of this machine and its network, not of the file being fetched, so
# it is worth saying once and loudly rather than 232 times.
_TLS_HELP = """TLS verification failed against disclosures.utah.gov (curl error 60).

This is almost always a TLS-inspecting proxy (Zscaler, Netskope, a corporate
appliance) re-signing HTTPS with an internal root CA. The tell is that the
same URL loads fine in your browser: the browser trusts that root from the
OS store, OpenSSL/libcurl does not.

This scraper already asks config.ca_bundle() for the OS trust store, so if
you are seeing this, either the cached bundle is stale or your root is not in
the store it exports. Try, in order:

  1. Rebuild the bundle (config caches it in your temp dir):
       python3 -c "import config; print(config.ca_bundle(refresh=True))"

  2. Ask config which verify= setting actually works here:
       python3 -c "import config; config.diagnose_tls('https://disclosures.utah.gov/')"

  3. If your organisation ships its own PEM, point at it explicitly --
     config.ca_bundle() honours these ahead of everything else:
       set REQUESTS_CA_BUNDLE=C:\\path\\to\\corp-roots.pem

This is a *verification* failure, not the WAF. If F5 ASM were blocking the
client you would get an HTML challenge page, not a curl (60) -- so bumping
IMPERSONATE will not help."""


# Printed when the source stops answering mid-transfer, repeatedly.
_RESET_HELP = """disclosures.utah.gov stopped completing responses.

{n} files in a row failed after {r} attempts each, the last with:
  {err}

The successful downloads are already recorded in data/Utah/manifest.csv, so
**re-running picks up exactly where this left off** — nothing is re-fetched
and nothing is lost. That is usually the whole fix: wait a while and run the
same command again.

curl (56) means the server accepted the request, began responding, then cut
the connection. It is not TLS and not the WAF's challenge page. On a real run
this appeared only after ~15 consecutive multi-megabyte downloads, which
points at a rate or volume threshold rather than anything about the specific
year that failed.

If it recurs immediately on a fresh run, slow the scraper down and narrow it:

  1. Longer gap between downloads (default {pause}s):
       python3 src/pipeline/scrapers/utah.py --pause 5

  2. Take it a few years at a time, letting the manifest accumulate:
       python3 src/pipeline/scrapers/utah.py --start-year 2015 --end-year 2018

  3. Fetch the entity roster separately from the bulk files, so a stall in one
     doesn't cost the other:
       python3 src/pipeline/scrapers/utah.py --entities
       python3 src/pipeline/scrapers/utah.py --transactions --pause 5

If it fails at the *same* file every time regardless of pacing, that file is
the problem, not the pacing — check whether that (type, year) is unusually
large or the server simply times out generating it. GenerateReport builds each
CSV on demand, so a big election year can take a while server-side."""


# curl exits that mean "the transfer broke", as opposed to "the server said
# no". These get retried; an HTTP error status does not.
def _is_transient(exc: Exception) -> bool:
    """True for a mid-transfer network failure worth retrying."""
    text = str(exc).lower()
    return any(sig in text for sig in (
        "connection was reset",     # curl 56, the one seen in practice
        "connection reset",
        "recv failure",
        "send failure",             # curl 55
        "empty reply",              # curl 52
        "transfer closed",          # curl 18
        "operation timed out",      # curl 28
        "timed out",
        "connection refused",
    ))


def _is_tls_failure(exc: Exception) -> bool:
    """True for a certificate-verification failure (curl error 60 and kin)."""
    if isinstance(exc, _CurlSSLError):
        return True
    # Defensive: curl_cffi has moved this class between modules across
    # releases, so fall back to the message rather than letting a real cert
    # problem be reported as an ordinary per-file download error. Matched as
    # a list of independent phrases, not "certificate AND verify" — OpenSSL
    # words this several ways and only some of them say "verify" ("unable to
    # get local issuer certificate" on its own is the common one).
    text = str(exc).lower()
    return any(sig in text for sig in (
        "certificate verify failed",
        "unable to get local issuer",
        "self signed certificate",
        "self-signed certificate",
        "certificate has expired",
        "unable to verify",
        "ssl certificate problem",
        "verify result",              # curl's "OpenSSL verify result: ..."
    ))


# disclosures.utah.gov closes a keep-alive connection after roughly 33
# requests. Measured, not guessed: on a real roster sweep the curl (56) resets
# landed at pages 4, 37 and 69 of one pass and 5, 38 of the next — a 32-33
# request stride, which is an IIS/F5 MaxKeepAliveRequests limit, not rate
# limiting. libcurl surfaces that close as an error instead of reconnecting
# transparently, so reconnect *before* the server forces it rather than
# treating a predictable server behaviour as a failure to retry through.
SESSION_MAX_REQUESTS = 25


def _new_session():
    # verify=config.ca_bundle() is the whole reason config is imported here.
    # On Windows it exports the OS trust store — the same one the browser
    # uses — to a PEM and returns its path; on macOS/Linux it returns True,
    # i.e. ordinary certifi verification. Without it, any machine behind a
    # TLS-inspecting proxy fails every request with curl (60): the browser
    # works and libcurl doesn't, which is the signature of an internal CA
    # sitting in the OS store but not in certifi's bundle.
    #
    # Never verify=False. config.ca_bundle() is documented not to return it:
    # on a network already known to be intercepting traffic, accepting any
    # certificate at all is the wrong trade.
    s = requests.Session(impersonate=IMPERSONATE, verify=config.ca_bundle())
    s.headers.update(SESSION_HEADERS)
    # Prime the ASM cookie jar with a GET of the search page itself before
    # POSTing to its partial-view endpoint. The site sets TSPD_101 on first
    # contact and echoes it back on subsequent requests; skipping this makes
    # the first POST look like it arrived out of nowhere.
    try:
        s.get(ADV, timeout=60)
    except Exception as e:
        # A failed warm-up is normally not fatal — the real requests will say
        # so. A TLS failure is the exception: it will fail every subsequent
        # request identically, and this is the earliest point it can be
        # caught, before 232 identical error lines scroll past.
        if _is_tls_failure(e):
            raise RuntimeError(_TLS_HELP) from e
    return s


class RecyclingSession:
    """A curl_cffi session that reconnects before the server hangs up.

    Proxies `.get`/`.post` and rebuilds the underlying session every
    SESSION_MAX_REQUESTS calls, so the keep-alive limit described above is
    never reached. `recycle()` forces one immediately, which is what a
    retry after a reset wants: libcurl pools connections, so a reset one
    keeps failing every subsequent request made through it.

    Wrapping rather than returning a new session from every helper keeps
    the call sites honest — there is no way to accidentally go on using a
    stale handle, which an explicit "remember to reassign" contract invites.
    """

    def __init__(self, max_requests: int = SESSION_MAX_REQUESTS):
        self._max = max(1, max_requests)
        self._s = _new_session()
        self._n = 1                     # the warm-up GET counts
        self.recycles = 0

    def recycle(self) -> None:
        self._s = _new_session()
        self._n = 1
        self.recycles += 1

    def _tick(self):
        if self._n >= self._max:
            self.recycle()
        self._n += 1
        return self._s

    def get(self, *a, **kw):
        return self._tick().get(*a, **kw)

    def post(self, *a, **kw):
        return self._tick().post(*a, **kw)


def _is_csv_response(resp) -> bool:
    """True when GenerateReport actually returned data rather than an HTML page.

    A (type, year) combination with no filings returns a normal HTML page
    with status 200, so the status code says nothing. Utah sets
    `Content-Type: application/csv` on the real thing; fall back to
    sniffing the body for an HTML doctype/tag if the header is missing.
    """
    ctype = (resp.headers.get("Content-Type") or "").lower()
    if "csv" in ctype:
        return True
    if "html" in ctype:
        return False
    head = resp.content[:512].lstrip().lower()
    return not (head.startswith(b"<!doctype") or head.startswith(b"<html"))


def _count_rows(blob: bytes) -> int:
    """Data rows in a CSV blob — newline count less the header, floored at 0.

    Approximate by design: Utah embeds raw newlines in a handful of
    free-text fields, so this can overcount slightly. It is a log/manifest
    figure only; the parser counts for real.
    """
    n = blob.count(b"\n")
    if blob and not blob.endswith(b"\n"):
        n += 1
    return max(n - 1, 0)


# ====================== bulk transaction downloads =====================

def _download_bulk(session, entity_code: str, year: int, dest: Path) -> int | None:
    """GET the (entity type, year) bulk CSV. Returns row count, or None if
    the site had no data for that combination (HTML response)."""
    resp = session.get(GENERATE,
                       params={"ReportYear": str(year), "EntityType": entity_code},
                       timeout=300)
    resp.raise_for_status()
    if not _is_csv_response(resp):
        return None
    # Written verbatim — no re-encoding, no CSV rewriting. The quote-escaping
    # defect in Utah's export is repaired at parse time so raw stays raw.
    tmp = dest.with_suffix(dest.suffix + ".part")
    tmp.write_bytes(resp.content)
    tmp.replace(dest)
    return _count_rows(resp.content)


# =========================== entity roster ============================

ENTITY_COLS = ["entity_id", "entity_name", "entity_type", "entity_type_label",
               "active", "ending_balance", "data_years"]

ENTITIES_FILE = "entities.csv"


def _parse_roster_page(html: str) -> list[dict]:
    """Parse one GetEntityReportList fragment into roster rows.

    Each row: entity_id, entity_name (closure marker stripped), active,
    entity_type_label, ending_balance, data_years (years the entity has
    downloadable data for).
    """
    soup = BeautifulSoup(html, "html.parser")
    tables = soup.find_all("table")
    if not tables:
        return []
    # The fragment normally contains exactly one table; a full-page response
    # (e.g. a challenge or error page) may contain several, so take the
    # largest by row count rather than assuming position.
    table = max(tables, key=lambda t: len(t.find_all("tr")))
    body = table.find("tbody") or table

    out = []
    for tr in body.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 3:
            continue                        # header/spacer row
        link = tds[0].find("a", href=_FOLDER_ID_RE)
        if not link:
            continue
        m = _FOLDER_ID_RE.search(link.get("href") or "")
        if not m:
            continue
        raw_name = link.get_text(strip=True)
        closed   = bool(_CLOSED_RE.search(raw_name))
        # Years this entity has a downloadable export for. Present on the
        # last cell as one <a> per year.
        years = sorted({
            ym.group(2)
            for a in tds[-1].find_all("a", href=True)
            for ym in [_REPORT_YEAR_RE.search(a["href"])]
            if ym
        })
        out.append({
            "entity_id":         m.group(1),
            "entity_name":       _CLOSED_RE.sub("", raw_name).strip(),
            "entity_type_label": tds[1].get_text(strip=True),
            "active":            "0" if closed else "1",
            "ending_balance":    tds[-2].get_text(strip=True),
            "data_years":        "|".join(years),
        })
    return out


def _fetch_roster_page(session, entity_code: str, year: int, page: int) -> list[dict]:
    body = {
        "Search": "",
        "EntityType": entity_code,
        "ReportYear": str(year),
        "HideContributions": "false",
        "HideExpenditures": "false",
        "PageNumber": str(page),
        # Sent as a form field as well as a header — the site's own JS does
        # both, and the controller has been seen to read the form value.
        "X-Requested-With": "XMLHttpRequest",
    }
    resp = session.post(
        ROSTER, data=body, timeout=120,
        headers={"X-Requested-With": "XMLHttpRequest",
                 "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8"},
    )
    resp.raise_for_status()
    return _parse_roster_page(resp.text)


def _load_entities() -> dict[str, dict]:
    path = RAW_DIR / ENTITIES_FILE
    if not path.exists() or path.stat().st_size == 0:
        return {}
    with open(path, newline="", encoding="utf-8") as f:
        return {r["entity_id"]: r for r in csv.DictReader(f) if r.get("entity_id")}


def _write_entities(entities: dict[str, dict]) -> int:
    path = RAW_DIR / ENTITIES_FILE
    tmp  = path.with_suffix(".csv.part")
    rows = sorted(entities.values(),
                  key=lambda r: (r.get("entity_type", ""), r.get("entity_name", "")))
    with open(tmp, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=ENTITY_COLS,
                           extrasaction="ignore", restval="")
        w.writeheader()
        w.writerows(rows)
    tmp.replace(path)
    return len(rows)


def _merge_entity(entities: dict[str, dict], row: dict,
                  entity_code: str) -> None:
    """Fold one roster row into the accumulated roster, unioning year sets.

    Still a merge rather than a plain insert even though each type is now
    swept once: an incremental run reads the existing entities.csv first, so
    a row already on disk contributes its `data_years` alongside this
    sighting's. The current sighting wins on the mutable fields (name,
    balance, closure) — it is the newer observation.
    """
    eid = row["entity_id"]
    prev = entities.get(eid)
    data_years = set(filter(None, (row.get("data_years") or "").split("|")))
    if prev:
        data_years |= set(filter(None, (prev.get("data_years") or "").split("|")))
    entities[eid] = {
        "entity_id":         eid,
        "entity_name":       row["entity_name"],
        "entity_type":       entity_code,
        "entity_type_label": row["entity_type_label"],
        "active":            row["active"],
        "ending_balance":    row["ending_balance"],
        "data_years":        "|".join(sorted(data_years)),
    }


def _roster_page_with_retry(session, log, entity_code: str, year: int,
                            page: int) -> list[dict]:
    """One roster page, retrying transient failures.

    The session recycles itself (see RecyclingSession), so a reset just
    needs a retry — there is no stale handle for the caller to reassign.
    """
    for attempt in range(1, DOWNLOAD_RETRIES + 1):
        try:
            return _fetch_roster_page(session, entity_code, year, page)
        except Exception as e:
            if _is_tls_failure(e):
                raise RuntimeError(_TLS_HELP) from e
            if not _is_transient(e) or attempt == DOWNLOAD_RETRIES:
                raise
            session.recycle()
            # A single reset here is routine, not alarming: it is the
            # keep-alive limit, and SESSION_MAX_REQUESTS exists to stay
            # under it. Only escalate to a warning once a retry has already
            # failed, which is when it stops being ordinary.
            wait = DOWNLOAD_BACKOFF * attempt
            msg = (f"    roster {entity_code} p{page}: {type(e).__name__} on "
                   f"attempt {attempt}/{DOWNLOAD_RETRIES} — reconnected, "
                   f"retrying in {wait:.0f}s")
            (log.debug if attempt == 1 else log.warning)(msg)
            time.sleep(wait)
    raise RuntimeError("unreachable")   # pragma: no cover


def _sweep_roster(session, log, entity_code: str, entities: dict[str, dict],
                  year: int) -> int:
    """Page through one entity type's roster. Returns rows found.

    ONE sweep per entity type, not per report year. `ReportYear` does not
    filter which entities the grid returns — it only selects which
    report-period columns are displayed. Proven twice over:

      * A saved results page whose columns are dated 4/16/18 (so, a 2018
        search) lists "1 Powerful Voice", whose only downloadable year is
        2026, and "Citizens For PARAT", whose only year is 2014.
      * A real per-year sweep ran to page 69 for 1998, page 70 for 1999 and
        page 69 for 2000 — the same result set three times — with the resets
        landing on the same page numbers each pass.

    Sweeping per year therefore did 29x the work for identical data, ~16,000
    requests where ~560 suffice, which is what tripped the source's limits.
    Each row already carries that entity's full year coverage in its
    "Download Data by Year" cell, so nothing is lost.

    `year` is still sent because the form requires it; any value does.
    """
    found = 0
    for page in range(1, ROSTER_MAX_PAGES + 1):
        rows = _roster_page_with_retry(session, log, entity_code, year, page)
        for row in rows:
            _merge_entity(entities, row, entity_code)
        found += len(rows)
        if len(rows) < ROSTER_PAGE_SIZE:
            break                       # short page == last page
        time.sleep(REQUEST_PAUSE)
    else:
        log.warning(f"  ! roster for {entity_code} hit the "
                    f"{ROSTER_MAX_PAGES}-page cap — results may be truncated")
    return found


# ============================== run ====================================

def run(
    force: bool = False,
    entities: bool = False,
    transactions: bool = False,
    start_year: int | None = None,
    end_year: int | None = None,
    contributions: bool = False,
    expenditures: bool = False,
    candidates: bool = False,
    committees: bool = False,
    pause: float | None = None,
):
    """
    Download Utah campaign finance data.

    Horizontal scope:
        (no flag)                    bulk transactions + entity roster
        --transactions               bulk transaction CSVs only
        --contributions              same as --transactions (see below)
        --expenditures               same as --transactions (see below)
        --entities                   entity roster only
        --candidates / --committees  same as --entities (see below)

      Utah's export does not split contributions from expenditures: one
      CSV per (entity type, report year) contains both, keyed by a
      TRAN_TYPE column. The same file is also the only source of committee
      *names*. So --contributions/--expenditures collapse onto
      --transactions, and --candidates/--committees collapse onto
      --entities, rather than pretending to a granularity the source
      doesn't offer.

    Vertical scope (report years):
        (no flag)          incremental — skip years already in the manifest,
                           always re-fetch the current year
        --start-year YYYY  re-fetch years >= YYYY (wipes their manifest rows)
        --end-year YYYY    re-fetch years <= YYYY (combine for a range)
        --force            wipe the manifest and re-fetch everything

    Party/office enrichment sources are NOT fetched here — they live
    behind `--party` (see run_party()).
    """
    log = get_logger("utah", "scrape")
    t0  = time.perf_counter()
    log._emit("scrape_started", force=force, entities=entities,
              transactions=transactions, start_year=start_year, end_year=end_year,
              pause=pause)

    # --pause dials the gap between bulk downloads up (never down) for a
    # network or a day where the default still trips the source's threshold.
    download_pause = DOWNLOAD_PAUSE if pause is None else max(pause, 0.0)

    do_all      = not any([entities, transactions, contributions, expenditures,
                           candidates, committees])
    do_txns     = do_all or transactions or contributions or expenditures
    do_entities = do_all or entities or candidates or committees

    current_year = int(time.strftime("%Y"))
    today        = time.strftime("%Y-%m-%d")

    lo = start_year if start_year is not None else EARLIEST_YEAR
    hi = end_year   if end_year   is not None else current_year
    years = list(range(lo, hi + 1))
    year_range_explicit = start_year is not None or end_year is not None

    # An out-of-range window (--end-year 1995, --start-year 2030) leaves this
    # empty. Return instead of falling through to `years[0]`, which raised a
    # bare IndexError that the CLI's `except Exception: sys.exit(1)` then
    # swallowed — the operator saw exit 1 with no output at all.
    if not years:
        log.warning(f"  No report years in scope ({lo}-{hi}); Utah publishes "
                    f"{EARLIEST_YEAR}-{current_year}. Nothing to do.")
        log._emit("scrape_completed", status="completed",
                  duration_s=round(time.perf_counter() - t0, 1),
                  files_ok=0, files_err=0, files_skip=0, files_empty=0,
                  roster_rows=0, reason="empty_year_range")
        return

    files_ok = files_err = files_skip = files_empty = 0
    consecutive_err = 0
    roster_rows = 0

    try:
        if force:
            strip_manifest(lambda _: False)
            done = {}
        elif year_range_explicit:
            # Wipe in-range rows so the manifest is the sole authority for
            # what still needs fetching; without this the file-existence
            # check below would keep skipping them.
            def _outside(r: dict) -> bool:
                try:
                    y = int(r.get("year") or "")
                except ValueError:
                    return True                 # non-year rows always kept
                return y < lo or y > hi
            strip_manifest(_outside)
            done = load_manifest()
        else:
            done = load_manifest()

        session = RecyclingSession()

        # ---------------------- bulk transactions ----------------------
        if do_txns:
            log.info(f"Bulk transaction exports — {len(ENTITY_TYPES)} entity types "
                     f"x {years[0]}-{years[-1]}")
            for code, _label, slug in ENTITY_TYPES:
                for year in years:
                    filename = f"transactions_{slug}_{year}.csv"
                    dest     = RAW_DIR / filename
                    is_current = (year == current_year)

                    # Current year is always refetched: filings land in it
                    # continuously, so a manifest hit from this morning is
                    # already stale by this afternoon.
                    already = (
                        filename in done
                        and not is_current
                        and not year_range_explicit
                    ) or (
                        not year_range_explicit
                        and not is_current
                        and filename not in done
                        and dest.exists() and dest.stat().st_size > 0
                    )
                    if already:
                        log.file_download_skip(filename=filename)
                        files_skip += 1
                        continue

                    log.file_download_start(filename=filename)
                    t_file = time.perf_counter()
                    n_rows = _SENTINEL = object()
                    last_err = None
                    for attempt in range(1, DOWNLOAD_RETRIES + 1):
                        try:
                            n_rows = _download_bulk(session, code, year, dest)
                            break
                        except Exception as e:
                            # A cert failure is a property of this machine and
                            # its network, not of this file. Carrying on through
                            # the other 231 combinations produces 231 identical
                            # errors and buries the one line that matters.
                            if _is_tls_failure(e):
                                raise RuntimeError(_TLS_HELP) from e
                            last_err = e
                            if not _is_transient(e) or attempt == DOWNLOAD_RETRIES:
                                break
                            # Reconnect before retrying rather than reusing the
                            # handle. libcurl pools connections, so a reset one
                            # keeps failing every subsequent request through it
                            # — exactly the cascade seen on the first real run,
                            # where one curl (56) was followed by nine more
                            # with no recovery. Reconnecting also re-runs the
                            # ASM warm-up GET, so an expired TSPD cookie is
                            # fixed by the same act.
                            session.recycle()
                            wait = DOWNLOAD_BACKOFF * attempt
                            msg = (f"    {filename}: {type(e).__name__} on "
                                   f"attempt {attempt}/{DOWNLOAD_RETRIES} — "
                                   f"reconnected, retrying in {wait:.0f}s")
                            # A first reset is routine (keep-alive limit), so
                            # don't shout about it until a retry has failed too.
                            (log.debug if attempt == 1 else log.warning)(msg)
                            time.sleep(wait)

                    if n_rows is _SENTINEL:
                        log.file_download_error(filename=filename, error=str(last_err))
                        files_err += 1
                        consecutive_err += 1
                        if consecutive_err >= MAX_CONSECUTIVE_ERRORS:
                            raise RuntimeError(_RESET_HELP.format(
                                n=consecutive_err, r=DOWNLOAD_RETRIES,
                                err=str(last_err), pause=download_pause))
                        time.sleep(download_pause)
                        continue
                    consecutive_err = 0

                    if n_rows is None:
                        # No filings of this type in this year — Utah answers
                        # with an HTML page. Record it in the manifest anyway
                        # (rows=0) so the next incremental run doesn't re-ask.
                        # Emitted to JSONL, not just log.debug: roughly 130 of
                        # ~230 combinations land here, and a file_download_start
                        # with no terminal event would leave the majority of
                        # this scraper's requests unaccounted for in the run log.
                        log._emit("file_download_empty", filename=filename,
                                  entity_type=code, year=year)
                        files_empty += 1
                        if not is_current:
                            upsert_manifest({
                                "relation_type": "transactions",
                                "entity_type": code, "year": str(year),
                                "filename": filename, "bytes": 0, "rows": 0,
                                "downloaded_at": today,
                            }, done)
                        time.sleep(download_pause)
                        continue

                    log.file_download_ok(
                        filename=filename, bytes=dest.stat().st_size, rows=n_rows,
                        duration_s=round(time.perf_counter() - t_file, 2))
                    upsert_manifest({
                        "relation_type": "transactions",
                        "entity_type": code, "year": str(year),
                        "filename": filename, "bytes": dest.stat().st_size,
                        "rows": n_rows, "downloaded_at": today,
                    }, done)
                    files_ok += 1
                    time.sleep(download_pause)

        # ------------------------ entity roster ------------------------
        if do_entities:
            # One sweep per entity type — NOT per year. See _sweep_roster for
            # the evidence that ReportYear doesn't filter the grid. The year
            # sent is the newest in scope, purely to satisfy the form.
            log.info(f"Entity roster sweep — {len(ENTITY_TYPES)} entity types, "
                     f"{ROSTER_PAGE_SIZE} rows/page (one pass per type: "
                     f"ReportYear does not filter this grid)")
            # Merge into whatever is already on disk so an incremental run
            # doesn't drop filers a previous run had found.
            roster = _load_entities()
            before = len(roster)
            # Timed and error-counted separately from the bulk phase — folding
            # t0/files_err in would report the whole run's duration and the
            # transaction downloads' failures as if they were the sweep's.
            t_roster = time.perf_counter()
            roster_err = 0
            form_year = years[-1]
            for code, _label, _slug in ENTITY_TYPES:
                try:
                    type_rows = _sweep_roster(session, log, code, roster,
                                              form_year)
                    consecutive_err = 0
                except Exception as e:
                    if _is_tls_failure(e):
                        raise RuntimeError(_TLS_HELP) from e
                    log.page_scrape_error(entity=f"roster/{code}",
                                          page_id=form_year, error=str(e))
                    files_err += 1
                    roster_err += 1
                    type_rows = 0
                    # Same stop condition as the download loop: a run of
                    # exhausted retries means the source has stopped talking
                    # to us, and an entities.csv written from a mostly-failed
                    # sweep would be worse than the one already on disk.
                    consecutive_err += 1
                    if consecutive_err >= MAX_CONSECUTIVE_ERRORS:
                        raise RuntimeError(_RESET_HELP.format(
                            n=consecutive_err, r=DOWNLOAD_RETRIES,
                            err=str(e), pause=download_pause)) from e
                roster_rows += type_rows
                log.info(f"    {code}: {type_rows:,} roster rows")
                time.sleep(REQUEST_PAUSE)

            n_entities = _write_entities(roster)
            log.page_scrape_complete(
                filename=str(RAW_DIR / ENTITIES_FILE), rows=n_entities,
                duration_s=round(time.perf_counter() - t_roster, 1),
                ok=roster_rows, err=roster_err)
            log.info(f"  {ENTITIES_FILE}: {n_entities:,} distinct entities "
                     f"({n_entities - before:+,} vs previous run)")
            upsert_manifest({
                "relation_type": "entities", "entity_type": "", "year": "all",
                "filename": ENTITIES_FILE,
                "bytes": (RAW_DIR / ENTITIES_FILE).stat().st_size,
                "rows": n_entities, "downloaded_at": today,
            }, done)

        duration = round(time.perf_counter() - t0, 1)
        log.info(f"Done in {duration}s — {files_ok} downloaded, {files_skip} skipped, "
                 f"{files_empty} empty type/year combos, {files_err} errors")
        log._emit("scrape_completed", status="completed", duration_s=duration,
                  files_ok=files_ok, files_err=files_err, files_skip=files_skip,
                  files_empty=files_empty, roster_rows=roster_rows)

    except KeyboardInterrupt:
        log.warning("Interrupted")
        log._emit("scrape_completed", status="interrupted",
                  duration_s=round(time.perf_counter() - t0, 1),
                  files_ok=files_ok, files_err=files_err)
        raise

    except Exception as e:
        log._emit("scrape_completed", status="error",
                  duration_s=round(time.perf_counter() - t0, 1),
                  files_ok=files_ok, files_err=files_err,
                  error_type=type(e).__name__, error=str(e))
        raise


# ============ Party/office enrichment sources (--party) ================
#
# Utah's disclosure data has no party, office or district: a PCC filer is
# a bare personal name. These two sources fill that gap. Both are
# best-effort — an unreachable source logs a warning and contributes zero
# rows rather than failing the run, because a blank party is a much better
# outcome than a wrong one (parsers/utah.py only ever *fills* a blank).

PARTY_MANIFEST      = PROJECT_ROOT / "data" / "Utah" / "party_manifest.csv"
PARTY_MANIFEST_COLS = ["source", "filename", "row_count", "fetched_at"]

PARTY_REQUEST_PAUSE = 0.3     # polite delay between requests to any one host

# https://data.openstates.org/people/current/ut.csv — the same CC0,
# no-key-required nightly export scrapers/new_york.py and scrapers/texas.py
# already use in place of the v3 REST API, for the same reason: v3 needs an
# OPENSTATES_API_KEY, which would make this source depend on a per-user
# credential nobody else running this pipeline has.
PARTY_OPENSTATES_CURRENT = "https://data.openstates.org/people/current/ut.csv"

# The CSV export uses the same upper/lower classification the v3 API's
# current_role.org_classification did.
_PARTY_OS_CHAMBERS = {"upper": "State Senator", "lower": "State Representative"}

OPENSTATES_FILE = "OpenStates_People.csv"
OPENSTATES_COLS = [
    "openstates_id", "name", "given_name", "family_name",
    "party", "chamber", "district", "entity_id",
]

CANVASS_INDEX = "https://vote.utah.gov/historical-election-results/"
CANVASS_FILE  = "UT_ElectionResults.csv"
CANVASS_COLS  = ["election_year", "stage", "race_raw", "office", "district",
                 "candidate_name", "party", "source_file"]

# Excel only. The PDF-only canvasses (2024, 2022, and everything before
# 2008 except 2000) are deliberately not parsed — pdfplumber on 60 years of
# inconsistent scanned canvass layouts is a project of its own, and the
# Excel span (2000, 2008-2020) already covers the years where Utah's
# itemized finance data is dense.
_CANVASS_EXT_RE = re.compile(r"\.(xlsx|xls)$", re.I)
_CANVASS_YEAR_RE = re.compile(r"\b(19|20)\d{2}\b")


def party_clean(val) -> str:
    return re.sub(r"\s+", " ", (val or "").strip())


def _party_session() -> "requests.Session":
    # Same OS-trust-store handling as _new_session() — see the comment there.
    s = requests.Session(impersonate=IMPERSONATE, verify=config.ca_bundle())
    s.headers.update({"Accept": "text/csv,text/html;q=0.9,*/*;q=0.8",
                      "Accept-Language": "en-US,en;q=0.9"})
    return s


def _party_get(sess, url: str, retries: int = 3, timeout: int = 60, **kwargs):
    """GET with linear backoff. Returns None on any failure — callers treat a
    miss as "this source has nothing to say", never as a reason to abort."""
    for attempt in range(retries):
        try:
            resp = sess.get(url, timeout=timeout, **kwargs)
            if resp.status_code == 404:
                return None
            resp.raise_for_status()
            return resp
        except Exception:                                  # noqa: BLE001
            time.sleep(1.5 * (attempt + 1))
    return None


def _party_iter_csv(text: str):
    """Yield dict rows from CSV text, tolerating a UTF-8 BOM."""
    return csv.DictReader((text or "").lstrip("﻿").splitlines())


def _party_upsert_manifest(record: dict) -> None:
    existing = []
    if PARTY_MANIFEST.exists():
        with open(PARTY_MANIFEST, newline="", encoding="utf-8") as f:
            existing = [r for r in csv.DictReader(f)
                        if r.get("source") != record["source"]]
    with open(PARTY_MANIFEST, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=PARTY_MANIFEST_COLS,
                           extrasaction="ignore", restval="")
        w.writeheader()
        w.writerows(existing)
        w.writerow(record)


def _write_party_csv(filename: str, cols: list[str], rows: list[dict]) -> int:
    out  = RAW_DIR / filename
    part = out.with_suffix(out.suffix + ".part")
    with open(part, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore", restval="")
        w.writeheader()
        w.writerows(rows)
    part.replace(out)
    return len(rows)


# ------------------------- Open States source -------------------------

def scrape_openstates_party(sess, log) -> int:
    """Open States nightly bulk CSV for Utah -> raw/OpenStates_People.csv.

    Beyond name/party/chamber/district, this pulls out the one thing that
    makes Utah's party join exact rather than a guess: every legislator's
    `links`/`sources` list includes their own
    disclosures.utah.gov/.../FolderDetails/<id> URL, which is the same
    entity_id the roster sweep records as state_filer_id. Written to the
    `entity_id` column; blank when a person's record has no such link.
    """
    resp = _party_get(sess, PARTY_OPENSTATES_CURRENT, timeout=120)
    if resp is None or not resp.text.strip():
        log.warning("  Open States bulk CSV unavailable — skipping this source")
        return 0

    rows = []
    for src in _party_iter_csv(resp.text):
        name = party_clean(src.get("name"))
        if not name:
            continue
        district = party_clean(src.get("current_district"))
        chamber  = party_clean(src.get("current_chamber")).lower()
        # Pull the disclosures entity id out of either link column — some
        # records carry it under `sources` rather than `links`.
        blob = f"{src.get('links') or ''};{src.get('sources') or ''}"
        m = _FOLDER_ID_RE.search(blob)
        rows.append({
            "openstates_id": party_clean(src.get("id")),
            "name":          name,
            "given_name":    party_clean(src.get("given_name")),
            "family_name":   party_clean(src.get("family_name")),
            "party":         party_clean(src.get("current_party")),
            "chamber":       _PARTY_OS_CHAMBERS.get(chamber, ""),
            "district":      district.lstrip("0") or district,
            "entity_id":     m.group(1) if m else "",
        })

    n_ids = sum(1 for r in rows if r["entity_id"])
    log.info(f"  Open States: {len(rows):,} people, {n_ids:,} carry a "
             f"disclosures.utah.gov entity id (exact-join candidates)")
    return _write_party_csv(OPENSTATES_FILE, OPENSTATES_COLS, rows)


# --------------------------- canvass source ---------------------------
#
# !! The workbook layouts below are UNVERIFIED — see the module docstring.
# Everything here is written so that a layout mismatch yields zero rows and
# a warning rather than plausible-looking nonsense.

# Party labels Utah's canvasses actually use. Deliberately a closed set: a
# cell only counts as a party cell if it matches one of these, so a stray
# "Total" or a county name can't be mistaken for a party.
_CANVASS_PARTY_LABELS = {
    "REP", "REPUBLICAN", "R",
    "DEM", "DEMOCRATIC", "DEMOCRAT", "D",
    "LIB", "LBT", "LIBERTARIAN",
    "CON", "CST", "CONSTITUTION", "CONSTITUTION PARTY",
    "IAP", "INDEPENDENT AMERICAN", "INDEPENDENT AMERICAN PARTY",
    "GRN", "GRE", "GREEN", "GREEN PARTY",
    "UUP", "UNITED UTAH", "UNITED UTAH PARTY",
    "UNA", "UNAFFILIATED", "IND", "INDEPENDENT",
    "WRITE-IN", "WRITE IN", "WRI",
    "NPA", "NONPARTISAN", "NON-PARTISAN",
}

# Utah's ballot abbreviations expanded to unambiguous names BEFORE anything
# reaches src/aliases/parties.csv. That table is national, not state-keyed,
# and two of Utah's codes collide with other states' parties there: "CON" is
# Utah's Constitution Party but New York's Conservative Party, and "IAP" is
# Utah's Independent American Party, which parties.csv doesn't know at all.
# Writing the expanded name into UT_ElectionResults.csv keeps the collision
# from ever reaching canonical_party().
# Every key must also be in _CANVASS_PARTY_LABELS or it is unreachable —
# _is_party_cell() is the gate before _canvass_party() is ever called. And
# every label variant must have an entry here, including the "... PARTY"
# long forms, or the same Utah party reaches candidates.party under two
# different strings depending on how one workbook happened to spell it.
_CANVASS_PARTY_CANON = {
    "R": "Republican", "REP": "Republican",
    "D": "Democratic", "DEM": "Democratic",
    "L": "Libertarian", "LIB": "Libertarian", "LBT": "Libertarian",
    "C": "Constitution", "CON": "Constitution", "CST": "Constitution",
    "CONSTITUTION PARTY": "Constitution",
    "IAP": "Independent American",
    "INDEPENDENT AMERICAN PARTY": "Independent American",
    "UUP": "United Utah", "UNITED UTAH PARTY": "United Utah",
    "G": "Green", "GRN": "Green", "GRE": "Green", "GREEN PARTY": "Green",
    # Rocky Anderson's Justice Party — on Utah ballots 2012-2016.
    "J": "Justice", "JP": "Justice", "JUSTICE PARTY": "Justice",
    # Single "U" on the 2008-2012 canvasses is Unaffiliated, not United Utah
    # (that party wasn't founded until 2017): the 2010 gubernatorial "Anderson
    # & Maxfield \"U\"" line is Farley Anderson, who ran unaffiliated.
    "U": "Unaffiliated", "UNA": "Unaffiliated",
    "WRI": "Write-in", "WRITE IN": "Write-in", "WRITE-IN": "Write-in",
    "NPA": "Nonpartisan", "NON-PARTISAN": "Nonpartisan",
    "PEACE AND FREEDOM": "Peace and Freedom",
    "PARTY FOR SOCIALISM AND LIBERATION": "Socialism and Liberation",
    "REPUBLICAN": "Republican", "DEMOCRATIC": "Democratic",
    "DEMOCRAT": "Democratic", "LIBERTARIAN": "Libertarian",
    "CONSTITUTION": "Constitution", "GREEN": "Green", "JUSTICE": "Justice",
    "UNAFFILIATED": "Unaffiliated", "INDEPENDENT AMERICAN": "Independent American",
    "UNITED UTAH": "United Utah", "NONPARTISAN": "Nonpartisan",
}


def _canvass_party(label: str) -> str:
    """Expand a Utah ballot party abbreviation; pass long forms through."""
    up = party_clean(label).upper().rstrip(".")
    if up.startswith("WRITE"):          # "Write-in", "Write In", "Write -in"
        return "Write-in"
    return _CANVASS_PARTY_CANON.get(up, party_clean(label))


# Words that disqualify a cell from being read as a person's name.
_CANVASS_NAME_STOPWORDS = {
    "TOTAL", "TOTALS", "VOTES", "CANDIDATE", "CANDIDATES", "PARTY", "RACE",
    "COUNTY", "PRECINCT", "DISTRICT", "OFFICE", "PERCENT", "REGISTERED",
    "TURNOUT", "BALLOTS", "CAST", "GRAND TOTAL", "STATEWIDE", "YES", "NO",
    "FOR", "AGAINST", "WRITE-IN", "WRITE IN", "UNOPPOSED", "NAME",
}

# "Jane Q. Public (REP)" — the other common shape, party inline in one cell.
_CANVASS_INLINE_RE = re.compile(
    r"^(?P<name>[^()]{3,60}?)\s*\(\s*(?P<party>[A-Za-z][A-Za-z .\-]{0,30})\s*\)\s*$")

_CANVASS_OFFICE_PATTERNS: list[tuple[re.Pattern, str]] = [
    # Longest / most specific first. Utah elects governor and lieutenant
    # governor on one ticket and the canvasses title that race
    # "Governor / Lieutenant Governor", so the joint pattern MUST be tested
    # before the bare "lieutenant governor" one or every gubernatorial
    # candidate is filed under Lt. Governor. Likewise "u.s. senate" must
    # beat "state senate".
    (re.compile(r"governor\s*(?:/|&|and)\s*(?:lieutenant|lt\.?)", re.I),
     "Governor/Lt. Governor Ticket"),
    (re.compile(r"lieutenant\s+governor|lt\.?\s*governor", re.I), "Lt. Governor"),
    (re.compile(r"\bgovernor\b", re.I),                    "Governor"),
    (re.compile(r"attorney\s+general", re.I),              "Attorney General"),
    (re.compile(r"state\s+auditor|\bauditor\b", re.I),     "State Auditor"),
    (re.compile(r"state\s+treasurer|\btreasurer\b", re.I), "State Treasurer"),
    (re.compile(r"(u\.?\s*s\.?|united\s+states)\s*senat", re.I), "U.S. Senator"),
    (re.compile(r"(u\.?\s*s\.?|united\s+states)\s*(house|representative)"
                r"|\bcongress(ional)?\b", re.I),           "U.S. Representative"),
    (re.compile(r"state\s+senat|senate\s+district", re.I), "State Senator"),
    (re.compile(r"state\s+(house|representative)|house\s+district"
                r"|state\s+legislative", re.I),            "State Representative"),
    (re.compile(r"(state\s+)?(board\s+of\s+education|school\s+board)", re.I),
     "State Board of Education"),
    (re.compile(r"supreme\s+court", re.I),                 "State Supreme Court Justice"),
    (re.compile(r"court\s+of\s+appeals", re.I),            "Court of Appeals Judge"),
    (re.compile(r"district\s+court", re.I),                "District Court Judge"),
    (re.compile(r"juvenile\s+court", re.I),                "District Court Judge"),
    (re.compile(r"county\s+attorney|district\s+attorney", re.I), "County Prosecutor"),
    (re.compile(r"\bsheriff\b", re.I),                     "County Sheriff"),
    (re.compile(r"county\s+commission", re.I),             "County Commissioner"),
    (re.compile(r"county\s+council", re.I),                "County Council Member"),
    (re.compile(r"county\s+clerk", re.I),                  "County Clerk"),
    (re.compile(r"county\s+auditor", re.I),                "County Auditor"),
    (re.compile(r"county\s+assessor", re.I),               "County Assessor"),
    (re.compile(r"county\s+treasurer", re.I),              "County Treasurer"),
    (re.compile(r"county\s+recorder", re.I),               "County Clerk"),
    (re.compile(r"county\s+surveyor", re.I),               "Other"),
    (re.compile(r"\bmayor\b", re.I),                       "Mayor"),
    (re.compile(r"city\s+council", re.I),                  "City Council Member"),
    (re.compile(r"\bpresident\b.*\bvice\b|president\s+of\s+the\s+united", re.I),
     "Other"),
]

_CANVASS_DISTRICT_RE = re.compile(
    r"district\s*(?:no\.?\s*|#\s*)?(\d{1,3})\b", re.I)
_CANVASS_SEAT_RE = re.compile(r"\b(?:seat|position)\s*([A-Z0-9]{1,3})\b", re.I)


def _canvass_office(race_raw: str) -> str:
    for rx, label in _CANVASS_OFFICE_PATTERNS:
        if rx.search(race_raw or ""):
            return label
    return ""


def _canvass_district(race_raw: str) -> str:
    m = _CANVASS_DISTRICT_RE.search(race_raw or "")
    if m:
        return m.group(1).lstrip("0") or m.group(1)
    m = _CANVASS_SEAT_RE.search(race_raw or "")
    return m.group(1) if m else ""


def _is_party_cell(text: str) -> bool:
    return party_clean(text).upper().rstrip(".") in _CANVASS_PARTY_LABELS


def _looks_like_person(text: str) -> bool:
    """Conservative name test: 2-6 alphabetic tokens, no digits, not a header."""
    t = party_clean(text)
    if not (4 <= len(t) <= 60) or any(ch.isdigit() for ch in t):
        return False
    up = t.upper()
    if up in _CANVASS_NAME_STOPWORDS or _is_party_cell(t):
        return False
    if any(w in up for w in ("TOTAL", "PRECINCT", "REGISTERED", "PERCENT")):
        return False
    toks = [tk for tk in re.split(r"[\s,.]+", t) if tk]
    if not (2 <= len(toks) <= 6):
        return False
    return all(re.fullmatch(r"[A-Za-z][A-Za-z'\-]*", tk) for tk in toks)


def _iter_sheets(path: Path):
    """Yield (sheet_name, rows) per worksheet, rows as lists of cell strings.

    Column positions are preserved (blanks included) because the layout is
    positional: a race title sits above the candidate columns it covers.
    Sheets are materialized rather than streamed — canvass sheets are tiny
    (tens of rows) and the parser needs random access to two rows at once.

    openpyxl handles .xlsx, xlrd handles legacy .xls; both are already in
    requirements.txt. An unreadable workbook yields nothing.
    """
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError:                                # pragma: no cover
            return
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            return
        try:
            for ws in wb.worksheets:
                yield ws.title, [["" if c is None else str(c) for c in row]
                                 for row in ws.iter_rows(values_only=True)]
        finally:
            wb.close()
    elif suffix == ".xls":
        try:
            import xlrd
        except ImportError:                                # pragma: no cover
            return
        try:
            wb = xlrd.open_workbook(str(path))
        except Exception:
            return
        for ws in wb.sheets():
            yield ws.name, [["" if v is None else str(v)
                             for v in ws.row_values(r)] for r in range(ws.nrows)]


# --- candidate cell -> (name, party) ------------------------------------
#
# Both markers occur, and BOTH forms appear with the party spelled either as
# a single letter or in full:
#     Rob Bishop "R"                      2008/2010 generals
#     McCain & Palin "Republican"         2008 president
#     Mitt Romney, Paul Ryan (R)          2012 generals
#     Loy Brunson (Write-In)              write-ins, either style
#
# Parenthesised form is tried FIRST because the double quote is overloaded:
# candidate nicknames are quoted too, and 'Michael L."Mike" Binyon (D)' would
# otherwise yield party='Mike'. For the quoted form the LAST quoted run wins,
# for the same reason.
_CAND_PAREN_RE  = re.compile(r'^(?P<name>.+?)\s*\((?P<party>[^()]{1,40})\)\s*$')
_CAND_QUOTED_RE = re.compile(r'^(?P<name>.+?)\s*"(?P<party>[^"]{1,40})"?\s*$')

# Header cells that are vote-tally columns rather than candidates.
_CANVASS_SKIP_COL_RE = re.compile(
    r'^\s*(county|active\s+registered|registered|votes?\s+cast|ballots?\s+cast|'
    r'total\s+ballots|percent|percentage|%|number|total)\b', re.I)

# Party named inside a race title — how primaries state it
# ("State House District 20 Republican", "Governor     Constitution").
_TITLE_PARTY_RE = re.compile(
    r'\b(republican|democratic|democrat|libertarian|constitution|green|justice|'
    r'independent\s+american|united\s+utah|unaffiliated)\b', re.I)


def _is_party_token(tok: str) -> bool:
    """True only for something that really names a party.

    Guards the quoted form against candidate nicknames: `David "Dave" L.
    Thomas` carries no party at all, but a regex looking for a trailing
    quoted run happily reads the text after the nickname and reports
    party='L. Thomas'. Requiring the token to be one we actually recognise
    turns that into "no party on this cell", which then correctly falls
    through to the race title.
    """
    up = party_clean(tok).upper().rstrip(".")
    return bool(up) and (up.startswith("WRITE")
                         or up in _CANVASS_PARTY_CANON
                         or up in _CANVASS_PARTY_LABELS)


# Utah's 29 counties. Used to find where a canvass sheet's data begins, which
# is a far sturdier anchor than the literal header string "COUNTY": the 2020
# primary's header row leaves column 0 EMPTY and starts straight in on "Beaver
# County", so a string match finds nothing and the whole workbook silently
# yields zero rows.
_UTAH_COUNTIES = {
    "BEAVER", "BOX ELDER", "CACHE", "CARBON", "DAGGETT", "DAVIS", "DUCHESNE",
    "EMERY", "GARFIELD", "GRAND", "IRON", "JUAB", "KANE", "MILLARD", "MORGAN",
    "PIUTE", "RICH", "SALT LAKE", "SAN JUAN", "SANPETE", "SEVIER", "SUMMIT",
    "TOOELE", "UINTAH", "UTAH", "WASATCH", "WASHINGTON", "WAYNE", "WEBER",
}


def _is_county_row(row: list) -> bool:
    """True when this row's first cell names a Utah county ('Beaver County')."""
    if not row:
        return False
    v = party_clean(row[0]).upper()
    v = re.sub(r"\s+COUNTY$", "", v)
    return v in _UTAH_COUNTIES


def _nearest_left(row: list, col: int) -> str:
    """Nearest non-empty cell at or to the left of `col`.

    Race titles and party labels are written once, at the first column of the
    span they cover, so a candidate three columns along still belongs to them.
    """
    for j in range(min(col, len(row) - 1), -1, -1):
        v = party_clean(row[j])
        if v:
            return v
    return ""


def _split_candidate(cell: str) -> tuple[str, str]:
    """('Candidate Name', 'party token') from a header cell, or ('', '')."""
    text = party_clean(cell)
    if not text or _CANVASS_SKIP_COL_RE.match(text):
        return "", ""
    for rx in (_CAND_PAREN_RE, _CAND_QUOTED_RE):
        m = rx.match(text)
        if m and _is_party_token(m.group("party")):
            return party_clean(m.group("name")), party_clean(m.group("party"))
    return text, ""             # a bare name — party may come from the title


def _ticket_lead(name: str) -> str:
    """First name of a joint ticket.

    'Huntsman & Herbert' -> 'Huntsman', 'Mitt Romney, Paul Ryan' -> 'Mitt
    Romney'. Canvass names are First-Last, so a comma separates running mates
    rather than surname-first — the opposite of the disclosure roster.
    """
    return party_clean(re.split(r"\s*[&,]\s*", name)[0]) if name else ""


def _extract_canvass(path: Path, year: str, stage: str) -> list[dict]:
    """Pull (candidate, party, race) out of one Utah canvass workbook.

    The layout, confirmed against all 16 real 2000-2020 files rather than
    assumed, is a county matrix — one worksheet per office group, one row per
    county, and **candidates in a header row rather than in any data row**:

        …        one or more context rows: race titles, and sometimes a
                 separate row of party labels
        hdr      the candidate row
        hdr+1…   one row per county, vote counts

    Anchoring is on the first row whose column 0 names a Utah county, so the
    candidate row is simply the row above it. The obvious alternative — look
    for a cell reading "COUNTY" — breaks on the 2020 primary, which leaves
    that cell empty.

    Two things vary and are both handled by walking upward from the candidate
    row:

      * **Where the party is stated.** On the cell itself in most years
        (`Rob Bishop "R"`, `Mitt Romney, Paul Ryan (R)`); in a dedicated row
        of its own in the 2020 primary; in the race title on other primaries
        (`State House District 20 Republican`).
      * **How far above the race title sits.** Directly above the candidates
        in the generals, three rows up in the 2020 primary.

    For each candidate column the first party-looking value found above it
    becomes the party and the first non-party-looking value becomes the race,
    each matched at or to the left of that column. A sheet with no county
    rows is skipped (cover pages, notes, ballot-total tabs).
    """
    out: list[dict] = []
    seen: set[tuple] = set()

    for sheet, rows in _iter_sheets(path):
        data_i = next((i for i, r in enumerate(rows) if _is_county_row(r)), None)
        if data_i is None:
            # Fallback for a sheet whose first data row isn't a recognizable
            # county: the classic "COUNTY" header cell.
            hdr_i = next((i for i, r in enumerate(rows[:15])
                          if r and party_clean(r[0]).upper() == "COUNTY"), None)
            if hdr_i is None:
                continue
        else:
            hdr_i = data_i - 1
        if hdr_i < 0 or hdr_i >= len(rows):
            continue

        header = rows[hdr_i]
        context = [rows[i] for i in range(hdr_i - 1, -1, -1)]   # nearest first

        for col, cell in enumerate(header):
            name, party_tok = _split_candidate(cell)
            if not name:
                continue

            race = ""
            for ctx in context:
                val = _nearest_left(ctx, col)
                if not val:
                    continue
                if not party_tok and _is_party_token(val):
                    party_tok = val
                elif not race and not _is_party_token(val):
                    race = val
                if race and party_tok:
                    break

            if not party_tok:
                # Primaries that name the party inside the race title.
                tm = _TITLE_PARTY_RE.search(race)
                if not tm:
                    continue    # nonpartisan/judicial race, or not a candidate
                party_tok = tm.group(1)

            party = _canvass_party(party_tok)
            lead = _ticket_lead(name)
            if not lead or len(lead) < 4:
                continue
            office = _canvass_office(race) or _canvass_office(sheet)
            district = _canvass_district(race)
            key = (lead.upper(), party.upper(), office, district)
            if key in seen:
                continue
            seen.add(key)
            out.append({
                "election_year":  year,
                "stage":          stage,
                "race_raw":       race or party_clean(sheet),
                "office":         office,
                "district":       district,
                "candidate_name": lead,
                "party":          party,
                "source_file":    path.name,
            })
    return out


def _canvass_stage(text: str) -> str:
    t = (text or "").lower()
    if "presidential primary" in t:
        return "presidential_primary"
    if "primary" in t:
        return "primary"
    if "special" in t:
        return "special"
    if "general" in t:
        return "general"
    return "other"


def _discover_canvass_links(sess, log) -> list[dict]:
    """Scrape vote.utah.gov's historical-results index for Excel canvasses.

    Returns [{url, year, stage, filename}, ...]. The index is a plain list
    of <li> items whose text carries the year and election stage and whose
    links point at /wp-content/uploads/... files, so year/stage come from
    the list item rather than the (inconsistent) filename.
    """
    resp = _party_get(sess, CANVASS_INDEX, timeout=90)
    if resp is None or not resp.text.strip():
        log.warning("  vote.utah.gov historical-results index unreachable "
                    "— skipping the canvass source")
        return []

    soup = BeautifulSoup(resp.text, "html.parser")
    found, seen_urls = [], set()
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if not _CANVASS_EXT_RE.search(href.split("?")[0]):
            continue
        url = urljoin(CANVASS_INDEX, href)
        if url in seen_urls:
            continue
        # Prefer the enclosing list item's text for year/stage; the anchor
        # text itself is often just "Excel".
        li = a.find_parent("li")
        context = party_clean(li.get_text(" ", strip=True)) if li else \
            party_clean(a.get_text(strip=True))
        ym = _CANVASS_YEAR_RE.search(context) or _CANVASS_YEAR_RE.search(href)
        if not ym:
            continue
        year  = ym.group(0)
        stage = _canvass_stage(context)
        ext   = _CANVASS_EXT_RE.search(href.split("?")[0]).group(1).lower()
        seen_urls.add(url)
        # The filename must be unique per URL, not per (year, stage): one
        # election's <li> can link more than one workbook (a statewide
        # canvass and a by-county one, say). Colliding names made the second
        # download see the first's file already on disk, skip fetching
        # entirely, and re-extract the wrong workbook twice.
        tag = hashlib.md5(url.encode()).hexdigest()[:6]
        found.append({
            "url": url, "year": year, "stage": stage,
            "filename": f"canvass_{year}_{stage}_{tag}.{ext}",
        })
    return found


def scrape_canvass_party(sess, log) -> int:
    """Download + flatten Utah's Excel election canvasses.

    Workbooks land in raw/canvass/ verbatim (kept on disk so the extractor
    can be re-run and improved without re-downloading); the flattened
    result is raw/UT_ElectionResults.csv.
    """
    links = _discover_canvass_links(sess, log)
    if not links:
        return 0
    CANVASS_DIR.mkdir(parents=True, exist_ok=True)
    log.info(f"  {len(links)} Excel canvass files listed on vote.utah.gov")

    all_rows: list[dict] = []
    for item in sorted(links, key=lambda d: (d["year"], d["stage"])):
        dest = CANVASS_DIR / item["filename"]
        if not (dest.exists() and dest.stat().st_size > 0):
            resp = _party_get(sess, item["url"], timeout=180)
            if resp is None or not resp.content:
                log.warning(f"    ! {item['filename']}: download failed, skipping")
                time.sleep(PARTY_REQUEST_PAUSE)
                continue
            tmp = dest.with_suffix(dest.suffix + ".part")
            tmp.write_bytes(resp.content)
            tmp.replace(dest)
            time.sleep(PARTY_REQUEST_PAUSE)

        rows = _extract_canvass(dest, item["year"], item["stage"])
        if not rows:
            # The single most likely cause is a sheet layout the scanner
            # doesn't recognize — say so explicitly rather than silently
            # contributing nothing.
            log.warning(f"    ! {item['filename']}: 0 candidate/party rows "
                        f"extracted — layout may not match the scanner "
                        f"(see _extract_canvass in scrapers/utah.py)")
        else:
            log.info(f"    {item['filename']}: {len(rows):,} rows")
        all_rows.extend(rows)

    if not all_rows:
        log.warning("  Canvass extraction produced no rows at all — "
                    "UT_ElectionResults.csv not written")
        return 0
    return _write_party_csv(CANVASS_FILE, CANVASS_COLS, all_rows)


def run_party(openstates: bool = False, canvass: bool = False):
    """Fetch Utah's external party/office sources into data/Utah/raw/.

    Sources are additive: no flag means all of them.
        --openstates   Open States nightly bulk CSV only (fast)
        --canvass      vote.utah.gov Excel canvasses only
    """
    log = get_logger("utah", "scrape")
    t0  = time.perf_counter()
    log.info("Starting Utah party-enrichment scraper")
    log._emit("scrape_started", source="utah_party",
              openstates=openstates, canvass=canvass)

    do_os      = openstates or not (openstates or canvass)
    do_canvass = canvass    or not (openstates or canvass)

    today = time.strftime("%Y-%m-%d")
    n_os = n_canvass = 0

    try:
        # Inside the try: a failure constructing the session must still emit
        # scrape_completed, since scrape_started has already fired.
        sess = _party_session()

        if do_os:
            log.info("Open States (data.openstates.org/people/current/ut.csv)")
            n_os = scrape_openstates_party(sess, log)
            # Both scrape_* helpers return 0 WITHOUT writing anything when the
            # source is unreachable. Stamping a manifest row and logging
            # "wrote ..." for a file that may not exist would make an outage
            # read like an empty source on the next run.
            if n_os:
                _party_upsert_manifest({"source": "openstates",
                                        "filename": OPENSTATES_FILE,
                                        "row_count": n_os, "fetched_at": today})
                log.info(f"  wrote {OPENSTATES_FILE} ({n_os:,} rows)")
            else:
                log.warning(f"  {OPENSTATES_FILE} not written — source returned "
                            f"nothing; any existing copy is left untouched")

        if do_canvass:
            log.info("Utah historical election canvasses (vote.utah.gov)")
            n_canvass = scrape_canvass_party(sess, log)
            if n_canvass:
                _party_upsert_manifest({"source": "canvass",
                                        "filename": CANVASS_FILE,
                                        "row_count": n_canvass,
                                        "fetched_at": today})
                log.info(f"  wrote {CANVASS_FILE} ({n_canvass:,} rows)")
            else:
                log.warning(f"  {CANVASS_FILE} not written — no rows extracted; "
                            f"any existing copy is left untouched")

        duration = round(time.perf_counter() - t0, 1)
        log.info(f"Done in {duration}s")
        log._emit("scrape_completed", status="completed", duration_s=duration,
                  source="utah_party", openstates_rows=n_os,
                  canvass_rows=n_canvass)

    except KeyboardInterrupt:
        log._emit("scrape_completed", status="interrupted",
                  duration_s=round(time.perf_counter() - t0, 1),
                  source="utah_party", openstates_rows=n_os,
                  canvass_rows=n_canvass)
        raise

    except Exception as e:
        log._emit("scrape_completed", status="error",
                  duration_s=round(time.perf_counter() - t0, 1),
                  source="utah_party", openstates_rows=n_os,
                  canvass_rows=n_canvass,
                  error_type=type(e).__name__, error=str(e))
        raise


# ================================ CLI ==================================

if __name__ == "__main__":
    import argparse
    from datetime import datetime

    # --party is a separate, manual mode with its own parser. It is not in
    # orc.py's forwarded flag set, so this branch is only ever reached by a
    # human running the file directly.
    if "--party" in sys.argv:
        pp = argparse.ArgumentParser(
            description="Download Utah party-enrichment sources "
                        "(Open States bulk CSV + vote.utah.gov canvasses).")
        pp.add_argument("--party",      action="store_true", help=argparse.SUPPRESS)
        pp.add_argument("--openstates", action="store_true",
                        help="Open States bulk CSV only")
        pp.add_argument("--canvass",    action="store_true",
                        help="vote.utah.gov Excel canvasses only")
        pargs, _ = pp.parse_known_args()
        try:
            run_party(openstates=pargs.openstates, canvass=pargs.canvass)
        except KeyboardInterrupt:
            sys.exit(130)
        except Exception as e:
            print(f"\n{e}\n", file=sys.stderr)
            sys.exit(1)
        sys.exit(0)

    ap = argparse.ArgumentParser(
        description="Download Utah campaign finance data from disclosures.utah.gov.")

    vert = ap.add_mutually_exclusive_group()
    vert.add_argument("--force",      action="store_true",
                      help="wipe the manifest and re-download everything in scope")
    vert.add_argument("--start-year", type=int, metavar="YYYY",
                      help="earliest report year to download (inclusive)")

    ap.add_argument("--end-year", type=int, metavar="YYYY",
                    help="latest report year to download (inclusive, <= current year)")

    ap.add_argument("--transactions",  action="store_true",
                    help="bulk transaction CSVs only")
    ap.add_argument("--entities",      action="store_true",
                    help="entity roster only")
    ap.add_argument("--contributions", action="store_true",
                    help="alias for --transactions (Utah's export doesn't split them)")
    ap.add_argument("--expenditures",  action="store_true",
                    help="alias for --transactions (Utah's export doesn't split them)")
    ap.add_argument("--candidates",    action="store_true",
                    help="alias for --entities (one roster covers both)")
    ap.add_argument("--committees",    action="store_true",
                    help="alias for --entities (one roster covers both)")

    # Not part of orc.py's forwarded flag set — a local escape hatch for a
    # network or a day where the default pacing still trips the source's
    # threshold. See _RESET_HELP.
    ap.add_argument("--pause", type=float, metavar="SECONDS",
                    help=f"seconds between bulk downloads "
                         f"(default {DOWNLOAD_PAUSE}; raise if you get curl 56)")

    args, _ = ap.parse_known_args()

    cy = datetime.today().year
    if args.end_year:
        if args.end_year > cy:
            ap.error(f"--end-year cannot exceed current year ({cy})")
        if args.start_year and args.start_year > args.end_year:
            ap.error("--start-year cannot be greater than --end-year")
    if args.force and args.end_year:
        ap.error("--force cannot be combined with --end-year")

    try:
        run(
            force=args.force,
            entities=args.entities,
            transactions=args.transactions,
            start_year=args.start_year,
            end_year=args.end_year,
            contributions=args.contributions,
            expenditures=args.expenditures,
            candidates=args.candidates,
            committees=args.committees,
            pause=args.pause,
        )
    except KeyboardInterrupt:
        sys.exit(130)
    except Exception as e:
        # The bare `except Exception: sys.exit(1)` the other scrapers use
        # loses the message entirely when the file is run directly — orc.py
        # captures a subprocess's stderr, but a terminal has nothing to
        # capture if nothing was written. A TLS failure in particular carries
        # several lines of actionable guidance; exiting silently on it wastes
        # the diagnosis. Exit code is unchanged, so orc.py behaves as before.
        print(f"\n{e}\n", file=sys.stderr)
        sys.exit(1)
