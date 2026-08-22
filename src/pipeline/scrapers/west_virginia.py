"""
scrapers/west_virginia.py — Download West Virginia campaign finance data.

Source: WV Secretary of State Campaign Finance Reporting System (CFRS), a
React single-page app at https://cfrs.wvsos.gov/ backed by a JSON service
at https://cfrs.wvsos.gov/api/Public-Service/.

No Playwright. Every request is a plain requests call — CFRS runs no JS
challenge or bot filter on the API service, only the SPA shell needs a
browser, and we skip the shell entirely.

────────────────────────────────────────────────────────────────────────────
Where the data comes from
────────────────────────────────────────────────────────────────────────────
CFRS was rebuilt after 2023, and the rebuild RETIRED the bulk-CSV export that
earlier tooling (and the Investigative Reporting Workshop's Accountability
Project through 2023) relied on. Verified against the live site 2026-08-02,
twice, on both plausible bases:

    /api/Public-Service/DataDownload    → HTTP 404
    /CFIS_APIService/api/DataDownload   → HTTP 200 serving the React shell,
                                          i.e. the SPA catch-all for an
                                          unrouted path

The second is the nastier one: a check that only looks at the status code
would read 200 and treat a dead route as live. That download path has been
removed rather than kept as a tripwire — the old service was replaced by a
different architecture, not relocated. docs/states/west_virginia.md records
how it worked, and git history has the implementation, if it ever returns.

Transactions therefore come from the paged JSON grid:

    POST /api/Public-Service/CommitteeTransactions/getAllPublicTransactionDataList

written as newline-delimited JSON (transactions_{CAT}_{year}.jsonl), streamed
page-by-page so memory stays flat — a single year runs to tens of MB and
hundreds of thousands of rows.

parsers/west_virginia.py still reads the legacy CON_*.csv / EXP_*.csv layouts,
so files already on disk from the bulk era (or from the IRW archive) continue
to parse. Only the download path is gone.

────────────────────────────────────────────────────────────────────────────
Confirmed routes (captured from the live site)
────────────────────────────────────────────────────────────────────────────
  POST /api/Public-Service/Committee/getPublicCandidatesCommitteeDataList
       body: pageNumber, pageSize, sortColumn="registrationDate",
             sortDirection="desc", plus a block of empty/null filters.
       → the candidate + committee registry behind /public/gettoknow?tab=candidate.
         This is what supplies office / district / party / registration status,
         none of which appear in the transaction files.

  POST /api/Public-Service/CommitteeTransactions/getAllPublicTransactionDataList
       body: pageNumber, pageSize, transactionCategory="CON"|"EXP",
             sortColumn="transactionDate", sortDirection="desc",
             transactionYear="YYYY"
       → the per-row transaction grid behind /public/trackfinance?tab=CON|EXP.

  GET  /api/Public-Service/CommitteeTransactions/getContributorTypeByTransactionType/{CON|EXP}
       → the contributor/payee type vocabulary. Saved as a lookup so
         src/aliases/contributor_types.csv can be filled from the source's
         own list rather than from values observed in the data.

────────────────────────────────────────────────────────────────────────────
Service topology
────────────────────────────────────────────────────────────────────────────
CFRS is not one API. Extracted from the app's own JS bundle, where routes are
assembled as `${base}${path}`:

    /api/Public-Service    public site — the only one this scraper depends on
    /api/Common-Service    shared lookups + grid/export plumbing (authenticated)
    /api/Auth-Service      login, users, roles, admin
    /api/MessageCentral    correspondence, help library
    /api/ExternalAuth      SSO

The SAME path can exist on more than one base as separate endpoints.
/CommitteeTransactions/getContributorTypeByTransactionType is wired to
Common-Service in the bundle, yet the public site calls it on Public-Service.
A path found in the bundle therefore cannot be assumed to work against
Public-Service — the bundle's routes are largely the logged-in portal's.

LOOKUP_ENDPOINTS holds reference vocabularies (offices, parties, org types,
transaction types, purposes) whose paths were read out of that bundle and are
real, but whose public service base is unconfirmed. Each is tried against
Public-Service then Common-Service. We never send credentials, so a 401 from
Common-Service is simply a miss. All of it is enrichment for src/aliases/ —
nothing here affects the pipeline's output contract.

Confirmed live 2026-08-02: parties, committee_types, transaction_types and
occupations all resolve on Common-Service WITHOUT credentials. offices,
elections, jurisdictions, contribution_purposes and violations do not respond
on either base — their real paths are still unknown.

────────────────────────────────────────────────────────────────────────────
Whole-grid export
────────────────────────────────────────────────────────────────────────────
    POST /api/Common-Service/DataGrid/generateExportGridDataExcel

Public and unauthenticated — "moduleType": "PUBLIC" in the body is what makes
that work; we send no credentials. One call returns an entire grid.

The response embeds the spreadsheet rather than handing back a link:

    {"isSuccess": true,
     "responseData": {"fileBytes": "<base64>"},
     "message": null, "skipRecords": null}

So there is NO CloudFront round-trip, contrary to what the app bundle's
admin routes suggested. _extract_file_bytes() reads responseData.fileBytes
(base64 or byte array) and the extension comes from magic bytes, since the
route is named ...Excel but the payload isn't guaranteed to be xlsx. The
handle/link path is retained only as a fallback for the admin variants.

gridName/pageName are opaque server-side identifiers that cannot be derived,
only captured per grid (DevTools → Network → click Export). GRID_EXPORTS has
the candidate/committee registry; the TRANSACTION grids are still missing,
and they are the ones that matter — a full paged backfill measured 44 minutes
against two export calls.

────────────────────────────────────────────────────────────────────────────
TLS
────────────────────────────────────────────────────────────────────────────
cfrs.wvsos.gov serves an incomplete certificate chain — browsers paper over
it, Python's certifi bundle does not, and the site's own documented curl
examples pass --insecure. VERIFY_SSL is therefore False and urllib3's
InsecureRequestWarning is suppressed. This only disables chain validation
for this one .gov host over public, already-published data; flip VERIFY_SSL
back to True if the state ever fixes the chain.

────────────────────────────────────────────────────────────────────────────
Project integration
────────────────────────────────────────────────────────────────────────────
Output   (data/West_Virginia/raw/):
    transactions_{CAT}_{year}.jsonl per-row grid, newline-delimited — the
                                    transaction source
    export_committees.xlsx          whole-grid export (when it succeeds)
    entities_committees.json        candidate + committee registry
    lookup_contributor_types_{CAT}.json
    lookup_{offices,elections,jurisdictions,parties,committee_types,
            transaction_types,contribution_purposes,occupations,
            violations}.json        reference vocabularies (when reachable)
    _endpoints.json                 cache of resolved routes + service bases
Manifest (data/West Virginia/manifest.csv):
    relation_type, year, filename, source_url, bytes, rows, scraped_at
Logging  src.reporting.logger.get_logger("west_virginia", "scrape")

CLI:
    (no flags)          incremental — fill manifest gaps, always refresh the
                        current calendar year (CFRS amends open years in place)
    --force             wipe manifest for everything in scope, re-download
    --start-year YYYY   re-download years >= YYYY
    --end-year YYYY     re-download years <= YYYY
    --transactions      CON + EXP only
    --entities          registry + lookups only
    --contributions     CON only
    --expenditures      EXP only
    --candidates        entity sweep (CFRS returns candidates and committees
    --committees        from one route, so both flags map to the same sweep)
"""

import base64
import binascii
import csv
import io
import json
import math
import re
import socket
import sys
import time
from datetime import datetime
from pathlib import Path

import requests
import urllib3

# Make project root importable before any src.* import
PROJECT_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(PROJECT_ROOT))
from src.reporting.logger import get_logger

try:
    from config import USER_AGENT
except ImportError:      # config.py is optional for a bare checkout
    USER_AGENT = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/151.0.0.0 Safari/537.36")

# CFRS serves an incomplete cert chain — see the TLS note in the module
# docstring. Suppress the per-request warning storm this would otherwise cause.
VERIFY_SSL = False
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# =============================== paths ================================
DATA_DIR = PROJECT_ROOT / "data" / "West_Virginia"
RAW_DIR  = DATA_DIR / "raw"
MANIFEST = DATA_DIR / "manifest.csv"
ENDPOINTS_CACHE = RAW_DIR / "_endpoints.json"

RAW_DIR.mkdir(parents=True, exist_ok=True)

MANIFEST_COLS = ["relation_type", "year", "filename", "source_url",
                 "bytes", "rows", "scraped_at"]

# ========================= source constants ===========================
ORIGIN   = "https://cfrs.wvsos.gov"
API_BASE = f"{ORIGIN}/api/Public-Service"

# CFRS is split across several services (extracted from the app's JS bundle:
# Public-Service, Common-Service, Auth-Service, MessageCentral, ExternalAuth).
# Only the first two are relevant here. Common-Service is the authenticated
# portal's API — we never send credentials to it, so a 401 there is just a
# miss — but several lookup vocabularies are re-exposed publicly and it is
# worth one unauthenticated attempt.
COMMON_BASE = f"{ORIGIN}/api/Common-Service"

# Order matters: the public site's own routes win over the portal's.
LOOKUP_BASES = [API_BASE, COMMON_BASE]


# CFRS mandated e-filing produces usable bulk files from 2018 onward. Earlier
# years exist in the system as scanned paper filings but were never published
# as structured data, so probing below this only burns requests.
EARLIEST_YEAR = 2018

# TransactionType code → the relation it feeds in the cleaned schema.
# "CON" maps to contributions even though the file also carries loan rows —
# parsers/west_virginia.py splits those out onto loans_debts using the
# Receipt Type / Contribution Type columns.
TRANSACTION_TYPES = {
    "CON": "contributions",
    "EXP": "expenditures",
}

REQUEST_TIMEOUT = 180        # bulk year files reach ~15 MB and CFRS is slow
DOWNLOAD_CHUNK  = 1 << 20    # 1 MiB
RATE_LIMIT_S    = 0.35       # courtesy pause between requests
MAX_RETRIES     = 3
RETRY_BACKOFF_S = 5

GRID_PAGE_SIZE     = 1000    # rows per POST for the paginated JSON grids
GRID_MAX_PAGES     = 5000    # ceiling so a route that ignores pageNumber can't loop forever

# The only pageSize CFRS has been directly observed to accept, taken from the
# site's own XHR. _fetch_grid falls back to it when a larger size returns an
# empty first page, which is how a pageSize allow-list would manifest.
OBSERVED_PAGE_SIZE = 10


# Emit a progress line at least this often while paging. Time-based rather
# than page-based: a page-count trigger stays silent through the first N
# pages no matter how slow they are.
PROGRESS_EVERY_S = 15

BASE_HEADERS = {
    "User-Agent":      USER_AGENT,
    "Accept":          "application/json",
    "Accept-Language": "en-US,en;q=0.9",
    "Origin":          ORIGIN,
    "Referer":         f"{ORIGIN}/public/gettoknow?tab=candidate",
}

# ======================== confirmed endpoints =========================
# Captured from the live site's own XHRs. Bodies are reproduced exactly as
# the app sends them (minus pageSize, which we raise) — CFRS rejects
# requests with unexpected or missing filter keys, so the empty-string and
# null placeholders below are load-bearing, not decoration.

REGISTRY_PATH = "Committee/getPublicCandidatesCommitteeDataList"
REGISTRY_BODY = {
    "sortColumn":            "registrationDate",
    "sortDirection":         "desc",
    "entityId":              "",
    "orgStatus":             "",
    "orgSubTypeCode":        "",
    "candidateName":         "",
    "officerName":           "",
    "orgName":               "",
    "orgType":               "",
    "registrationStartDate": None,
    "registrationEndDate":   None,
    "electionID":            None,
    "officeID":              None,
    "districtID":            None,
    "partyCode":             None,
    "reportingCycleId":      None,
    "isJointFundrisingOrg":  "",
}

TRANSACTION_GRID_PATH = "CommitteeTransactions/getAllPublicTransactionDataList"
TRANSACTION_GRID_BODY = {
    "sortColumn":    "transactionDate",
    "sortDirection": "desc",
}

CONTRIBUTOR_TYPE_PATH = "CommitteeTransactions/getContributorTypeByTransactionType"

# ========================= grid export (tier 0) =======================
# Captured from the live site's Export button. This is the whole-grid export
# that replaced the retired bulk-CSV route, and the reason it matters is
# arithmetic: paging a single WV year at the fallback pageSize of 10 takes
# ~18,800 requests and over two hours, while this is two requests.
#
# Three details that are easy to get wrong:
#
#   1. It lives on Common-Service, NOT Public-Service — even though it is a
#      public, unauthenticated call. `moduleType: "PUBLIC"` is what makes it
#      work without a bearer token; we send no credentials.
#   2. `filterRequest` uses PascalCase keys (SortColumn, EntityId, OrgStatus)
#      whereas the paged grid at Public-Service uses camelCase for the very
#      same fields. Reusing one body for both silently returns an unfiltered
#      or empty result.
#   3. There is no pageNumber/pageSize. The response is a handle to a
#      generated file, not the rows.
EXPORT_PATH        = "DataGrid/generateExportGridDataExcel"
DOWNLOAD_LINK_PATH = "AmazonCloudFront/getDownloadLinkWithoutCookies"

# One entry per exportable grid. `gridName`/`pageName` are opaque server-side
# identifiers — they cannot be derived, only captured from the site's own
# Export request (DevTools → Network → Fetch/XHR → click Export).
GRID_EXPORTS: dict[str, dict] = {
    # CONFIRMED — captured 2026-08-02 from /public/gettoknow?tab=candidate.
    "committees": {
        "gridName":  "GETTOKNOW_CANDIDATECOMMITTEES",
        "pageName":  "PUB_GTK_CNCM",
        "fieldType": "G",
        "filterRequest": {
            "SortColumn":            "registrationDate",
            "SortDirection":         "desc",
            "EntityId":              "",
            "OrgStatus":             "",
            "OrgSubTypeCode":        "",
            "CandidateName":         "",
            "OfficerName":           "",
            "OrgName":               "",
            "OrgType":               "",
            "RegistrationStartDate": None,
            "RegistrationEndDate":   None,
            "ElectionID":            None,
            "OfficeID":              None,
            "DistrictID":            None,
            "PartyCode":             None,
            "ReportingCycleId":      None,
            "IsJointFundrisingOrg":  "",
        },
    },
    # CONFIRMED — captured 2026-08-02 from /public/trackfinance?tab=CON.
    #
    # Note what is ABSENT relative to the registry export above: no pageName,
    # no fieldType. Those are per-grid, not universal, which is why
    # export_grid() only sends them when a spec supplies them. It also adds
    # IsPublic, and TransactionYear is a STRING.
    #
    # {year} is substituted per call by export_transactions_year().
    "transactions_CON": {
        "gridName": "TrackFinance_Contributions",
        "filterRequest": {
            "TransactionCategory": "CON",
            "SortColumn":          "transactionDate",
            "SortDirection":       "desc",
            "TransactionYear":     "{year}",
            "IsPublic":            True,
        },
    },
    # INFERRED from the CON grid's naming, not captured. If this 404s or
    # returns isSuccess=false, the run falls back to paging that year and the
    # error names the grid — capture Export on ?tab=EXP and correct the name.
    "transactions_EXP": {
        "gridName": "TrackFinance_Expenditures",
        "filterRequest": {
            "TransactionCategory": "EXP",
            "SortColumn":          "transactionDate",
            "SortDirection":       "desc",
            "TransactionYear":     "{year}",
            "IsPublic":            True,
        },
    },
}

# Grids exported once (not per year) during the entity sweep.
ENTITY_EXPORTS = ("committees",)

# Use the whole-grid export for TRANSACTIONS? Default False.
#
# The export is dramatically cheaper — seconds per year against ~44 minutes
# for a full paged backfill — but a live comparison of the two payloads shows
# it is also materially thinner, and the gaps hit exactly the fields this
# pipeline depends on:
#
#   JSON grid                      CSV export
#   ─────────────────────────────  ──────────────────────────────────────────
#   52 fields                      11 (CON) / 10 (EXP)
#   transactionID                  ABSENT -> no filing_id, and no cross-year
#                                  dedup. CFRS republishes amended rows under
#                                  their original ID in later year files, and
#                                  that ID is the only thing that catches it.
#   orgID = 1517                   RegistrantID = 1020001517 — a DIFFERENT id,
#                                  so state_filer_id (and therefore person_id)
#                                  would not line up across sources.
#   amendedFlag                    ABSENT
#   electionYear (int)             ABSENT (falls back to the file's year)
#   city / stateCode / zipCode     one combined address; parses for ~74% of
#                                  non-empty values (18,832 of 25,517 in 2018)
#
# Loans are the one thing the export does keep: ContributionType carries
# "Loans" / "Loan Payment" / "Loan Forgiveness".
#
# Set True to trade that completeness for speed. The registry export
# (ENTITY_EXPORTS) is unaffected and stays on.
EXPORT_TRANSACTIONS_READY = False

# ========================= lookup endpoints ===========================
# Every path below was extracted from the site's own JS bundle
# (/assets/index-*.js), where routes are assembled as `${base}${path}`. They
# are REAL routes, not invented ones — an earlier version of this file
# guessed names like "Common/getOfficeList" that almost certainly never
# existed.
#
# What is still unconfirmed is which *service* serves each one publicly. The
# bundle wires most of these to Common-Service, but Common-Service is the
# authenticated portal's API: the app attaches a bearer token from
# sessionStorage to those calls. The public site re-exposes some of the same
# paths under Public-Service without auth — confirmed for
# getContributorTypeByTransactionType, whose captured public URL is on
# Public-Service even though the bundle shows it on Common-Service.
#
# So each spec is tried against Public-Service first, then Common-Service.
# We never attempt to authenticate; a 401 from Common-Service is simply a
# miss, logged and skipped. All of this is enrichment — nothing here affects
# the pipeline's output contract, and every one of them may miss.
LOOKUP_ENDPOINTS: dict[str, list[dict]] = {
    # Feeds src/aliases/office_types.csv, which is currently unpopulated for WV.
    "offices": [
        {"path": "Election/getOfficeTypeList",   "method": "GET"},
    ],
    "elections": [
        {"path": "Election/getElectionTypeList", "method": "GET"},
        {"path": "Election/getReportingCycle",   "method": "GET"},
    ],
    "jurisdictions": [
        {"path": "Election/getJurisdictionList", "method": "GET"},
    ],
    # The registry returns party as a display string; this is the source's
    # own controlled list.
    "parties": [
        {"path": "CommitteeRegistration/getAllPartyDistrictList", "method": "GET"},
    ],
    # Feeds src/aliases/committee_types.csv — the authoritative org-type
    # vocabulary, replacing the values inferred from WV Code 3-8.
    "committee_types": [
        {"path": "CommitteeRegistration/getAllOrgType",    "method": "GET"},
        {"path": "CommitteeRegistration/getAllOrgSubType", "method": "GET"},
    ],
    # Feeds src/aliases/transaction_categories.csv and expenditure_categories.csv.
    "transaction_types": [
        {"path": "CommitteeTransactions/getTransactionType", "method": "GET"},
    ],
    # Expenditure purpose vocabulary — the `purpose` / `category` columns.
    "contribution_purposes": [
        {"path": "CommitteeTransactions/getContributionPurpose",         "method": "GET"},
        {"path": "CommitteeTransactions/getContributionPurposeCategory", "method": "GET"},
    ],
    "occupations": [
        {"path": "CommitteeRegistration/getAllOccupations", "method": "GET"},
    ],
    # The one entry here still not corroborated by a bundle constant beyond
    # the bare path "/Transactions/violation" — shape and method are guesses.
    "violations": [
        {"path": "Transactions/violation", "method": "GET"},
        {"path": "Committee/getPublicViolationDataList", "method": "POST",
         "body": {"sortColumn": "violationDate", "sortDirection": "desc"},
         "paged": True},
    ],
}


# ============================ http helpers ============================

def _session() -> requests.Session:
    s = requests.Session()
    s.headers.update(BASE_HEADERS)
    s.verify = VERIFY_SSL
    return s


class HostUnreachable(RuntimeError):
    """CFRS could not be resolved or connected to at all."""


# Substrings identifying a DNS/name-resolution failure across platforms.
# Retrying these is pointless: the host name did not resolve, and it will not
# resolve 5 seconds later either. Worth special-casing because the entity
# probe issues ~11 requests, and at MAX_RETRIES x RETRY_BACKOFF_S each that
# turns "you have no network" into three minutes of silent grinding before
# the first error surfaces.
_DNS_MARKERS = (
    "name or service not known",
    "temporary failure in name resolution",
    "nodename nor servname",
    "getaddrinfo failed",
    "failed to resolve",
    "name resolution",
)


def _is_dns_failure(exc: BaseException) -> bool:
    """True if exc (or anything in its cause chain) is a name-resolution error."""
    seen = set()
    while exc is not None and id(exc) not in seen:
        seen.add(id(exc))
        if isinstance(exc, socket.gaierror):
            return True
        if any(m in str(exc).lower() for m in _DNS_MARKERS):
            return True
        exc = exc.__cause__ or exc.__context__
    return False


def _request(session: requests.Session, method: str, url: str,
             json_body: dict | None = None, params: dict | None = None,
             stream: bool = False, timeout: int = REQUEST_TIMEOUT,
             log=None, label: str = ""):
    """HTTP with bounded retries on transient failures.

    Retries network errors and 429/5xx only. Two things are deliberately NOT
    retried:
      - a 404, which means the route is wrong; retrying it is pure noise, and
        the unverified endpoint probe requests routes that mostly don't exist.
      - a DNS failure, which is not transient in any useful sense and would
        otherwise multiply across every probed endpoint.
    """
    last_exc = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = session.request(method, url, json=json_body, params=params,
                                   stream=stream, timeout=timeout,
                                   verify=VERIFY_SSL)
            if resp.status_code in (429, 500, 502, 503, 504) and attempt < MAX_RETRIES:
                resp.close()
                delay = RETRY_BACKOFF_S * attempt
                if log:
                    # Silent retries are why a stalled first request looked
                    # like a hang: 3 attempts at a 120s timeout is 6 minutes
                    # with nothing on screen.
                    log.info(f"  {label or url}: HTTP {resp.status_code}, "
                             f"retry {attempt}/{MAX_RETRIES} in {delay}s")
                time.sleep(delay)
                continue
            return resp
        except requests.RequestException as e:
            if _is_dns_failure(e):
                raise HostUnreachable(
                    f"cannot resolve {ORIGIN} — no DNS for the CFRS host. "
                    f"Check network access; this is not a scraper bug."
                ) from e
            last_exc = e
            if attempt < MAX_RETRIES:
                delay = RETRY_BACKOFF_S * attempt
                if log:
                    log.info(f"  {label or url}: {type(e).__name__}, "
                             f"retry {attempt}/{MAX_RETRIES} in {delay}s")
                time.sleep(delay)
                continue
            raise
    raise last_exc  # pragma: no cover — the loop always returns or raises


def _preflight(session: requests.Session, log) -> None:
    """Fail fast and legibly if the CFRS host isn't reachable at all.

    Without this, an offline run spends minutes in retry backoff and then
    reports a pile of per-file download errors, which reads like the state
    changed its API rather than like a local network problem.
    """
    host = ORIGIN.split("://", 1)[-1]
    try:
        socket.getaddrinfo(host, 443)
    except socket.gaierror as e:
        msg = (f"cannot resolve {host} — no DNS for the CFRS host. "
               f"Check network access; this is not a scraper bug.")
        # The CLI's `except Exception: sys.exit(1)` (the house pattern across
        # every scraper) swallows the traceback, so without this the operator
        # sees a bare non-zero exit and nothing else.
        log.error(f"  {msg}")
        raise HostUnreachable(msg) from e


def _decode(raw: bytes) -> str:
    """Decode a CFRS body, honouring a UTF-16 or UTF-8 BOM if present.

    .NET services on several state disclosure sites emit UTF-16 for CSV
    exports (see docs/contributing.md's Encoding note). CFRS has been
    observed serving UTF-8, but the BOM check is cheap insurance against a
    silent mojibake regression.
    """
    if raw[:2] in (b"\xff\xfe", b"\xfe\xff"):
        return raw.decode("utf-16")
    if raw[:3] == b"\xef\xbb\xbf":
        return raw[3:].decode("utf-8", errors="replace")
    return raw.decode("utf-8", errors="replace")


# Envelope keys CFRS wraps result arrays in. The grids return a paged
# envelope; the small lookup routes return a bare array.
_ENVELOPE_KEYS = ("data", "Data", "items", "Items", "results", "Results",
                  "records", "Records", "list", "List", "rows", "Rows")

_TOTAL_KEYS = ("totalRecords", "TotalRecords", "totalCount", "TotalCount",
               "total", "Total", "recordCount", "RecordCount")


def _unwrap(payload, _depth: int = 0) -> list | None:
    """Pull the record array out of a JSON response, or None if there isn't one.

    Three passes, most specific first:
      1. A bare array — the small lookup routes return this.
      2. A known envelope key (_ENVELOPE_KEYS), recursing one level for
         shapes like {"data": {"items": [...]}}.
      3. Structural fallback: ANY value that is a list of dicts.

    Pass 3 exists because the envelope key names were inferred, not observed.
    Relying on a fixed key list means an unrecognized envelope silently reads
    as "zero records" — indistinguishable from a genuinely empty result, and
    the failure mode that made this scraper's first live run report no data
    for every year. Matching on shape instead of name degrades gracefully
    when CFRS renames a field.
    """
    if isinstance(payload, list):
        return payload
    if not isinstance(payload, dict) or _depth > 3:
        return None

    # Pass 2 — known envelope names.
    for key in _ENVELOPE_KEYS:
        val = payload.get(key)
        if isinstance(val, list):
            return val
        if isinstance(val, dict):
            inner = _unwrap(val, _depth + 1)
            if inner is not None:
                return inner

    # Pass 3 — structural. Prefer the longest list of dicts, so a payload
    # carrying both a records array and some small metadata array picks the
    # records. Ignores lists of scalars, which are never record sets.
    best: list | None = None
    for val in payload.values():
        if isinstance(val, list) and val and all(isinstance(x, dict) for x in val):
            if best is None or len(val) > len(best):
                best = val
    if best is not None:
        return best

    # Pass 3b — recurse into nested dicts under unknown key names.
    for val in payload.values():
        if isinstance(val, dict):
            inner = _unwrap(val, _depth + 1)
            if inner:
                return inner
    return None


def _shape_of(payload) -> str:
    """One-line description of an unrecognized response, for diagnostics.

    Printed when a grid's first page yields nothing. Without it, diagnosing a
    response-shape change means opening DevTools; with it, the scraper's own
    output says what came back.
    """
    if isinstance(payload, list):
        keys = sorted(payload[0].keys()) if payload and isinstance(payload[0], dict) else []
        return f"array[{len(payload)}] first-record keys={keys[:25]}"
    if isinstance(payload, dict):
        parts = []
        for k, v in list(payload.items())[:25]:
            if isinstance(v, list):
                inner = sorted(v[0].keys())[:25] if v and isinstance(v[0], dict) else []
                parts.append(f"{k}=array[{len(v)}]{' first-record keys=' + str(inner) if inner else ''}")
            elif isinstance(v, dict):
                parts.append(f"{k}=object(keys={sorted(v.keys())[:15]})")
            else:
                parts.append(f"{k}={type(v).__name__}({str(v)[:40]})")
        return "object{" + ", ".join(parts) + "}"
    return f"{type(payload).__name__}({str(payload)[:120]})"


def _total_of(payload) -> int | None:
    """Server-reported total row count, when the envelope advertises one.

    Two passes. _TOTAL_KEYS is an inferred list, so an exact-name match alone
    is fragile — `recordsTotal`, for instance, is a perfectly ordinary name
    that isn't in it. The second pass matches any integer field whose name
    looks like a count, which is what keeps a renamed field from silently
    reading as "no total" and disabling the progress bar, the ETA and the
    page-cap sizing all at once.
    """
    if not isinstance(payload, dict):
        return None

    for key in _TOTAL_KEYS:
        val = payload.get(key)
        if isinstance(val, int) and not isinstance(val, bool):
            return val
        if isinstance(val, dict):
            nested = _total_of(val)
            if nested is not None:
                return nested

    # Structural fallback: an int field whose name reads like a row count.
    # Deliberately excludes page-shaped fields (pageNumber, pageSize,
    # pageCount) — those are request echoes, not the result size.
    for k, v in payload.items():
        if not isinstance(v, int) or isinstance(v, bool):
            continue
        name = k.lower()
        if re.search(r"page", name):
            continue
        if re.search(r"(total|count|records|rows)", name):
            return v

    for v in payload.values():
        if isinstance(v, dict):
            nested = _total_of(v)
            if nested is not None:
                return nested
    return None


# tqdm is optional (already a project dependency, but the scraper must not
# hard-fail without it). docs/contributing.md recommends a progress bar for
# page sweeps, and this sweep can legitimately run for tens of minutes.
try:
    from tqdm import tqdm as _tqdm
except ImportError:
    _tqdm = None


def _seen(records: list, fetched: int) -> int:
    """Rows handled so far, whether retained in memory or streamed to a sink."""
    return fetched if fetched else len(records)


def _progress_bar(label: str, total: int):
    """A tqdm row-counter for one grid sweep, or None if tqdm is unavailable."""
    if _tqdm is None:
        return None
    return _tqdm(total=total, desc=f"  {label}", unit="row",
                 dynamic_ncols=True, leave=False)


def _close_bar(bar) -> None:
    if bar is not None:
        try:
            bar.close()
        except Exception:
            pass


def _fetch_grid(session: requests.Session, path: str, body: dict, log,
                label: str, page_size: int = GRID_PAGE_SIZE,
                _is_retry: bool = False, sink=None) -> tuple[list, bool]:
    """Page a POST grid endpoint to exhaustion.

    If `sink` is given it is called with each page's rows and NOTHING is kept
    in memory — the returned list is empty and only `complete` is meaningful.
    That matters at WV's real volumes: 2018 contributions are 34k rows / 57 MB,
    so 2022's 188k rows is ~320 MB of JSON, which accumulating in a list and
    then json.dumps-ing would balloon to multiple GB.

    Returns (records, complete). `complete` is False when paging stopped
    early — a page cap or a mid-sweep HTTP error — meaning the caller is
    holding a PARTIAL result and must not record it as a finished download.
    That distinction is the whole point of the tuple: an earlier version
    returned a bare list, so a truncated sweep was indistinguishable from a
    finished one and would have been written to the manifest as complete.

    Stops on: a short page, an empty page, reaching the server-reported
    total, or the page cap. The belt-and-braces conditions exist because a
    grid that silently ignored pageNumber would otherwise refetch page 1
    forever, which is worth being paranoid about when one run can issue
    thousands of requests.
    """
    url = f"{API_BASE}/{path.lstrip('/')}"
    records: list = []
    total: int | None = None
    complete = True
    max_pages = GRID_MAX_PAGES
    bar = None
    fetched = 0
    t_start = time.perf_counter()
    last_report = t_start

    page = 0
    while page < max_pages:
        page += 1
        payload = dict(body)
        payload.update({"pageNumber": page, "pageSize": page_size})

        if page == 1:
            # Heartbeat before the first request. Without it there is no
            # output at all between "Downloading..." and the first response,
            # which on a slow query is minutes of apparent deadlock.
            log.info(f"  {label}: requesting page 1 (pageSize={page_size})...")
        resp = _request(session, "POST", url, json_body=payload, timeout=120,
                        log=log, label=label)
        if resp.status_code != 200:
            if page == 1:
                raise requests.HTTPError(
                    f"{label}: HTTP {resp.status_code} from {url}")
            log.warning(f"  {label}: HTTP {resp.status_code} on page {page} — "
                        f"keeping {_seen(records, fetched):,} row(s) already fetched "
                        f"(PARTIAL)")
            complete = False
            break

        parsed = json.loads(_decode(resp.content))
        batch  = _unwrap(parsed)
        if page == 1:
            total = _total_of(parsed)
            if total:
                # Size the cap to the work actually required rather than a
                # blanket constant. At the fallback pageSize of 10 a real WV
                # year (188k contribution rows in 2022) needs ~18,800 pages,
                # far past GRID_MAX_PAGES — which would have silently dropped
                # three quarters of the data.
                needed = math.ceil(total / page_size) + 2
                max_pages = max(GRID_MAX_PAGES, needed)
                eta = needed * (RATE_LIMIT_S + 0.5)
                log.info(f"  {label}: {total:,} row(s) server-side, "
                         f"~{needed:,} page(s) at pageSize={page_size}"
                         + (f" — roughly {eta/60:.0f} min" if eta > 90 else ""))
                bar = _progress_bar(label, total)
            else:
                # No recognized total key. _TOTAL_KEYS is an inferred list, so
                # this is a likely-wrong-guess situation rather than a server
                # that omits the count — print the envelope's actual keys so
                # the right one can be added.
                #
                # This branch previously did not exist: ALL page-1 output sat
                # behind `if total:`, so an unrecognized envelope produced no
                # ETA line, no progress bar, and nothing at all until page 25.
                # On a year needing fewer than 25 pages that meant total
                # silence from "Downloading..." to completion, which is
                # exactly how a working scraper came to look like a hang.
                keys = sorted(parsed.keys()) if isinstance(parsed, dict) else "(array)"
                log.info(f"  {label}: page 1 returned {len(batch):,} row(s) in "
                         f"{time.perf_counter() - t_start:.1f}s; no row-count "
                         f"field recognized — envelope keys: {keys}")

        if not batch:
            if page == 1:
                # An empty first page is ambiguous: the query legitimately has
                # no rows, the response shape changed, or the server rejected
                # our pageSize. Try the one page size we have actually observed
                # CFRS serve (10, from the site's own XHR) before concluding
                # anything — some .NET grids validate pageSize against an
                # allow-list and return an empty set rather than an error.
                if page_size != OBSERVED_PAGE_SIZE and not _is_retry:
                    log.info(f"  {label}: empty at pageSize={page_size}, "
                             f"retrying at the site's own pageSize="
                             f"{OBSERVED_PAGE_SIZE}")
                    retry, retry_complete = _fetch_grid(
                        session, path, body, log, label,
                        page_size=OBSERVED_PAGE_SIZE, _is_retry=True)
                    if retry:
                        log.warning(f"  {label}: pageSize={page_size} is "
                                    f"rejected by CFRS — lower GRID_PAGE_SIZE "
                                    f"to {OBSERVED_PAGE_SIZE} to avoid this "
                                    f"retry on every call")
                        _close_bar(bar)
                        return retry, retry_complete
                # Genuinely nothing, at any page size. Print what came back so
                # a shape change is diagnosable from the scraper's own output
                # rather than needing DevTools. Suppressed on the retry pass so
                # an empty year logs this once, not twice.
                if not _is_retry:
                    log.warning(f"  {label}: no records returned — "
                                f"response shape: {_shape_of(parsed)}")
            break

        if sink is not None:
            sink(batch)          # streamed straight to disk, not retained
            fetched += len(batch)
        else:
            records.extend(batch)
        if bar is not None:
            bar.update(len(batch))
        elif time.perf_counter() - last_report >= PROGRESS_EVERY_S:
            # Time-based, not page-count-based. A page-count trigger says
            # nothing for the first N pages however long they take, and this
            # loop's whole problem is that slow silence is indistinguishable
            # from a hang.
            last_report = time.perf_counter()
            log.info(f"  {label}: page {page:,}, {_seen(records, fetched):,} row(s) "
                     f"in {last_report - t_start:.0f}s")

        if len(batch) < page_size:
            break
        if total is not None and _seen(records, fetched) >= total:
            break

        time.sleep(RATE_LIMIT_S)
    else:
        # Loop exhausted the cap without a natural stop.
        complete = False

    _close_bar(bar)

    if not complete or (total is not None and _seen(records, fetched) < total):
        complete = False
        expected = f"{total:,}" if total is not None else "an unknown number of"
        log.warning(f"  {label}: INCOMPLETE — got {_seen(records, fetched):,} of "
                    f"{expected} row(s) after {page:,} page(s). Not recording "
                    f"as a finished download; re-run to retry.")

    return records, complete



# ==================== transactions: paged JSON grid ===================

def download_grid_year(session: requests.Session, ttype: str, year: str,
                       log) -> tuple[bool, int, int]:
    """Fetch one (category, year) slice of the transaction grid.

    Returns (ok, bytes_written, row_count).
    """
    body = dict(TRANSACTION_GRID_BODY)
    body.update({"transactionCategory": ttype, "transactionYear": str(year)})

    label = f"transactions_{ttype}_{year}"

    # Streamed to newline-delimited JSON via a .part file. Two reasons:
    #   - memory: a full year held in a list and json.dumps-ed at the end is
    #     several GB at WV's real volumes (2018 CON alone is 34k rows / 50 MB).
    #   - the .part rename keeps the "never leave a partial year on disk"
    #     guarantee intact, which matters because run()'s existence
    #     check would otherwise treat a truncated file as a finished download
    #     and skip it forever.
    dest = RAW_DIR / f"{label}.jsonl"
    tmp  = dest.with_suffix(".jsonl.part")
    rows = 0

    try:
        with open(tmp, "w", encoding="utf-8") as fh:
            def _sink(batch):
                nonlocal rows
                for rec in batch:
                    fh.write(json.dumps(rec, separators=(",", ":")) + "\n")
                    rows += 1

            _, complete = _fetch_grid(session, TRANSACTION_GRID_PATH, body, log,
                                      label, sink=_sink)
    except Exception as e:
        tmp.unlink(missing_ok=True)
        log.file_download_error(filename=f"{label}.jsonl", error=str(e))
        return False, 0, 0

    if not rows:
        # No rows for that year is a legitimate answer, not a failure — don't
        # write an empty file the parser would have to special-case. Emit a
        # distinct event: this used to reuse file_download_skip, whose console
        # text reads "already downloaded, skipping", which actively misleads
        # (it made a run where every year returned zero rows look like a
        # fully-cached no-op). _fetch_grid has already logged the response
        # shape if this was a shape problem rather than an empty year.
        tmp.unlink(missing_ok=True)
        log.info(f"  – {label}.jsonl: API returned no rows for this year")
        log._emit("file_download", status="empty", filename=f"{label}.jsonl",
                  year=year, transaction_type=ttype)
        return False, 0, 0

    if not complete:
        # Deliberately do NOT keep a partial year: the .part file is discarded
        # rather than renamed into place.
        #
        # Keeping it would be worse than useless — it wouldn't be in the
        # manifest, but run()'s existence fallback treats any file on disk as
        # done, so a re-run would skip it forever and the truncation would
        # quietly become permanent.
        tmp.unlink(missing_ok=True)
        log.file_download_error(
            filename=f"{label}.jsonl",
            error=f"incomplete sweep ({rows:,} rows) — discarded, "
                  f"will retry on next run")
        log._emit("file_download", status="incomplete", filename=f"{label}.jsonl",
                  year=year, transaction_type=ttype, rows_fetched=rows)
        return False, 0, 0

    tmp.replace(dest)
    return True, dest.stat().st_size, rows


# ============================== entities ==============================

def fetch_registry(session: requests.Session, log) -> tuple[bool, int]:
    """Download the full candidate + committee registry.

    This is the only source for office, district, party and registration
    status — none of which appear anywhere in the transaction feed — so a
    failure here degrades entity quality but must not abort the run.
    """
    try:
        records, complete = _fetch_grid(session, REGISTRY_PATH, REGISTRY_BODY,
                                        log, "entities_committees")
    except Exception as e:
        log.file_download_error(filename="entities_committees.json", error=str(e))
        return False, 0

    if not records:
        log.warning("  registry returned no records — parser will fall back to "
                    "transaction-derived entities")
        return False, 0

    if not complete:
        # Unlike a transaction year, a partial registry IS worth keeping: it's
        # enrichment, not a primary relation, and committees it does cover get
        # correct office/district/party. Warn loudly so the gap is visible.
        log.warning(f"  registry is PARTIAL ({len(records):,} rows) — some "
                    f"committees will fall back to transaction-derived "
                    f"entities with no office/district/party")

    dest = RAW_DIR / "entities_committees.json"
    dest.write_text(json.dumps(records, indent=1), encoding="utf-8")
    log.file_download_ok(filename=dest.name, bytes=dest.stat().st_size,
                         rows=len(records), duration_s=0.0)
    return True, len(records)


def fetch_contributor_types(session: requests.Session, log) -> int:
    """Download the contributor/payee type vocabulary for CON and EXP.

    Saved so src/aliases/contributor_types.csv can be filled from the
    source's own controlled list rather than from whatever values happen to
    appear in the data.
    """
    ok = 0
    for ttype in sorted(TRANSACTION_TYPES):
        url = f"{API_BASE}/{CONTRIBUTOR_TYPE_PATH}/{ttype}"
        try:
            time.sleep(RATE_LIMIT_S)
            resp = _request(session, "GET", url, timeout=60, log=log,
                            label=f"contributor_types/{ttype}")
            if resp.status_code != 200:
                continue
            records = _unwrap(json.loads(_decode(resp.content)))
            if not records:
                continue
        except (requests.RequestException, json.JSONDecodeError, ValueError) as e:
            log.warning(f"  contributor types {ttype}: {e}")
            continue

        dest = RAW_DIR / f"lookup_contributor_types_{ttype}.json"
        dest.write_text(json.dumps(records, indent=1), encoding="utf-8")
        log.file_download_ok(filename=dest.name, bytes=dest.stat().st_size,
                             rows=len(records), duration_s=0.0)
        ok += 1
    return ok


# ========================== whole-grid export =========================

# Keys the generate step might return the file handle under. Ordered most
# specific first. Unknown shapes fall through to _shape_of() diagnostics
# rather than a silent miss.
_FILE_REF_KEYS = ("filePath", "fileName", "fileKey", "key", "url", "fileUrl",
                  "downloadUrl", "path", "name", "data", "result")

_BINARY_CONTENT_TYPES = ("spreadsheet", "excel", "octet-stream", "text/csv",
                         "application/zip")


def _extract_file_bytes(payload, _depth: int = 0):
    """Pull inline file content out of an export response, or None.

    CFRS's observed shape is:

        {"isSuccess": true,
         "responseData": {"fileBytes": "<base64>"},
         "message": null, "skipRecords": null}

    i.e. the spreadsheet comes back embedded in the JSON — there is no file
    handle and no CloudFront round-trip, which is what the handle-searching
    path was originally built for. Accepts base64 text or a raw byte array,
    since .NET serializes byte[] either way depending on configuration.
    """
    if not isinstance(payload, dict) or _depth > 4:
        return None
    for k, v in payload.items():
        if _norm(k) in ("filebytes", "filecontent", "filedata", "content", "bytes"):
            if isinstance(v, str) and v.strip():
                try:
                    return base64.b64decode(v, validate=False)
                except (binascii.Error, ValueError):
                    return None
            if isinstance(v, list) and v and all(isinstance(x, int) for x in v):
                try:
                    return bytes(v)
                except ValueError:
                    return None
        if isinstance(v, dict):
            inner = _extract_file_bytes(v, _depth + 1)
            if inner:
                return inner
    return None


def _norm(k: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (k or "").lower())


# CFRS only emits Excel — there is no CSV route and no format parameter
# anywhere in the app bundle (checked: only generateExportGridDataExcel,
# /Common/generateExportGridData and /ExportGridData/... exist, none taking a
# fileType/exportFormat). So the conversion happens here, at write time.
#
# This is the one place the scraper does NOT keep a raw file byte-for-byte as
# served. The tradeoff is deliberate: a single-sheet grid export converts to
# CSV losslessly, CSV is diffable and greppable where .xlsx is an opaque zip,
# and the rest of this pipeline is CSV-native. Set EXPORT_AS_CSV = False to
# keep the original .xlsx instead.
EXPORT_AS_CSV = True


def _xlsx_to_csv(blob: bytes) -> tuple[bytes, int] | None:
    """Convert a single-sheet xlsx export to UTF-8 CSV.

    Returns (csv_bytes, data_row_count) — the count excludes the header — or
    None if the payload isn't a readable workbook. Counting here is free: we
    are already walking every row, and it is the only place the true row
    count is known, since the API reports nothing about the export's size.

    Streams with read_only/values_only so a large export doesn't build a full
    cell object graph. Dates are emitted ISO-8601 (the form CFRS's own JSON
    grid uses, and one parse_date already accepts) rather than Excel's
    locale-dependent rendering.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    import datetime as _dt

    def _cell(v):
        if v is None:
            return ""
        if isinstance(v, (_dt.datetime, _dt.date)):
            return v.isoformat()
        if isinstance(v, bool):
            return "True" if v else "False"
        return str(v)

    buf = io.StringIO(newline="")
    writer = csv.writer(buf, lineterminator="\n")
    try:
        wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    except Exception:
        return None
    try:
        ws = wb[wb.sheetnames[0]]
        n = 0
        for row in ws.iter_rows(values_only=True):
            if row is None:
                continue
            writer.writerow([_cell(v) for v in row])
            n += 1
        if not n:
            return None
        rows = max(n - 1, 0)          # drop the header
    except Exception:
        return None
    finally:
        wb.close()

    return buf.getvalue().encode("utf-8"), rows


def _write_export(stem: str, blob: bytes, log, relation: str) -> tuple[bool, int, int]:
    """Persist an export blob, converting xlsx -> csv when EXPORT_AS_CSV.

    Falls back to writing the original bytes if conversion isn't possible
    (openpyxl missing, or the payload wasn't a readable workbook) — better a
    usable .xlsx than a failed download.
    """
    ext = _guess_ext(blob)
    if EXPORT_AS_CSV and ext in (".xlsx", ".xls"):
        converted = _xlsx_to_csv(blob)
        if converted is not None:
            csv_bytes, rows = converted
            dest = RAW_DIR / f"{stem}.csv"
            dest.write_bytes(csv_bytes)
            log.info(f"  export/{relation}: {len(blob)/1024:.0f} KB {ext} -> "
                     f"{len(csv_bytes)/1024:.0f} KB csv, {rows:,} rows ({dest.name})")
            return True, dest.stat().st_size, rows
        log.warning(f"  export/{relation}: could not convert {ext} to CSV — "
                    f"keeping the original")
    # Not converted (kept as-is): the row count isn't known without opening
    # the workbook, and -1 distinguishes "unknown" from a genuine zero.
    dest = RAW_DIR / f"{stem}{ext}"
    dest.write_bytes(blob)
    return True, dest.stat().st_size, -1


def _guess_ext(blob: bytes) -> str:
    """File extension from magic bytes — xlsx is a zip, CSV is plain text."""
    if blob[:4] == b"PK\x03\x04":
        return ".xlsx"
    if blob[:2] == b"\xd0\xcf":          # legacy OLE2 .xls
        return ".xls"
    return ".csv"


def _extract_file_ref(payload) -> str | None:
    """Pull a file handle (S3 key, filename or URL) out of the generate step.

    The exact response shape of generateExportGridDataExcel has not been
    observed, so this searches rather than assumes: a bare string, then a
    known key, then any string value that looks like a filename or URL.
    """
    if isinstance(payload, str):
        return payload.strip() or None
    if not isinstance(payload, dict):
        return None
    for key in _FILE_REF_KEYS:
        for k, v in payload.items():
            if k.lower() == key.lower():
                if isinstance(v, str) and v.strip():
                    return v.strip()
                if isinstance(v, dict):
                    inner = _extract_file_ref(v)
                    if inner:
                        return inner
    for v in payload.values():
        if isinstance(v, str) and re.search(r"(\.(xlsx?|csv|zip)$|^https?://)",
                                            v.strip(), re.I):
            return v.strip()
    return None


def export_grid(session: requests.Session, relation: str, spec: dict,
                log, stem: str | None = None) -> tuple[bool, int, int]:
    """Whole-grid export: generate a file server-side, then download it.

    Returns (ok, bytes_written, rows) — rows is 0 because the payload is a
    spreadsheet this function does not parse.

    Deliberately defensive: neither step's response shape has been observed,
    so every branch that can't proceed logs what it actually received via
    _shape_of() instead of failing silently.
    """
    # pageName / fieldType are per-grid, not universal: the registry export
    # sends them, the TrackFinance exports do not. Sending keys a grid didn't
    # ask for is exactly the kind of thing this API rejects, so only include
    # what the spec declares.
    body = {
        "moduleType":    "PUBLIC",     # what makes this work without a token
        "gridName":      spec["gridName"],
        "filterRequest": spec["filterRequest"],
    }
    if spec.get("pageName"):
        body["pageName"] = spec["pageName"]
    if spec.get("fieldType"):
        body["fieldType"] = spec["fieldType"]

    gen_url = f"{COMMON_BASE}/{EXPORT_PATH}"

    try:
        resp = _request(session, "POST", gen_url, json_body=body, timeout=300,
                        log=log, label=f"export/{relation}")
    except requests.RequestException as e:
        log.file_download_error(filename=f"export_{relation}", error=str(e))
        return False, 0, 0

    if resp.status_code != 200:
        log.file_download_error(
            filename=f"export_{relation}",
            error=f"generate step returned HTTP {resp.status_code}")
        return False, 0, 0

    ctype = resp.headers.get("Content-Type", "").lower()

    # Case 1 — the generate step streamed the file itself.
    if any(t in ctype for t in _BINARY_CONTENT_TYPES):
        log.info(f"  export/{relation}: served inline ({ctype})")
        ok, nbytes, nrows = _write_export(stem or f"export_{relation}",
                                          resp.content, log, relation)
        return ok, nbytes, nrows

    # Case 2 — a JSON handle we must exchange for a download link.
    try:
        parsed = json.loads(_decode(resp.content))
    except (json.JSONDecodeError, ValueError):
        log.file_download_error(
            filename=f"export_{relation}",
            error=f"generate step returned neither a file nor JSON "
                  f"[{ctype}]: {_decode(resp.content)[:160]!r}")
        return False, 0, 0

    # Case 2a — the file is embedded in the JSON. This is what CFRS actually
    # does (responseData.fileBytes); the handle/CloudFront path below is kept
    # only because the bundle's admin routes clearly support it too.
    blob = _extract_file_bytes(parsed)
    if blob:
        ok, nbytes, nrows = _write_export(stem or f"export_{relation}", blob, log, relation)
        return ok, nbytes, nrows

    # Case 2b — a handle to exchange for a download link.
    ref = _extract_file_ref(parsed)
    if not ref:
        ok_flag = parsed.get("isSuccess") if isinstance(parsed, dict) else None
        msg     = parsed.get("message") if isinstance(parsed, dict) else None
        log.file_download_error(
            filename=f"export_{relation}",
            error=f"no file bytes or handle in generate response "
                  f"(isSuccess={ok_flag!r}, message={msg!r}) — "
                  f"shape: {_shape_of(parsed)}")
        return False, 0, 0

    log.info(f"  export/{relation}: generated {ref[:80]}")

    link = ref if ref.lower().startswith("http") else None
    if link is None:
        link_url = f"{COMMON_BASE}/{DOWNLOAD_LINK_PATH}"
        for attempt in ({"fileName": ref}, {"filePath": ref}, {"key": ref}):
            try:
                lresp = _request(session, "POST", link_url,
                                 json_body=attempt, timeout=120)
                if lresp.status_code != 200:
                    continue
                lparsed = json.loads(_decode(lresp.content))
            except (requests.RequestException, json.JSONDecodeError, ValueError):
                continue
            cand = _extract_file_ref(lparsed)
            if cand and cand.lower().startswith("http"):
                link = cand
                break
        if link is None:
            log.file_download_error(
                filename=f"export_{relation}",
                error=f"could not exchange handle {ref[:60]!r} for a download "
                      f"link via {DOWNLOAD_LINK_PATH}")
            return False, 0, 0

    # CloudFront is a different origin, so this is a plain GET without the
    # API headers.
    try:
        fresp = requests.get(link, timeout=REQUEST_TIMEOUT, verify=VERIFY_SSL,
                             headers={"User-Agent": USER_AGENT})
        fresp.raise_for_status()
    except requests.RequestException as e:
        log.file_download_error(filename=f"export_{relation}",
                                error=f"download link fetch failed: {e}")
        return False, 0, 0

    ok, nbytes, nrows = _write_export(stem or f"export_{relation}",
                                      fresp.content, log, relation)
    return ok, nbytes, nrows


def export_transactions_year(session: requests.Session, ttype: str, year: str,
                             log) -> tuple[bool, int, int]:
    """Export one (category, year) transaction grid in a single call.

    Returns (ok, bytes_written, rows). The whole point of this path is cost: the
    equivalent paged sweep is hundreds of requests and, measured end to end
    across 2018–2026, 44 minutes.
    """
    spec = GRID_EXPORTS.get(f"transactions_{ttype}")
    if not spec:
        return False, 0, 0

    # Deep-copy the filter and substitute the year — the template lives in
    # GRID_EXPORTS and must not be mutated across calls.
    filt = {k: (v.format(year=year) if isinstance(v, str) and "{year}" in v else v)
            for k, v in spec["filterRequest"].items()}
    call_spec = {**spec, "filterRequest": filt}

    return export_grid(session, f"transactions_{ttype}", call_spec, log,
                       stem=f"export_{ttype}_{year}")


# =========================== lookup probing ===========================

def _try_lookup(session: requests.Session, spec: dict, base: str,
                log) -> list | None:
    """Fetch one lookup spec against one service base. Returns records or None.

    Never raises — probing a route that may not exist on this base is
    inherently best-effort and must not be able to fail a run.
    """
    url    = f"{base}/{spec['path'].lstrip('/')}"
    method = spec.get("method", "GET")
    try:
        if method == "POST" and spec.get("paged"):
            if base != API_BASE:
                return None      # paged grids are a Public-Service concept
            recs, _ = _fetch_grid(session, spec["path"], spec.get("body") or {},
                                  log, spec["path"])
            return recs or None
        resp = _request(session, method, url,
                        json_body=spec.get("body") if method == "POST" else None,
                        timeout=60)
        if resp.status_code != 200:
            return None
        return _unwrap(json.loads(_decode(resp.content))) or None
    except (requests.RequestException, json.JSONDecodeError, ValueError):
        return None


def fetch_lookups(session: requests.Session, log,
                  cache: dict) -> tuple[int, int]:
    """Fetch the reference vocabularies and remaining entity tabs.

    Each relation's specs are tried against each service base until one
    returns records. A successful (path, base) pair is cached so steady-state
    runs make one request per relation instead of re-walking the matrix.

    Returns (ok, miss). Misses are expected and harmless — these feed the
    alias CSVs and office_types.csv, not the pipeline's output contract.
    """
    resolved = cache.setdefault("lookups", {})
    ok = miss = 0

    for relation, specs in LOOKUP_ENDPOINTS.items():
        prev = resolved.get(relation)

        attempts: list[tuple[dict, str]] = []
        if prev:
            attempts.append(({k: v for k, v in prev.items() if k != "base"},
                             prev.get("base", API_BASE)))
        for spec in specs:
            for base in LOOKUP_BASES:
                if prev and spec["path"] == prev.get("path") and base == prev.get("base"):
                    continue     # already queued first
                attempts.append((spec, base))

        records = used = used_base = None
        for spec, base in attempts:
            time.sleep(RATE_LIMIT_S)
            records = _try_lookup(session, spec, base, log)
            if records:
                used, used_base = spec, base
                break

        if not records or not used:
            miss += 1
            resolved.pop(relation, None)
            log.info(f"  – lookups/{relation}: no route responded (enrichment only)")
            continue

        dest = RAW_DIR / f"lookup_{relation}.json"
        dest.write_text(json.dumps(records, indent=1), encoding="utf-8")
        resolved[relation] = {**used, "base": used_base}
        ok += 1
        service = used_base.rsplit("/", 1)[-1]
        log.file_download_ok(filename=dest.name, bytes=dest.stat().st_size,
                             rows=len(records), duration_s=0.0)
        log.info(f"    ({service}/{used['path']})")
    return ok, miss


# ========================= manifest helpers ===========================

def _key(relation_type: str, year: str, filename: str) -> str:
    return f"{relation_type}|{year}|{filename}"


def load_manifest() -> dict[str, dict]:
    if not MANIFEST.exists():
        return {}
    with open(MANIFEST, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None or "relation_type" not in reader.fieldnames:
            print(f"WARNING: {MANIFEST} exists but doesn't look like a "
                  f"west_virginia.py manifest (missing 'relation_type' column) "
                  f"— ignoring it and starting fresh.")
            return {}
        return {_key(r["relation_type"], r["year"], r["filename"]): r
                for r in reader}


def write_manifest(rows: dict[str, dict]) -> None:
    with open(MANIFEST, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=MANIFEST_COLS, extrasaction="ignore")
        w.writeheader()
        for row in sorted(rows.values(),
                          key=lambda r: (r["relation_type"], r["year"], r["filename"])):
            w.writerow(row)


def strip_manifest(manifest: dict[str, dict], predicate) -> dict[str, dict]:
    """Drop manifest entries where predicate(row) is True — used to force a
    re-download for --force / --start-year / --end-year (see
    docs/contributing.md, "Manifest clearing for year flags")."""
    return {k: v for k, v in manifest.items() if not predicate(v)}


def _load_cache() -> dict:
    if not ENDPOINTS_CACHE.exists():
        return {}
    try:
        return json.loads(ENDPOINTS_CACHE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def _save_cache(cache: dict) -> None:
    ENDPOINTS_CACHE.write_text(json.dumps(cache, indent=2), encoding="utf-8")


# ================================ run =================================

def run(
    force: bool = False,
    entities: bool = False,
    transactions: bool = False,
    start_year: int | None = None,
    end_year: int | None = None,
    contributions: bool = False,
    expenditures: bool = False,
    candidates: bool = False,
    committees: bool = False,
):
    """Download WV CFRS transactions and entity feeds.

    Horizontal scope resolves to a set of transaction type codes plus a
    boolean for the entity sweep. CFRS serves candidates and committees from
    a single registry route, so --candidates and --committees both enable
    the same sweep.
    """
    log = get_logger("west_virginia", "scrape")
    t0  = time.perf_counter()
    log._emit("scrape_started", force=force, entities=entities,
              transactions=transactions, start_year=start_year, end_year=end_year,
              contributions=contributions, expenditures=expenditures,
              candidates=candidates, committees=committees)

    # ---- resolve horizontal scope ----
    any_horizontal = any([transactions, entities, contributions,
                          expenditures, candidates, committees])
    if not any_horizontal:
        want_types    = set(TRANSACTION_TYPES)
        want_entities = True
    else:
        want_types = set()
        if transactions:
            want_types |= set(TRANSACTION_TYPES)
        if contributions:
            want_types.add("CON")
        if expenditures:
            want_types.add("EXP")
        want_entities = entities or candidates or committees

    # ---- resolve vertical scope ----
    current_year        = datetime.today().year
    year_range_explicit = start_year is not None or end_year is not None

    manifest = load_manifest()
    cache    = _load_cache()

    if force or year_range_explicit:
        lo = start_year if start_year is not None else EARLIEST_YEAR
        hi = end_year   if end_year   is not None else current_year

        def _in_scope(row: dict) -> bool:
            try:
                yr = int(row["year"])
            except (ValueError, KeyError, TypeError):
                return False   # entity rows carry no year and are never wiped by year flags
            return lo <= yr <= hi

        manifest = strip_manifest(manifest, _in_scope)

    files_ok = files_err = files_skipped = 0
    entities_ok = entities_miss = 0
    session = _session()

    try:
        # Cheap up-front reachability check so an offline run says so once,
        # immediately, instead of surfacing as a slow pile of download errors.
        _preflight(session, log)

        # ================== transactions ==================
        if want_types:
            lo = start_year if start_year is not None else EARLIEST_YEAR
            hi = end_year   if end_year   is not None else current_year

            for year in range(lo, hi + 1):
                for ttype in sorted(want_types):
                    relation = TRANSACTION_TYPES[ttype]
                    filename = f"transactions_{ttype}_{year}.jsonl"
                    key  = _key(relation, str(year), filename)
                    dest = RAW_DIR / filename

                    # A .json left by an earlier scraper build still counts as
                    # having this year — the parser reads both formats — so
                    # changing the on-disk format doesn't silently force a full
                    # re-download of everything already fetched.
                    legacy  = RAW_DIR / f"transactions_{ttype}_{year}.json"
                    # An export file satisfies the year too — otherwise every
                    # run would re-export years already on disk.
                    candidates_on_disk = [dest, legacy] + sorted(
                        RAW_DIR.glob(f"export_{ttype}_{year}.*"))
                    on_disk = next((f for f in candidates_on_disk
                                    if f.exists() and f.stat().st_size > 0), None)

                    # The current calendar year is always refetched: CFRS
                    # amends the open year in place as reports land, so a
                    # manifest hit from earlier in the year is stale by
                    # definition.
                    is_current   = year == current_year
                    already_done = key in manifest or (
                        not year_range_explicit and not force
                        and on_disk is not None
                    )
                    if already_done and not is_current and not force:
                        log.file_download_skip(filename=filename)
                        files_skipped += 1
                        continue

                    t_file = time.perf_counter()

                    # Whole-grid export first — one request instead of
                    # hundreds. Any failure (wrong gridName, isSuccess=false,
                    # unreadable body) falls straight through to the paged
                    # sweep, so this can only make a run faster, never fail it.
                    exported = False
                    if EXPORT_TRANSACTIONS_READY:
                        try:
                            exported, xbytes, xrows = export_transactions_year(
                                session, ttype, str(year), log)
                        except Exception as e:
                            log.warning(f"  export {ttype} {year} failed "
                                        f"({e}) — falling back to paging")
                            exported = False

                    if exported:
                        files_ok += 1
                        xname = next((f.name for f in
                                      RAW_DIR.glob(f"export_{ttype}_{year}.*")),
                                     f"export_{ttype}_{year}")
                        log.file_download_ok(
                            filename=xname, bytes=xbytes,
                            rows=max(xrows, 0),
                            duration_s=time.perf_counter() - t_file)
                        manifest[_key(relation, str(year), xname)] = {
                            "relation_type": relation, "year": str(year),
                            "filename": xname,
                            "source_url": f"{COMMON_BASE}/{EXPORT_PATH}",
                            "bytes": xbytes,
                            # -1 means "written but not counted" (kept as xlsx);
                            # blank is more honest there than a fake 0.
                            "rows": xrows if xrows >= 0 else "",
                            "scraped_at": datetime.now().isoformat(timespec="seconds"),
                        }
                        write_manifest(manifest)
                        continue

                    log.file_download_start(filename=filename)
                    ok, nbytes, nrows = download_grid_year(
                        session, ttype, str(year), log)
                    if not ok:
                        files_err += 1
                        continue

                    files_ok += 1
                    log.file_download_ok(filename=filename, bytes=nbytes,
                                         rows=nrows,
                                         duration_s=time.perf_counter() - t_file)
                    manifest[key] = {
                        "relation_type": relation, "year": str(year),
                        "filename": filename,
                        "source_url": f"{API_BASE}/{TRANSACTION_GRID_PATH}",
                        "bytes": nbytes, "rows": nrows,
                        "scraped_at": datetime.now().isoformat(timespec="seconds"),
                    }
                    write_manifest(manifest)   # per-file rewrite — keeps re-runs resumable

        # ==================== entities ====================
        if want_entities:
            reg_ok, reg_rows = fetch_registry(session, log)
            if reg_ok:
                entities_ok += 1
            else:
                entities_miss += 1

            entities_ok += fetch_contributor_types(session, log)

            # Whole-grid export. Confirmed working for the candidate/committee
            # registry; transaction grids need their gridName/pageName captured
            # (see GRID_EXPORTS). Best-effort — the paged registry above has
            # already run, so a failure here costs nothing.
            for relation in ENTITY_EXPORTS:
                spec = GRID_EXPORTS.get(relation)
                if not spec:
                    continue
                try:
                    e_ok, e_bytes, e_rows = export_grid(session, relation, spec, log)
                except Exception as e:
                    log.warning(f"  export/{relation} failed: {e}")
                    continue
                if e_ok:
                    entities_ok += 1
                    ename = next((f.name for f in
                                  RAW_DIR.glob(f"export_{relation}.*")),
                                 f"export_{relation}")
                    log.file_download_ok(filename=ename, bytes=e_bytes,
                                         rows=max(e_rows, 0), duration_s=0.0)

            u_ok, u_miss = fetch_lookups(session, log, cache)
            entities_ok   += u_ok
            entities_miss += u_miss
            _save_cache(cache)

            resolved_lookups = cache.get("lookups", {})
            for path in sorted(RAW_DIR.glob("entities_*.json")) + \
                        sorted(RAW_DIR.glob("lookup_*.json")):
                relation = path.stem.replace("entities_", "").replace("lookup_", "")
                # Record the base the route actually resolved on, not a blanket
                # API_BASE — a lookup served by Common-Service should say so,
                # otherwise the manifest misattributes where the data came from.
                spec = resolved_lookups.get(relation) or {}
                source = (f"{spec['base']}/{spec['path']}"
                          if spec.get("base") and spec.get("path") else API_BASE)
                manifest[_key(relation, "", path.name)] = {
                    "relation_type": relation, "year": "",
                    "filename": path.name, "source_url": source,
                    "bytes": path.stat().st_size, "rows": "",
                    "scraped_at": datetime.now().isoformat(timespec="seconds"),
                }
            write_manifest(manifest)

        duration = round(time.perf_counter() - t0, 1)
        log.info(f"Done in {duration}s — {files_ok} downloaded, "
                 f"{files_skipped} skipped, {files_err} failed, "
                 f"{entities_ok} entity feed(s) captured. Files in {RAW_DIR}/")
        log._emit("scrape_completed", status="completed", duration_s=duration,
                  files_ok=files_ok, files_err=files_err,
                  files_skipped=files_skipped,
                  entities_ok=entities_ok, entities_miss=entities_miss)

    except KeyboardInterrupt:
        write_manifest(manifest)
        _save_cache(cache)
        log._emit("scrape_completed", status="interrupted",
                  duration_s=round(time.perf_counter() - t0, 1),
                  files_ok=files_ok, files_err=files_err,
                  files_skipped=files_skipped)
        raise

    except Exception as e:
        write_manifest(manifest)
        _save_cache(cache)
        log._emit("scrape_completed", status="error",
                  duration_s=round(time.perf_counter() - t0, 1),
                  files_ok=files_ok, files_err=files_err,
                  files_skipped=files_skipped,
                  error_type=type(e).__name__, error=str(e))
        raise


# ================================ CLI =================================

if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(
        description="Download West Virginia CFRS campaign finance data."
    )

    vert = ap.add_mutually_exclusive_group()
    vert.add_argument("--force", action="store_true",
                      help="re-download all years in scope, wipe manifest")
    vert.add_argument("--start-year", type=int, metavar="YYYY",
                      help="earliest year to download (inclusive); wipes manifest for range")
    ap.add_argument("--end-year", type=int, metavar="YYYY",
                    help="latest year to download (inclusive, <= current year)")

    ap.add_argument("--transactions",  action="store_true", help="contributions + expenditures")
    ap.add_argument("--entities",      action="store_true", help="registry + lookups only")
    ap.add_argument("--contributions", action="store_true", help="CON only")
    ap.add_argument("--expenditures",  action="store_true", help="EXP only")
    ap.add_argument("--candidates",    action="store_true", help="entity sweep (see docstring)")
    ap.add_argument("--committees",    action="store_true", help="entity sweep (see docstring)")

    # parse_known_args — orc.py forwards the full flag set to every scraper
    args, _ = ap.parse_known_args()

    cy = datetime.today().year
    if args.end_year:
        if args.end_year > cy:
            ap.error(f"--end-year cannot exceed current year ({cy})")
        if args.start_year and args.start_year > args.end_year:
            ap.error("--start-year cannot be greater than --end-year")
    if args.force and args.end_year:
        ap.error("--force cannot be combined with --end-year")

    try:
        run(
            force=args.force,
            entities=args.entities,
            transactions=args.transactions,
            start_year=args.start_year,
            end_year=args.end_year,
            contributions=args.contributions,
            expenditures=args.expenditures,
            candidates=args.candidates,
            committees=args.committees,
        )
    except KeyboardInterrupt:
        sys.exit(130)
    except Exception:
        sys.exit(1)
