"""
parsers/west_virginia.py — Parse West Virginia CFRS campaign finance data.

Reads whatever scrapers/west_virginia.py left in data/West_Virginia/raw/ and
writes the five canonical CSVs to data/West_Virginia/cleaned/.

────────────────────────────────────────────────────────────────────────────
Input files
────────────────────────────────────────────────────────────────────────────
  CON_{year}.csv                  30 cols — "Contributions and Loans"
  EXP_{year}.csv                  28 cols — "Expenditures"
  transactions_{CAT}_{year}.jsonl per-row JSON grid, newline-delimited — the
                                  CURRENT production source for transactions,
                                  since CFRS retired the bulk CSVs. Streamed
                                  line-by-line: a single year runs to hundreds
                                  of MB. (.json from earlier builds still reads.)
  export_{CAT}_{year}.csv         whole-grid export — one request per year
                                  instead of hundreds, but MUCH thinner: 11
                                  columns (CON) / 10 (EXP) against the grid's
                                  52. No transactionID (so no filing_id and no
                                  cross-year dedup), no amendedFlag, no
                                  electionYear, RegistrantID instead of orgID,
                                  and city/state/zip only inside one combined
                                  address. Off by default — see
                                  EXPORT_TRANSACTIONS_READY in the scraper.
  export_committees.csv           registry export
  entities_committees.json        candidate + committee registry
  entities_{offices,elections,parties,violations}.json   enrichment, optional
  lookup_contributor_types_{CAT}.json                    vocabulary, optional

Column order for the two CSVs is fixed by the state's published record
layouts (linked in docs/states/west_virginia.md) and is read POSITIONALLY,
by index, not by header name. That is deliberate, for two reasons:

  1. EXP_*.csv has *two* columns both headed "Expenditure Type" (index 17,
     the schedule grouping; index 27, the monetary/in-kind classification).
     csv.DictReader silently collapses duplicate headers, so the second
     value would overwrite the first and one of the two fields would vanish
     without any error. Positional reads keep both.
  2. It makes the parser immune to header casing/whitespace drift, which has
     already been observed to differ between year files.

────────────────────────────────────────────────────────────────────────────
The quoting problem
────────────────────────────────────────────────────────────────────────────
CFRS wraps character columns in double-quotes but does NOT escape
double-quotes occurring *inside* them, e.g.

    ...,"Friends of "Big Jim" Smith","Charleston",...

There is no CSV dialect that reads this correctly, so _split_row() runs a
staged repair per line: parse as-is, and if that yields either the wrong
field count OR any field still containing a `"`, rewrite interior quotes
(any `"` not adjacent to a comma or a line boundary) to apostrophes and
re-parse. The second condition matters — Python's csv reader is non-strict
and will happily return the right number of fields with a mangled value
inside one of them, which a count check alone would wave through. A
minority of CON rows are also short
one trailing column — the state omits the final "Contribution Type" rather
than emitting an empty field — so a row one short of the expected width is
padded rather than discarded. Rows that survive neither pass are counted and
reported as `malformed`, never silently dropped.

Raw files are left byte-for-byte as downloaded; all repair happens here, in
memory, so raw_file + row_num still point at the real source line.

────────────────────────────────────────────────────────────────────────────
Notable transformations
────────────────────────────────────────────────────────────────────────────
  committee_name is REQUIRED by the schema but is empty on every candidate-
  committee transaction row in the source — CFRS populates candidate_name
  instead and leaves the committee blank. It is resolved in three steps:
  the row's own value, then the registry keyed on ORG ID, then the
  candidate's name. Without this the state fails tier-1 validation outright.

  Loans are split out of CON. The file layout is "Contributions and Loans";
  rows whose Receipt Type or Contribution Type mentions a loan, or that
  carry a Forgiven Loan value, are routed to loans_debts.csv.gz instead of
  contributions.csv.gz.

  Names arrive split across LAST/FIRST/MIDDLE/SUFFIX, with non-individuals
  carrying the whole entity name in LAST NAME and a literal " " (single
  space, not empty) in the other three. _person_name() rebuilds a single
  display name and treats whitespace-only parts as absent.

  election_year comes from the leading year of REPORT NAME ("2022 4th
  Quarter Report" → 2022), falling back to the source file's year. It is not
  taken from the transaction date, which frequently lands in the year before
  the cycle it was reported under.

  person_id uses id_model="committee": WV issues a new ORG ID per committee
  registration, so the same candidate has different IDs across cycles and
  assign_person_ids collapses them on (state, candidate_name, office,
  district).

────────────────────────────────────────────────────────────────────────────
Known limitations
────────────────────────────────────────────────────────────────────────────
  - Office, district and party exist only in the registry feed. If
    entities_committees.json is absent, candidates are reconstructed from
    transaction rows and those three fields come out empty (the same
    tradeoff new_hampshire.py documents).
  - CFRS bulk coverage starts at 2018; earlier WV filings were paper.
  - contributor_city / _zip are missing on roughly half of CON rows in the
    source itself — not a parsing artifact. Measured against 5,000 real 2018
    JSON rows: city 66%, state/zip 78%, employer 29%, occupation 26%,
    candidate_name 82%. Everything the schema requires is 100%.
  - `election_year` is CFRS's own `electionYear`, which is the COMMITTEE'S
    registration cycle, not the transaction's year. Verified: the 44,503
    contributions carrying election_year=2012 have transaction dates spanning
    2018-01-15 to 2025-05-07, and belong to committees like "MORRISEY FOR AG
    2012" that registered for the 2012 cycle and kept filing for years after.
    The value is real source data and is left as-is rather than silently
    reinterpreted — but for "transactions in year N", use `date`, not
    `election_year`.
  - The CSV export is a viable but lossy alternative to the grid. Measured
    on the real 2018 export, 4,000 rows: committee_name / amount / date /
    transaction_type / contributor_name / contributor_type all 100%, loans
    still split (96 of 4,000), but filing_id is 0% (the column doesn't
    exist) and city/state/zip recover to only ~61% for CON and ~87% for EXP
    because they must be parsed back out of a combined address string.
  - The JSON grid's field names are OBSERVED, not inferred (see
    _json_common). An earlier inferred set matched amount and date but
    missed the name, type, employer, occupation and amended flag — which is
    the worst case, because amount+date resolving made the mapping look
    healthy. _parse_json_file now guards on the contributor/payee name
    separately for exactly that reason.
"""

import csv
import gzip
import json
import re
import sys
import time
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

# Some CFRS description/explanation fields run long; lift the default limit
# before any reader is constructed.
csv.field_size_limit(10 * 1024 * 1024)

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))
from src.reporting.logger import get_logger

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import columns as C
import utils

try:
    from tqdm import tqdm
except ImportError:                              # tqdm is optional
    def tqdm(it, **kwargs):                      # type: ignore
        return it

PROJECT_ROOT = Path(__file__).resolve().parents[3]
RAW_DIR      = PROJECT_ROOT / "data" / "West_Virginia" / "raw"
CLEAN_DIR    = PROJECT_ROOT / "data" / "West_Virginia" / "cleaned"
CLEAN_DIR.mkdir(parents=True, exist_ok=True)

STATE = "WV"


# ====================== source column positions =======================
# Index-based, per the state's published record layouts. See the module
# docstring for why these are positional rather than header-keyed.

CON_WIDTH = 30
CON = {
    "org_id": 0, "amount": 1, "date": 2,
    "last_name": 3, "first_name": 4, "middle_name": 5, "suffix": 6,
    "address1": 7, "address2": 8, "city": 9, "state": 10, "zip": 11,
    "description": 12, "receipt_id": 13, "filed_date": 14,
    "source_type": 15, "amended": 16, "receipt_type": 17,
    "committee_type": 18, "committee_name": 19, "candidate_name": 20,
    "employer": 21, "occupation": 22, "occupation_comment": 23,
    "forgiven_loan": 24,
    "fundraiser_date": 25, "fundraiser_type": 26, "fundraiser_place": 27,
    "report_name": 28, "contribution_type": 29,
}

EXP_WIDTH = 28
EXP = {
    "org_id": 0, "amount": 1, "date": 2,
    "last_name": 3, "first_name": 4, "middle_name": 5, "suffix": 6,
    "address1": 7, "address2": 8, "city": 9, "state": 10, "zip": 11,
    "explanation": 12, "expenditure_id": 13, "filed_date": 14,
    "purpose": 15, "amended": 16,
    "schedule_type": 17,          # first "Expenditure Type" — schedule grouping
    "committee_type": 18, "committee_name": 19, "candidate_name": 20,
    "fundraiser_date": 21, "fundraiser_type": 22, "fundraiser_place": 23,
    "support_or_oppose": 24, "candidate": 25, "report_name": 26,
    "monetary_type": 27,          # second "Expenditure Type" — Monetary/In-Kind/...
}

# Receipt/contribution type values that mean "this is a loan, not a gift".
# Matched as a case-insensitive substring so variants like "Loan Received"
# and "Repayment of Loan" are all caught.
LOAN_MARKERS = ("loan",)


# =========================== small helpers ============================

def clean(val) -> str:
    """Strip whitespace and coerce None to empty string.

    CFRS uses a literal single space rather than an empty field for "not
    provided" in the name and address columns, so stripping is what turns
    those into true blanks.
    """
    return (val or "").strip()


def parse_amount(val) -> str:
    """Parse a dollar amount to a plain numeric string. Returns '' on failure."""
    v = str(val or "").strip().replace("$", "").replace(",", "")
    if not v:
        return ""
    if v.startswith("(") and v.endswith(")"):
        v = "-" + v[1:-1]                 # accounting-style negative
    try:
        float(v)
        return v
    except ValueError:
        return ""


_DATE_FORMATS = (
    "%m/%d/%Y %I:%M:%S %p",   # CFRS's normal form: 11/30/2022 12:00:00 AM
    "%m/%d/%Y %H:%M:%S",
    "%m/%d/%Y",
    "%Y-%m-%dT%H:%M:%S",      # the JSON grid tier serves ISO
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d",
    "%m-%d-%Y",
)


def parse_date(val) -> str:
    """Normalize a date to YYYY-MM-DD. Returns '' on failure or implausible year."""
    v = str(val or "").strip()
    if not v:
        return ""
    if "." in v and "T" in v:             # trim ISO fractional seconds
        v = v.split(".", 1)[0]
    for fmt in _DATE_FORMATS:
        try:
            d = datetime.strptime(v, fmt)
        except ValueError:
            continue
        if d.year < 1990 or d.year > date.today().year + 2:
            return ""
        return d.strftime("%Y-%m-%d")
    return ""


def _year_of(val: str) -> str:
    """Leading 4-digit year in a string, or '' — used on REPORT NAME."""
    m = re.search(r"(19|20)\d{2}", val or "")
    return m.group(0) if m else ""


def _year_from_filename(name: str) -> str:
    m = re.search(r"(19|20)\d{2}", name)
    return m.group(0) if m else ""


def _person_name(last: str, first: str, middle: str, suffix: str) -> str:
    """Rebuild one display name from CFRS's split name columns.

    Non-individual sources (businesses, PACs, unions) put the entire entity
    name in LAST NAME and a literal single space in the other three, so an
    entity comes back unchanged while a person comes back as
    "First Middle Last Suffix".
    """
    last, first, middle, suffix = (clean(x) for x in (last, first, middle, suffix))
    if not (first or middle or suffix):
        return utils.clean_name(last)
    parts = [p for p in (first, middle, last, suffix) if p]
    return utils.clean_name(" ".join(parts))


def amended_flag(val) -> str:
    """Normalize CFRS's Y/N amendment flag to the schema's 0/1.

    Matches the convention the other states use (alabama's yn_to_int,
    arkansas' amended_flag) so the column is comparable in the aggregate DB;
    validate.py flags raw Y/N as a tier-2 warning. Unrecognized values become
    empty rather than a guessed 0.
    """
    v = clean(val).upper()
    if v in ("Y", "YES", "TRUE", "1"):
        return "1"
    if v in ("N", "NO", "FALSE", "0"):
        return "0"
    return ""


def _is_loan(receipt_type: str, contribution_type: str, forgiven: str) -> bool:
    """True if a CON row belongs in loans_debts rather than contributions."""
    blob = f"{receipt_type} {contribution_type}".lower()
    if any(m in blob for m in LOAN_MARKERS):
        return True
    return bool(clean(forgiven))


def open_writer(filename: str, fieldnames: list):
    fh = gzip.open(CLEAN_DIR / filename, "wt", encoding="utf-8", newline="")
    w  = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore", restval="")
    w.writeheader()
    return fh, w


def raw_files(pattern: str) -> list[Path]:
    """Non-empty raw files matching a glob, oldest year first."""
    return sorted((f for f in RAW_DIR.glob(pattern) if f.stat().st_size > 0),
                  key=lambda p: p.name)


# ======================== CSV row repair ==============================

# Any double-quote that is NOT at a line boundary and NOT adjacent to a comma
# is an unescaped interior quote. Rewriting those to apostrophes is what makes
# the line parseable; the same rule the IRW Accountability Project arrived at
# independently for this source.
_INTERIOR_QUOTE = re.compile(r'(?<!^)(?<!,)"(?!,)(?!$)')


def _split_row(line: str, expected: int) -> tuple[list[str] | None, str]:
    """Parse one raw CSV line into exactly `expected` fields.

    Returns (fields, status) where status is one of "clean", "repaired",
    "padded", "truncated" or "malformed". Nothing is dropped silently — the
    caller counts every status and reports the non-clean ones.
    """
    line = line.rstrip("\r\n")
    if not line.strip():
        return None, "malformed"

    # Stage 1 — parse as served.
    try:
        row = next(csv.reader([line]))
    except csv.Error:
        row = None

    # A correct field count is NOT sufficient to call the line clean. Python's
    # csv reader is non-strict: given `"Friends of "Big Jim" Smith"` it closes
    # the quote early and appends the remainder literally, producing the right
    # number of fields with a mangled value inside. A surviving double-quote in
    # any parsed field is the tell — CFRS never escapes them, so one that made
    # it through parsing is always damage, never data.
    if row is not None and len(row) == expected and not any('"' in v for v in row):
        return row, "clean"

    # Stage 2 — rewrite interior quotes and re-parse.
    try:
        repaired = next(csv.reader([_INTERIOR_QUOTE.sub("'", line)]))
    except csv.Error:
        repaired = None

    if repaired is not None and len(repaired) == expected:
        return repaired, "repaired"

    # Stage 2b — the repair changed the field count (rare, and it means the
    # line's quoting is beyond reconstruction). If the naive parse at least
    # had the right shape, keep it and scrub the leftover quote characters
    # so they don't leak into names.
    if row is not None and len(row) == expected:
        return [v.replace('"', "'") for v in row], "repaired"

    # Stage 3 — accept a near-miss. Prefer whichever attempt got closer.
    best = None
    for cand in (repaired, row):
        if cand is None:
            continue
        if best is None or abs(len(cand) - expected) < abs(len(best) - expected):
            best = cand
    if best is None:
        return None, "malformed"

    if len(best) == expected - 1:
        # Known CFRS defect: the trailing column is omitted rather than
        # emitted empty on a minority of CON rows.
        return best + [""], "padded"
    if len(best) < expected:
        return best + [""] * (expected - len(best)), "padded"
    return best[:expected], "truncated"


def _read_rows(path: Path, expected: int):
    """Yield (row_num, fields, status) for every data line in a CFRS CSV.

    row_num is 1-based including the header, matching the convention in
    docs/contributing.md (enumerate from 2), so a row_num in the database
    points at that exact line of the raw file.
    """
    with open(path, "r", encoding="utf-8", errors="replace", newline="") as f:
        header = f.readline()
        if not header:
            return
        for row_num, line in enumerate(f, start=2):
            fields, status = _split_row(line, expected)
            yield row_num, fields, status


# ===================== JSON key access (tier 2) =======================

def _norm_key(k: str) -> str:
    """Normalize a JSON key for matching: lowercase, punctuation stripped."""
    return re.sub(r"[^a-z0-9]", "", (k or "").lower())


def _flatten(rec: dict, _prefix: str = "", _depth: int = 0) -> dict:
    """Flatten a nested record to {normalized_leaf_key: scalar}.

    CFRS's grid rows average ~1.7 KB, which is far too fat to be flat, and a
    top-level-only lookup silently returns nothing for every nested field —
    producing empty output rather than an error. Matching on the leaf name
    means the mapping works whether a field arrives as `orgName` or
    `committee.orgName`.

    Shallower keys win on collision, so a top-level `amount` is preferred
    over a nested `detail.amount`.
    """
    out: dict = {}
    if not isinstance(rec, dict) or _depth > 4:
        return out
    for k, v in rec.items():
        if isinstance(v, dict):
            for nk, nv in _flatten(v, k, _depth + 1).items():
                out.setdefault(nk, nv)
        elif isinstance(v, list):
            continue                       # repeating groups aren't scalars
        else:
            out.setdefault(_norm_key(k), v)   # shallower wins
    return out


def _pick(rec: dict, *aliases: str, default: str = "") -> str:
    """Case-insensitive, nesting-tolerant lookup across possible key spellings.

    CFRS has already been observed switching between PascalCase and camelCase
    for the same fields between API generations (and the export endpoint uses
    PascalCase while the paged grid uses camelCase), so matching on a
    normalized key rather than an exact one is load-bearing, not defensive
    styling.
    """
    if not isinstance(rec, dict):
        return default
    norm = rec.get("__flat__") if isinstance(rec.get("__flat__"), dict) else _flatten(rec)
    for alias in aliases:
        key = _norm_key(alias)
        if key in norm and norm[key] is not None:
            val = str(norm[key]).strip()
            if val and val.lower() not in ("null", "none"):
                return val
    return default


# ============================= registry ===============================

class Registry:
    """ORG ID → committee/candidate metadata, loaded from the registry feed.

    Supplies the three fields transactions don't carry (office, district,
    party) and, critically, the committee_name that CFRS leaves blank on
    candidate-committee transaction rows.
    """

    def __init__(self):
        self.by_org: dict[str, dict] = {}
        self.rows:   list[dict] = []
        self.loaded = False

    def load(self, log) -> int:
        path = RAW_DIR / "entities_committees.json"
        if not path.exists():
            log.info("  no entities_committees.json — candidates/committees "
                     "will be reconstructed from transactions")
            return 0

        try:
            records = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as e:
            log.file_parse_error(path.name, f"unreadable registry: {e}")
            return 0

        if not isinstance(records, list):
            log.file_parse_error(path.name, "registry is not a JSON array")
            return 0

        for rec in records:
            org_id = _pick(rec, "orgId", "entityId", "organizationId",
                           "committeeId", "id")
            if not org_id:
                continue
            entry = {
                "state_filer_id": org_id,
                "committee_name": utils.clean_name(
                    _pick(rec, "orgName", "committeeName", "organizationName",
                          "name")),
                "committee_type": _pick(rec, "orgType", "orgSubTypeCode",
                                        "committeeType", "organizationType"),
                "candidate_name": utils.clean_name(
                    _pick(rec, "candidateName", "candidateFullName")),
                "treasurer_name": _pick(rec, "officerName", "treasurerName"),
                "office":         _pick(rec, "officeName", "office", "officeTitle"),
                "district":       _pick(rec, "districtName", "district",
                                        "districtId"),
                "jurisdiction":   _pick(rec, "countyName", "jurisdiction"),
                "party":          _pick(rec, "partyName", "party", "partyCode"),
                "city":           _pick(rec, "city", "orgCity", "mailingCity"),
                "zip":            utils.clean_zip(
                    _pick(rec, "zip", "zipCode", "orgZip", "postalCode")),
                "election_year":  _year_of(_pick(rec, "electionYear",
                                                 "electionName",
                                                 "registrationDate")),
                "status":         _pick(rec, "orgStatus", "status"),
            }
            # First registration wins on duplicate ORG IDs — the feed is
            # sorted newest-first, and later pages can repeat a row when the
            # underlying grid shifts between requests.
            self.by_org.setdefault(org_id, entry)
            self.rows.append(entry)

        self.loaded = bool(self.by_org)
        log.registry_loaded(path.name, entries=len(self.by_org),
                            relation="committees")
        return len(self.by_org)

    def get(self, org_id: str) -> dict:
        return self.by_org.get(clean(org_id), {})


# ==================== transaction-derived entities ====================

class EntityAccumulator:
    """Builds candidate/committee rows out of transaction rows.

    Every CON and EXP row carries ORG ID, COMMITTEE TYPE, COMMITTEE NAME and
    CANDIDATE NAME, which is enough to reconstruct a usable registry when the
    registry feed is missing — the approach new_hampshire.py uses. When the
    feed IS present this still runs, and fills in any ORG ID that appears in
    transactions but not in the registry (committees that have since been
    dissolved and dropped off the public grid).
    """

    def __init__(self):
        # org_id → observed values, most recent election year wins
        self.orgs: dict[str, dict] = {}
        self.years: dict[str, str] = defaultdict(str)

    def observe(self, org_id: str, committee_name: str, committee_type: str,
                candidate_name: str, election_year: str) -> None:
        org_id = clean(org_id)
        if not org_id:
            return
        entry = self.orgs.setdefault(org_id, {
            "state_filer_id": org_id,
            "committee_name": "",
            "committee_type": "",
            "candidate_name": "",
            "election_year":  "",
        })
        # Keep the first non-empty value seen for each field; prefer a later
        # election year for election_year itself so the row reflects the most
        # recent cycle the committee was active in.
        if committee_name and not entry["committee_name"]:
            entry["committee_name"] = committee_name
        if committee_type and not entry["committee_type"]:
            entry["committee_type"] = committee_type
        if candidate_name and not entry["candidate_name"]:
            entry["candidate_name"] = candidate_name
        if election_year and election_year > entry["election_year"]:
            entry["election_year"] = election_year


# ========================== row builders ==============================

def _resolve_committee_name(row_value: str, reg: dict, candidate_name: str) -> str:
    """committee_name is required by the schema; CFRS often leaves it blank.

    Resolution order: the row's own value → the registry entry for that ORG
    ID → the candidate's own name. The last step is what keeps the tier-1
    fill-rate check passing on candidate-committee rows, which are the
    majority of the file and are blank in the source 100% of the time.
    """
    name = utils.clean_name(row_value)
    if name:
        return name
    if reg.get("committee_name"):
        return reg["committee_name"]
    return utils.clean_name(candidate_name)


def _contribution_row(f: list[str], src: str, row_num: int,
                      reg: dict, file_year: str) -> dict:
    election_year = _year_of(f[CON["report_name"]]) or file_year
    candidate = utils.clean_name(f[CON["candidate_name"]])
    return {
        "state":             STATE,
        "committee_name":    _resolve_committee_name(
                                 f[CON["committee_name"]], reg, candidate),
        "amount":            parse_amount(f[CON["amount"]]),
        "date":              parse_date(f[CON["date"]]),
        "transaction_type":  clean(f[CON["contribution_type"]])
                             or clean(f[CON["receipt_type"]]),
        "contributor_name":  _person_name(f[CON["last_name"]], f[CON["first_name"]],
                                          f[CON["middle_name"]], f[CON["suffix"]]),
        "contributor_type":  clean(f[CON["source_type"]]),
        "contributor_city":  clean(f[CON["city"]]),
        "contributor_state": clean(f[CON["state"]]).upper()[:2],
        "contributor_zip":   utils.clean_zip(f[CON["zip"]]),
        "employer":          clean(f[CON["employer"]]),
        "occupation":        clean(f[CON["occupation"]])
                             or clean(f[CON["occupation_comment"]]),
        "candidate_name":    candidate,
        "office":            reg.get("office", ""),
        "election_year":     election_year,
        "amended":           amended_flag(f[CON["amended"]]),
        "filing_id":         clean(f[CON["receipt_id"]]),
        "raw_file":          src,
        "row_num":           row_num,
    }


def _loan_row(f: list[str], src: str, row_num: int,
              reg: dict, file_year: str) -> dict:
    election_year = _year_of(f[CON["report_name"]]) or file_year
    candidate = utils.clean_name(f[CON["candidate_name"]])

    # record_type normally comes straight from the source. The exception is a
    # row that landed here only because FORGIVEN LOAN was populated — its
    # Receipt Type still reads "Contributions", which would be actively
    # misleading on a loans_debts row.
    record_type = clean(f[CON["receipt_type"]]) or clean(f[CON["contribution_type"]])
    blob = f"{f[CON['receipt_type']]} {f[CON['contribution_type']]}".lower()
    if clean(f[CON["forgiven_loan"]]) and not any(m in blob for m in LOAN_MARKERS):
        record_type = "Forgiven Loan"

    return {
        "state":               STATE,
        "committee_name":      _resolve_committee_name(
                                   f[CON["committee_name"]], reg, candidate),
        "original_amount":     parse_amount(f[CON["amount"]]),
        "date":                parse_date(f[CON["date"]]),
        "record_type":         record_type,
        "counterparty_name":   _person_name(f[CON["last_name"]], f[CON["first_name"]],
                                            f[CON["middle_name"]], f[CON["suffix"]]),
        "counterparty_city":   clean(f[CON["city"]]),
        "counterparty_state":  clean(f[CON["state"]]).upper()[:2],
        "counterparty_zip":    utils.clean_zip(f[CON["zip"]]),
        "candidate_name":      candidate,
        "election_year":       election_year,
        "amended":             amended_flag(f[CON["amended"]]),
        "filing_id":           clean(f[CON["receipt_id"]]),
        "raw_file":            src,
        "row_num":             row_num,
    }


def _expenditure_row(f: list[str], src: str, row_num: int,
                     reg: dict, file_year: str) -> dict:
    election_year = _year_of(f[EXP["report_name"]]) or file_year
    candidate = utils.clean_name(f[EXP["candidate_name"]])
    return {
        "state":            STATE,
        "committee_name":   _resolve_committee_name(
                                f[EXP["committee_name"]], reg, candidate),
        "amount":           parse_amount(f[EXP["amount"]]),
        "date":             parse_date(f[EXP["date"]]),
        # Column 27 is the monetary classification (Monetary / In-Kind /
        # Disbursement of Excess Funds); column 17 is the schedule grouping
        # and is the better fallback when 27 is blank.
        "transaction_type": clean(f[EXP["monetary_type"]])
                            or clean(f[EXP["schedule_type"]]),
        "payee_name":       _person_name(f[EXP["last_name"]], f[EXP["first_name"]],
                                         f[EXP["middle_name"]], f[EXP["suffix"]]),
        "purpose":          clean(f[EXP["purpose"]]) or clean(f[EXP["explanation"]]),
        "category":         clean(f[EXP["schedule_type"]]),
        "payee_city":       clean(f[EXP["city"]]),
        "payee_state":      clean(f[EXP["state"]]).upper()[:2],
        "payee_zip":        utils.clean_zip(f[EXP["zip"]]),
        "candidate_name":   candidate,
        "office":           reg.get("office", ""),
        "election_year":    election_year,
        "amended":          amended_flag(f[EXP["amended"]]),
        "filing_id":        clean(f[EXP["expenditure_id"]]),
        "raw_file":         src,
        "row_num":          row_num,
    }


# ===================== tier-2 JSON row builders =======================

_ADDRESS_TAIL = re.compile(
    r",\s*(?P<city>[^,]+?)\s*,\s*(?P<state>[A-Za-z]{2})\s*,\s*(?P<zip>\d{5}(?:-\d{4})?)\s*$")


def _split_address(rec: dict) -> tuple[str, str, str]:
    """Recover (city, state, zip) from the export's combined address field.

    The CSV export ships "500 Virginia St, Charleston, WV, 25322" as one
    column where the JSON grid has three. Only a trailing
    ", city, ST, zip" is matched — anything else returns blanks rather than
    guessing, because a wrong city is worse than a missing one. Measured on
    the real 2018 export this recovers ~74% of non-empty addresses.
    """
    raw = _pick(rec, "contributorAddress", "recipientAddress", "address")
    if not raw:
        return "", "", ""
    m = _ADDRESS_TAIL.search(raw)
    if not m:
        return "", "", ""
    return m.group("city").strip(), m.group("state").upper(), m.group("zip")


def _json_common(rec: dict) -> dict:
    """Fields shared by both categories in the JSON grid tier.

    Names below are OBSERVED, taken from a real CON_2018 payload — not
    inferred. The earlier inferred set got `transactionAmount`,
    `transactionDate`, `committeeName`, `orgID` and `transactionID` right and
    everything else wrong, which was the dangerous outcome: amount and date
    resolving meant the mapping looked healthy while contributor name, type,
    employer, occupation and the amended flag all came through empty.

    Differences from the legacy CSV layout worth knowing:
      - the contributor/payee arrives as ONE field (contributorPayeeName),
        not split across LAST/FIRST/MIDDLE/SUFFIX, so no name assembly.
      - electionYear is a real integer field, so the cycle no longer has to
        be scraped out of the report name — which matters because EXP rows
        carry names like "Final Report" with no year in them at all.
      - amendedFlag is a JSON boolean, not "Y"/"N".
    """
    return {
        # RegistrantID is the export's filer id and is NOT the grid's orgID
        # (1020001517 vs 1517). Kept last so the grid's value always wins when
        # both are present.
        "org_id":         _pick(rec, "orgID", "orgId", "entityID",
                                "organizationID", "registrantID"),
        "amount":         _pick(rec, "transactionAmount", "amount"),
        "date":           _pick(rec, "transactionDate", "date"),
        # Single combined name; entities keep their full name here too.
        "payee_name":     _pick(rec, "contributorPayeeName", "contributorPayee",
                                "contributorName", "payeeName",
                                "contributor", "recipientName"),
        # The export ships one combined address instead of separate parts;
        # _split_address recovers what it can.
        "city":           _pick(rec, "city") or _split_address(rec)[0],
        "state":          _pick(rec, "stateCode", "state") or _split_address(rec)[1],
        "zip":            _pick(rec, "zipCode", "zip") or _split_address(rec)[2],
        "committee_name": _pick(rec, "committeeName", "orgName", "committee"),
        "candidate_name": _pick(rec, "candidateName", "candidate"),
        "committee_type": _pick(rec, "orgType", "committeeType"),
        "committee_subtype": _pick(rec, "orgSubType", "committeeSubType"),
        "entity_type":    _pick(rec, "entityTypeDesc", "entityType",
                                "contributorType", "payeeType", "recipientType"),
        # transactionCategoryDesc is the money classification (Monetary,
        # In-Kind, Other Income); transactionTypeDesc is the schedule
        # grouping (Contributions, Expenditures, Loans).
        "category_desc":  _pick(rec, "transactionCategoryDesc",
                                "transactionCategory", "contributionType",
                                "expenditureType"),
        "type_desc":      _pick(rec, "transactionTypeDesc", "transactionType"),
        "purpose":        _pick(rec, "transactionPurpose", "purpose",
                                "expenditurePurpose"),
        "employer":       _pick(rec, "employerName", "employer"),
        "occupation":     _pick(rec, "employerOccupation", "occupation"),
        "report_name":    _pick(rec, "reportFileName", "reportName", "report"),
        "election_year":  _pick(rec, "electionYear"),
        "amended":        _pick(rec, "amendedFlag", "amended"),
        "filing_id":      _pick(rec, "transactionID", "transactionId",
                                "transactionNumber"),
    }


def _json_election_year(b: dict, file_year: str) -> str:
    """Cycle year for a JSON-tier row.

    Prefers the source's own electionYear integer. Falls back to the leading
    year of the report name, then the file's year — the fallbacks matter
    because EXP report names are frequently just "Final Report".
    """
    yr = clean(b.get("election_year"))
    if yr.isdigit() and 1990 <= int(yr) <= date.today().year + 2:
        return yr
    return _year_of(b.get("report_name", "")) or file_year


def _json_contribution_row(rec: dict, src: str, row_num: int,
                           reg: dict, file_year: str) -> dict:
    b = _json_common(rec)
    candidate = utils.clean_name(b["candidate_name"])
    return {
        "state":             STATE,
        "committee_name":    _resolve_committee_name(b["committee_name"], reg, candidate),
        "amount":            parse_amount(b["amount"]),
        "date":              parse_date(b["date"]),
        "transaction_type":  b["category_desc"] or b["type_desc"],
        "contributor_name":  utils.clean_name(b["payee_name"]),
        "contributor_type":  b["entity_type"],
        "contributor_city":  b["city"],
        "contributor_state": b["state"].upper()[:2],
        "contributor_zip":   utils.clean_zip(b["zip"]),
        "employer":          b["employer"],
        "occupation":        b["occupation"],
        "candidate_name":    candidate,
        "office":            reg.get("office", ""),
        "election_year":     _json_election_year(b, file_year),
        "amended":           amended_flag(b["amended"]),
        "filing_id":         b["filing_id"],
        "raw_file":          src,
        "row_num":           row_num,
    }


def _json_loan_row(rec: dict, src: str, row_num: int,
                   reg: dict, file_year: str) -> dict:
    """A CON-category row whose schedule grouping marks it as a loan."""
    b = _json_common(rec)
    candidate = utils.clean_name(b["candidate_name"])
    return {
        "state":              STATE,
        "committee_name":     _resolve_committee_name(b["committee_name"], reg, candidate),
        "original_amount":    parse_amount(b["amount"]),
        "date":               parse_date(b["date"]),
        "record_type":        b["type_desc"] or b["category_desc"],
        "counterparty_name":  utils.clean_name(b["payee_name"]),
        "counterparty_city":  b["city"],
        "counterparty_state": b["state"].upper()[:2],
        "counterparty_zip":   utils.clean_zip(b["zip"]),
        "candidate_name":     candidate,
        "election_year":      _json_election_year(b, file_year),
        "amended":            amended_flag(b["amended"]),
        "filing_id":          b["filing_id"],
        "raw_file":           src,
        "row_num":            row_num,
    }


def _json_expenditure_row(rec: dict, src: str, row_num: int,
                          reg: dict, file_year: str) -> dict:
    b = _json_common(rec)
    candidate = utils.clean_name(b["candidate_name"])
    return {
        "state":            STATE,
        "committee_name":   _resolve_committee_name(b["committee_name"], reg, candidate),
        "amount":           parse_amount(b["amount"]),
        "date":             parse_date(b["date"]),
        # Mirrors the CSV tier: the money classification is transaction_type,
        # the schedule grouping is category.
        "transaction_type": b["category_desc"] or b["type_desc"],
        "payee_name":       utils.clean_name(b["payee_name"]),
        "purpose":          b["purpose"],
        "category":         b["type_desc"],
        "payee_city":       b["city"],
        "payee_state":      b["state"].upper()[:2],
        "payee_zip":        utils.clean_zip(b["zip"]),
        "candidate_name":   candidate,
        "office":           reg.get("office", ""),
        "election_year":    _json_election_year(b, file_year),
        "amended":          amended_flag(b["amended"]),
        "filing_id":        b["filing_id"],
        "raw_file":         src,
        "row_num":          row_num,
    }


# ============================== parsing ===============================

def _parse_con_file(path: Path, cont_w, loan_w, reg: Registry,
                    acc: EntityAccumulator, seen: set, log) -> tuple[int, int, int]:
    """Parse one CON_{year}.csv. Returns (contributions, loans, skipped)."""
    t0 = time.perf_counter()
    file_year = _year_from_filename(path.name)
    n_cont = n_loan = n_skip = 0
    status_counts: dict[str, int] = defaultdict(int)

    for row_num, fields, status in tqdm(
            _read_rows(path, CON_WIDTH), desc=f"  {path.name}",
            unit="row", dynamic_ncols=True, leave=False):
        status_counts[status] += 1
        if fields is None:
            n_skip += 1
            continue

        org_id   = clean(fields[CON["org_id"]])
        reg_e    = reg.get(org_id)
        filing_id = clean(fields[CON["receipt_id"]])

        # CFRS republishes an amended transaction under its original Receipt
        # ID in the year file it was amended into, so the same ID can appear
        # in more than one year's export. Keep the first occurrence.
        if filing_id:
            dedup_key = ("CON", filing_id)
            if dedup_key in seen:
                n_skip += 1
                continue
            seen.add(dedup_key)

        election_year = _year_of(fields[CON["report_name"]]) or file_year
        acc.observe(org_id,
                    utils.clean_name(fields[CON["committee_name"]]),
                    clean(fields[CON["committee_type"]]),
                    utils.clean_name(fields[CON["candidate_name"]]),
                    election_year)

        if _is_loan(fields[CON["receipt_type"]],
                    fields[CON["contribution_type"]],
                    fields[CON["forgiven_loan"]]):
            loan_w.writerow(_loan_row(fields, path.name, row_num, reg_e, file_year))
            n_loan += 1
        else:
            cont_w.writerow(_contribution_row(fields, path.name, row_num,
                                              reg_e, file_year))
            n_cont += 1

    _log_file(log, path, "contributions", n_cont + n_loan, n_skip,
              status_counts, t0)
    return n_cont, n_loan, n_skip


def _parse_exp_file(path: Path, expn_w, reg: Registry,
                    acc: EntityAccumulator, seen: set, log) -> tuple[int, int]:
    """Parse one EXP_{year}.csv. Returns (expenditures, skipped)."""
    t0 = time.perf_counter()
    file_year = _year_from_filename(path.name)
    n_exp = n_skip = 0
    status_counts: dict[str, int] = defaultdict(int)

    for row_num, fields, status in tqdm(
            _read_rows(path, EXP_WIDTH), desc=f"  {path.name}",
            unit="row", dynamic_ncols=True, leave=False):
        status_counts[status] += 1
        if fields is None:
            n_skip += 1
            continue

        org_id    = clean(fields[EXP["org_id"]])
        filing_id = clean(fields[EXP["expenditure_id"]])
        if filing_id:
            dedup_key = ("EXP", filing_id)
            if dedup_key in seen:
                n_skip += 1
                continue
            seen.add(dedup_key)

        election_year = _year_of(fields[EXP["report_name"]]) or file_year
        acc.observe(org_id,
                    utils.clean_name(fields[EXP["committee_name"]]),
                    clean(fields[EXP["committee_type"]]),
                    utils.clean_name(fields[EXP["candidate_name"]]),
                    election_year)

        expn_w.writerow(_expenditure_row(fields, path.name, row_num,
                                         reg.get(org_id), file_year))
        n_exp += 1

    _log_file(log, path, "expenditures", n_exp, n_skip, status_counts, t0)
    return n_exp, n_skip


def _iter_export_csv(path: Path):
    """Yield each row of a grid-export CSV as a dict keyed by column header.

    Distinct from the legacy CON_*.csv / EXP_*.csv readers, which are
    positional: those are the state's own malformed export, this is a file
    this pipeline wrote itself from the xlsx, so it is well-formed CSV and
    DictReader is safe. The two never collide — export files are prefixed
    "export_", legacy files start with the category code.
    """
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        for row in csv.DictReader(f):
            yield row


def _iter_xlsx(path: Path):
    """Yield each row of a CFRS grid export as a dict keyed by column header.

    The export is what `generateExportGridDataExcel` returns. Its column
    headers are the grid's *display* labels rather than API field names, but
    _norm_key() strips spaces and punctuation and lowercases, so
    "Transaction Amount" collapses onto the same key as `transactionAmount`.
    That's why one alias table serves both the JSON grid and the spreadsheet
    without a second mapping.

    read_only + values_only so a large export streams instead of building a
    full cell object graph.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "reading CFRS .xlsx exports needs openpyxl (already in "
            "requirements.txt) — pip install openpyxl") from e

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return
        keys = [str(h).strip() if h is not None else "" for h in header]
        for values in rows:
            if values is None or all(v is None or str(v).strip() == "" for v in values):
                continue
            yield {k: v for k, v in zip(keys, values) if k}
    finally:
        wb.close()


def _iter_jsonl(path: Path):
    """Yield each record of a newline-delimited JSON file.

    Blank lines are skipped; a malformed line is skipped rather than killing
    the whole file, since one bad row shouldn't cost an entire year.
    """
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                yield json.loads(line)
            except json.JSONDecodeError:
                continue


def _parse_json_file(path: Path, writer, kind: str, reg: Registry,
                     acc: EntityAccumulator, seen: set, log,
                     loan_writer=None) -> tuple[int, int]:
    """Parse a tier-2 transactions_{CAT}_{year}.json file."""
    t0 = time.perf_counter()
    file_year = _year_from_filename(path.name)
    n_ok = n_skip = 0

    # .jsonl is streamed line-by-line; .json (whole-array, written by earlier
    # scraper builds) is loaded entirely. Streaming matters at WV's volumes —
    # a single year runs to hundreds of MB, which as one parsed array is
    # several GB resident.
    try:
        if path.suffix in (".xlsx", ".xls"):
            records = _iter_xlsx(path)
        elif path.suffix == ".csv":
            records = _iter_export_csv(path)
        elif path.suffix == ".jsonl":
            records = _iter_jsonl(path)
        else:
            records = json.loads(path.read_text(encoding="utf-8"))
            if not isinstance(records, list):
                log.file_parse_error(path.name, "expected a JSON array")
                return 0, 0
    except (OSError, json.JSONDecodeError) as e:
        log.file_parse_error(path.name, str(e))
        return 0, 0

    build = _json_contribution_row if kind == "contributions" else _json_expenditure_row
    n_loan = 0

    # The field aliases in _json_common() were inferred, not observed. Print
    # the record's actual (flattened) keys once per file so a mismatch is
    # immediately diagnosable instead of silently yielding empty columns.
    first_keys: list = []
    mapped_ok = 0          # rows where amount + date resolved
    named_ok  = 0          # rows where the contributor/payee name resolved

    for row_num, rec in enumerate(records, start=1):
        if row_num == 1 and isinstance(rec, dict):
            # The aliases in _json_common() were inferred, not observed. Print
            # the record's actual (flattened) keys once per file so a mismatch
            # is diagnosable instead of silently yielding empty columns.
            first_keys = sorted(_flatten(rec).keys())
            log.info(f"  {path.name}: record keys = {first_keys}")
        b = _json_common(rec)
        if b["amount"] and b["date"]:
            mapped_ok += 1
        # Tracked separately: amount+date matched even in the half-broken
        # alias table, so they alone are not evidence the mapping is sound.
        if b["payee_name"]:
            named_ok += 1
        if b["filing_id"]:
            dedup_key = (kind[:3].upper(), b["filing_id"])
            if dedup_key in seen:
                n_skip += 1
                continue
            seen.add(dedup_key)

        acc.observe(b["org_id"], utils.clean_name(b["committee_name"]),
                    b["committee_type"], utils.clean_name(b["candidate_name"]),
                    _json_election_year(b, file_year))

        # CON carries loans alongside contributions here just as it does in
        # the legacy CSVs; transactionTypeDesc is the schedule grouping that
        # tells them apart. Without this the JSON tier would file every loan
        # as a contribution.
        if (kind == "contributions" and loan_writer is not None
                and _is_loan(b["type_desc"], b["category_desc"], "")):
            loan_writer.writerow(_json_loan_row(rec, path.name, row_num,
                                                reg.get(b["org_id"]), file_year))
            n_loan += 1
        else:
            writer.writerow(build(rec, path.name, row_num,
                                  reg.get(b["org_id"]), file_year))
        n_ok += 1

    # A mapping failure is not a parse failure — every row still gets written,
    # just with empty amount/date, which sails through as "valid but useless"
    # and only shows up as a validator fill-rate warning much later. Fail
    # loudly here instead, naming the keys that were actually available.
    if n_ok and named_ok / n_ok < 0.5:
        raise ValueError(
            f"{path.name}: field mapping is resolving amount/date but not the "
            f"contributor/payee name ({named_ok:,} of {n_ok:,}). That pattern "
            f"means the alias table half-matches this feed. Available keys: "
            f"{first_keys}")

    if n_ok and mapped_ok / n_ok < 0.5:
        available = first_keys
        raise ValueError(
            f"{path.name}: field mapping failed — only {mapped_ok:,} of "
            f"{n_ok:,} rows resolved both amount and date. The aliases in "
            f"_json_common() don't match this feed. Available keys: {available}"
        )

    log.file_parsed(path.name, kind, n_ok, skipped=n_skip,
                    duration_s=time.perf_counter() - t0,
                    bytes=path.stat().st_size,
                    mapped_ok=mapped_ok, named_ok=named_ok, loans=n_loan)
    return n_ok, n_skip


def _log_file(log, path: Path, relation: str, rows: int, skipped: int,
              status_counts: dict, t0: float) -> None:
    """Emit the per-file parse event, surfacing repair counts.

    Repaired/padded/truncated counts are worth seeing in the run report: a
    sudden jump means CFRS changed its export, and this is the only place
    that would show it before the data looked wrong downstream.
    """
    repaired  = status_counts.get("repaired", 0)
    padded    = status_counts.get("padded", 0)
    truncated = status_counts.get("truncated", 0)
    malformed = status_counts.get("malformed", 0)

    log.file_parsed(path.name, relation, rows, skipped=skipped,
                    duration_s=time.perf_counter() - t0,
                    bytes=path.stat().st_size,
                    repaired=repaired, padded=padded,
                    truncated=truncated, malformed=malformed)

    if malformed or truncated:
        log.warning(f"    {path.name}: {malformed:,} malformed, "
                    f"{truncated:,} over-wide row(s)")


# ======================= entity output writing ========================

def _write_entities(cand_w, cmte_w, reg: Registry,
                    acc: EntityAccumulator, log) -> tuple[int, int]:
    """Write candidates.csv.gz and committees.csv.gz.

    The registry feed is authoritative when present. Any ORG ID seen only in
    transactions is then appended from the accumulator, so a committee that
    has since dropped off the public grid still gets a row rather than
    leaving its transactions orphaned.
    """
    n_cand = n_cmte = 0
    written_orgs: set[str] = set()

    for entry in reg.rows:
        org_id = entry["state_filer_id"]
        if org_id in written_orgs:
            continue
        written_orgs.add(org_id)

        cmte_w.writerow({
            "state":          STATE,
            "committee_name": entry["committee_name"],
            "committee_type": entry["committee_type"],
            "election_year":  entry["election_year"],
            "candidate_name": entry["candidate_name"],
            "treasurer_name": entry["treasurer_name"],
            "city":           entry["city"],
            "zip":            entry["zip"],
            # CFRS exposes a free-text status; map the affirmative values to
            # 1 and everything else (Terminated, Inactive, ...) to 0. Left
            # blank when the feed didn't carry a status at all.
            "active":         _active_flag(entry["status"]),
            "state_filer_id": org_id,
            "raw_file":       "entities_committees.json",
            "row_num":        n_cmte + 2,
        })
        n_cmte += 1

        if entry["candidate_name"]:
            first, last = _split_candidate(entry["candidate_name"])
            cand_w.writerow({
                "state":           STATE,
                "candidate_name":  entry["candidate_name"],
                "candidate_first": first,
                "candidate_last":  last,
                "office":          entry["office"],
                "district":        entry["district"],
                "jurisdiction":    entry["jurisdiction"],
                "party":           entry["party"],
                "election_year":   entry["election_year"],
                "state_filer_id":  org_id,
                "raw_file":        "entities_committees.json",
                "row_num":         n_cand + 2,
            })
            n_cand += 1

    # Office/district/party for a candidate seen only in transactions.
    #
    # This matters more than it looks: assign_person_ids(id_model="committee")
    # groups on (state, candidate_name, office, district), so a derived row
    # with those fields blank would NOT merge with the same person's registry
    # row and the candidate would end up with two different person_ids. Only
    # unambiguous names are carried over — if two registry candidates share a
    # name, leaving the fields blank is the safer error.
    name_meta: dict[str, dict | None] = {}
    for entry in reg.rows:
        cname = entry["candidate_name"]
        if not cname:
            continue
        meta = {k: entry[k] for k in ("office", "district", "jurisdiction", "party")}
        if cname in name_meta and name_meta[cname] != meta:
            name_meta[cname] = None      # ambiguous — don't guess
        else:
            name_meta.setdefault(cname, meta)

    # Backfill anything the registry didn't cover.
    backfilled_c = backfilled_p = 0
    for org_id, entry in sorted(acc.orgs.items()):
        if org_id in written_orgs:
            continue
        cmte_w.writerow({
            "state":          STATE,
            "committee_name": entry["committee_name"] or entry["candidate_name"],
            "committee_type": entry["committee_type"],
            "election_year":  entry["election_year"],
            "candidate_name": entry["candidate_name"],
            "state_filer_id": org_id,
            "raw_file":       "derived_from_transactions",
            "row_num":        n_cmte + 2,
        })
        n_cmte += 1
        backfilled_c += 1

        if entry["candidate_name"]:
            first, last = _split_candidate(entry["candidate_name"])
            meta = name_meta.get(entry["candidate_name"]) or {}
            cand_w.writerow({
                "state":           STATE,
                "candidate_name":  entry["candidate_name"],
                "candidate_first": first,
                "candidate_last":  last,
                "office":          meta.get("office", ""),
                "district":        meta.get("district", ""),
                "jurisdiction":    meta.get("jurisdiction", ""),
                "party":           meta.get("party", ""),
                "election_year":   entry["election_year"],
                "state_filer_id":  org_id,
                "raw_file":        "derived_from_transactions",
                "row_num":         n_cand + 2,
            })
            n_cand += 1
            backfilled_p += 1

    log.enrichment_summary(
        registry_committees=len(reg.rows),
        derived_committees=backfilled_c,
        derived_candidates=backfilled_p,
        source="registry" if reg.loaded else "transactions",
    )
    return n_cand, n_cmte


_ACTIVE_VALUES = {"active", "open", "a", "y", "yes", "true", "1"}


def _active_flag(status: str) -> str:
    """Map CFRS's free-text org status onto the schema's 0/1 `active` column."""
    s = clean(status).lower()
    if not s:
        return ""
    return "1" if s in _ACTIVE_VALUES else "0"


def _split_candidate(name: str) -> tuple[str, str]:
    """Split a candidate name into (first, last).

    CFRS's registry writes candidates as "Justice, James C., II" in some
    feeds and "Lissa Lucas" in others, so both orderings are handled.
    """
    name = clean(name)
    if not name:
        return "", ""
    if "," in name:
        last, _, rest = name.partition(",")
        first = rest.strip().split(" ")[0] if rest.strip() else ""
        return first.strip(), last.strip()
    parts = name.split()
    if len(parts) == 1:
        return "", parts[0]
    return parts[0], parts[-1]


# ================================ run =================================

def run():
    log = get_logger("west_virginia", "parse")
    t0  = time.perf_counter()
    log._emit("parse_started")

    total_contributions = 0
    total_expenditures  = 0
    total_loans         = 0
    total_skipped       = 0
    committees_written  = 0
    candidates_written  = 0
    file_handles        = []

    try:
        cont_fh, cont_w = open_writer("contributions.csv.gz", C.CONTRIBUTIONS)
        expn_fh, expn_w = open_writer("expenditures.csv.gz",  C.EXPENDITURES)
        cand_fh, cand_w = open_writer("candidates.csv.gz",    C.CANDIDATES)
        cmte_fh, cmte_w = open_writer("committees.csv.gz",    C.COMMITTEES)
        loan_fh, loan_w = open_writer("loans_debts.csv.gz",   C.LOANS_DEBTS)
        file_handles = [cont_fh, expn_fh, cand_fh, cmte_fh, loan_fh]

        registry = Registry()
        registry.load(log)

        acc  = EntityAccumulator()
        seen: set = set()

        # ---- contributions + loans ----
        con_files = raw_files("CON_*.csv")
        for path in con_files:
            n_cont, n_loan, n_skip = _parse_con_file(
                path, cont_w, loan_w, registry, acc, seen, log)
            total_contributions += n_cont
            total_loans         += n_loan
            total_skipped       += n_skip

        # ---- expenditures ----
        exp_files = raw_files("EXP_*.csv")
        for path in exp_files:
            n_exp, n_skip = _parse_exp_file(path, expn_w, registry, acc, seen, log)
            total_expenditures += n_exp
            total_skipped      += n_skip

        # ---- tier-2 JSON fallback ----
        # Only present when the scraper couldn't reach the bulk CSVs. Parsed
        # after the CSVs and through the same `seen` set, so if both exist for
        # a year the richer CSV rows win and the JSON adds nothing duplicate.
        for path in list(raw_files("transactions_CON_*.json*")) + \
                    list(raw_files("export_CON_*.csv")) + \
                    list(raw_files("export_CON_*.xls*")):
            n_ok, n_skip = _parse_json_file(path, cont_w, "contributions",
                                            registry, acc, seen, log,
                                            loan_writer=loan_w)
            total_contributions += n_ok
            total_skipped       += n_skip
        for path in list(raw_files("transactions_EXP_*.json*")) + \
                    list(raw_files("export_EXP_*.csv")) + \
                    list(raw_files("export_EXP_*.xls*")):
            n_ok, n_skip = _parse_json_file(path, expn_w, "expenditures",
                                            registry, acc, seen, log)
            total_expenditures += n_ok
            total_skipped      += n_skip

        if not con_files and not exp_files and not total_contributions:
            log.warning("  no raw transaction files found in "
                        f"{RAW_DIR} — run the scraper first")

        # ---- entities ----
        candidates_written, committees_written = _write_entities(
            cand_w, cmte_w, registry, acc, log)

        # Close before person-ID assignment — those helpers rewrite the
        # candidate/committee files in place and would read a truncated gzip
        # stream if the handles were still open.
        for fh in file_handles:
            fh.close()
        file_handles = []

        # WV issues a new ORG ID per committee registration, so the same
        # person carries different IDs across cycles — the "committee" model
        # collapses them on (state, candidate_name, office, district).
        utils.assign_person_ids(CLEAN_DIR / "candidates.csv.gz",
                                id_model="committee")
        utils.assign_committee_person_ids(CLEAN_DIR / "committees.csv.gz",
                                          CLEAN_DIR / "candidates.csv.gz")

        def _bytes(name: str) -> int:
            p = CLEAN_DIR / name
            return p.stat().st_size if p.exists() else 0

        log.file_parsed("contributions.csv.gz", "contributions",
                        total_contributions, role="output",
                        bytes=_bytes("contributions.csv.gz"))
        log.file_parsed("expenditures.csv.gz", "expenditures",
                        total_expenditures, role="output",
                        bytes=_bytes("expenditures.csv.gz"))
        log.file_parsed("candidates.csv.gz", "candidates",
                        candidates_written, role="output",
                        bytes=_bytes("candidates.csv.gz"))
        log.file_parsed("committees.csv.gz", "committees",
                        committees_written, role="output",
                        bytes=_bytes("committees.csv.gz"))
        log.file_parsed("loans_debts.csv.gz", "loans_debts",
                        total_loans, role="output",
                        bytes=_bytes("loans_debts.csv.gz"))

        duration = round(time.perf_counter() - t0, 1)
        log.info(f"Done in {duration}s — {total_contributions:,} contributions, "
                 f"{total_expenditures:,} expenditures, {total_loans:,} loans, "
                 f"{candidates_written:,} candidates, {committees_written:,} committees")
        log._emit("parse_completed", status="completed", duration_s=duration,
                  contributions=total_contributions,
                  expenditures=total_expenditures,
                  loans_debts=total_loans, skipped=total_skipped,
                  committees=committees_written, candidates=candidates_written)

    except KeyboardInterrupt:
        log._emit("parse_completed", status="interrupted",
                  duration_s=round(time.perf_counter() - t0, 1),
                  contributions=total_contributions,
                  expenditures=total_expenditures,
                  committees=committees_written, candidates=candidates_written)
        raise

    except Exception as e:
        log._emit("parse_completed", status="error",
                  duration_s=round(time.perf_counter() - t0, 1),
                  contributions=total_contributions,
                  expenditures=total_expenditures,
                  committees=committees_written, candidates=candidates_written,
                  error_type=type(e).__name__, error=str(e))
        raise

    finally:
        for fh in file_handles:
            try:
                fh.close()
            except Exception:
                pass


# ================================ CLI =================================

if __name__ == "__main__":
    try:
        run()
    except KeyboardInterrupt:
        sys.exit(130)
    except Exception:
        sys.exit(1)
