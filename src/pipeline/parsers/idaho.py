"""
idaho.py — Parse Idaho campaign finance data into canonical cleaned CSVs.

Raw files (data/Idaho/raw/):
  portal_candidates.csv                 — current portal candidate registry (2020-2026,
                                           keyed by "Filing Entity Id", one row per
                                           registration/election year)
  portal_committees.csv                 — current portal PAC / Central Committee registry
  portal_contributions_{YYYY}.csv       — 2020-2026 contributions (incl. loans received)
  portal_expenditures_{YYYY}.csv        — 2020-2026 expenditures (incl. debt/loan payments)
  electionstats_activity.csv            — canvass.sos.idaho.gov "donate"/"spend" activity
                                           feed, full pull (382,124 rows, 2020-2024).
                                           Only ES_YEAR_MIN..ES_YEAR_MAX (2020-2022) rows
                                           are emitted — that's the gap the portal
                                           (real data from 2023) and archive (<=2018)
                                           don't cover. 2023+ rows are skipped to avoid
                                           double-counting against the portal.
  archive_{YYYY}_{cand|comm}_{cont|exp}.{xls,xlsx} — SOS archive, 2000-2018, 26 files
                                           with ~6 distinct header layouts across years

id_model = "name_hash"
  Idaho has no person-level ID anywhere in its 26-year dataset:
    - Portal "Filing Entity Id" is per-registration/cycle, not per-person —
      empirically confirmed 34 distinct (candidate, office, district) groups
      have multiple Filing Entity Ids across election years (e.g. Kelly Anthon:
      588 in 2024, 209 in 2024, 209 in 2026).
    - The 2000-2018 archive and the 2020-2022 electionstats_activity feed
      carry no numeric ID at all.
  person_id is therefore derived from MD5(state + normalized candidate_name)
  for every candidate row (utils.assign_person_ids(id_model="name_hash")),
  which unifies a person across cycles/sources purely by name.

Naming convention: candidate_name / committee_name / contributor_name /
payee_name are all written UPPERCASE via utils.clean_name(), unlike some
other state parsers which preserve source casing on transaction rows. This
is deliberate — Idaho candidate names must be matched (via name_to_committee
and the registry dicts below) across four very differently-formatted source
families, and a single normalized case avoids silent match failures.

Loan/debt routing:
  Portal contributions Transaction Type in {Loan Received, Loan Forgiven,
    Outstanding Loan}                          -> loans_debts.csv.gz
  Portal expenditures Transaction Type in {Debt Payment, Loan Payment, Debt,
    Outstanding Debt}                          -> loans_debts.csv.gz
  electionstats donate_type == "Loan"          -> loans_debts.csv.gz
  electionstats spend_type in {"Loan Payment", "Loan Interest",
    "Credit Card Payment", "Credit Card Interest/Fee"}
                                                -> loans_debts.csv.gz
  Archive cont Type code in {"L","Loan"}       -> loans_debts.csv.gz
  Archive exp  Type code == "Repayment"        -> loans_debts.csv.gz

2020-2022 gap (see docs/states/idaho.md): RESOLVED 2026-06-13. Previously,
2020-2022 donations TO PAC/Central Committees and ALL spending by
Candidate/PAC committees were unavailable — the only third-party source
(id.electionstats.com, non-.gov, donate->Candidate only) has been removed
per Henry's .gov-only policy and replaced with the full
electionstats_activity.csv pull above, which covers donate AND spend for
both Candidate and PAC committees in 2020-2022.

Output (data/Idaho/cleaned/):
  contributions.csv.gz, expenditures.csv.gz, loans_debts.csv.gz,
  committees.csv.gz, candidates.csv.gz
"""

import csv
import gzip
import os
import pickle
import re
import sys
import time
from pathlib import Path
from datetime import date, datetime

import openpyxl
import xlrd

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))
from src.reporting.logger import get_logger

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import columns as C
import utils

csv.field_size_limit(sys.maxsize)

# =============================== Paths ================================
PROJECT_ROOT = Path(__file__).resolve().parents[3]
RAW_DIR      = PROJECT_ROOT / "data" / "Idaho" / "raw"
CLEAN_DIR    = PROJECT_ROOT / "data" / "Idaho" / "cleaned"
CLEAN_DIR.mkdir(parents=True, exist_ok=True)

STATE          = "ID"
# 1999 (not 2000) because the 2000 archive files (covering the 2000 election
# cycle) legitimately contain ~8.3k contribution/expenditure rows dated in
# late 1999; a handful of clearly-erroneous outliers (1930s-1990s, <10 rows
# total) are still rejected by this floor.
EARLIEST_YEAR  = 1999
MAX_VALID_YEAR = date.today().year + 4

# Portal Transaction Type values that represent loan/debt activity rather
# than ordinary contributions/expenditures.
LOAN_INCOME_TYPES  = {"Loan Received", "Loan Forgiven", "Outstanding Loan"}
LOAN_EXPENSE_TYPES = {"Debt Payment", "Loan Payment", "Debt", "Outstanding Debt"}

# Archive Type-code -> canonical transaction_type. Codes vary by year
# (single-letter pre-2012, blank/"In Kind"/"Repayment" 2012+). Loan-related
# codes ("L","Loan","Repayment") are intercepted before this map is consulted.
CONT_TYPE_MAP = {
    "C": "Contribution",
    "I": "In-Kind",
    "In Kind": "In-Kind",
    "R": "Return Contribution",
    "P": "Contribution",   # rare/ambiguous (~60 rows/yr, 2004 & 2010 only)
    "": "Contribution",
}
EXP_TYPE_MAP = {
    "N": "Expenditure",
    "I": "In-Kind",
    "In Kind": "In-Kind",
    "R": "Return Expenditure",
    "": "Expenditure",
}


# ============================== Helpers ===============================
def clean(val) -> str:
    """Strip whitespace and coerce None to empty string."""
    return (val or "").strip()


def parse_amount(val: str) -> str:
    """Parse a dollar amount string to a plain numeric string; parentheses become negative. Returns '' on failure."""
    v = (val or "").strip().replace("$", "").replace(",", "")
    if not v:
        return ""
    if v.startswith("(") and v.endswith(")"):
        v = "-" + v[1:-1]
    try:
        float(v)
        return v
    except ValueError:
        return ""


def parse_date(val) -> str:
    """Parse 'MM/DD/YYYY', 'MM-DD-YYYY', or 'YYYY-MM-DD[ HH:MM:SS]' -> 'YYYY-MM-DD'. Returns '' on failure or out-of-range year."""
    v = clean(str(val)) if val is not None else ""
    if not v:
        return ""
    m = re.match(r"^(\d{4}-\d{2}-\d{2})", v)
    if m:
        try:
            d = datetime.strptime(m.group(1), "%Y-%m-%d")
        except ValueError:
            return ""
    else:
        d = None
        for fmt in ("%m/%d/%Y", "%m-%d-%Y"):
            try:
                d = datetime.strptime(v, fmt)
                break
            except ValueError:
                continue
        if d is None:
            return ""
    if d.year < EARLIEST_YEAR or d.year > MAX_VALID_YEAR:
        return ""
    return d.strftime("%Y-%m-%d")


def bool01(val) -> str:
    """Map Y/N (any case) -> 1/0; pass through existing 0/1; anything else -> ''."""
    v = clean(val).upper()
    if v == "Y":
        return "1"
    if v == "N":
        return "0"
    if v in ("0", "1"):
        return v
    return ""


def split_name(raw: str) -> tuple[str, str]:
    """'Last, First Middle' -> (first_middle, last). No comma -> (\"\", raw)."""
    raw = clean(raw)
    if "," in raw:
        last, _, first = raw.partition(",")
        return first.strip(), last.strip()
    return "", raw


def format_name(raw: str) -> str:
    """'Last, First Middle' -> 'First Middle Last'. No comma -> raw unchanged."""
    first, last = split_name(raw)
    if first and last:
        return f"{first} {last}"
    return last or first


def clean_zip_field(val) -> str:
    """Strip an Excel formula wrapper (=\"83706\") if present, then normalize via utils.clean_zip."""
    v = clean(val)
    m = re.match(r'^="?([^"]*)"?$', v)
    if m:
        v = m.group(1)
    return utils.clean_zip(v)


def raw_files(pattern: str) -> list[Path]:
    """Return non-empty raw files matching a glob pattern, sorted by name.

    IDAHO_PHASE is a temporary three-call checkpoint/resume mechanism for
    environments with a hard wall-clock cap per invocation (the full run is
    ~50-60s, well over a 45s cap). Stages: "early" (portal/electionstats/cand_cont/
    cand_exp), "comm_cont" (archive comm_cont, all years incl. 2018),
    "comm_exp" (archive comm_exp, all years incl. 2018, + final flush).
    IDAHO_PHASE=1 -> only "early"; IDAHO_PHASE=2 -> only "comm_cont";
    IDAHO_PHASE=3 -> only "comm_exp". Unset (normal/full run): no filtering,
    identical to the original behavior.
    """
    phase = os.environ.get("IDAHO_PHASE", "")
    if "comm_cont" in pattern:
        stage = "comm_cont"
    elif "comm_exp" in pattern:
        stage = "comm_exp"
    elif "legacy_2019_pdfs" in pattern:
        stage = "legacy_2019"
    else:
        stage = "early"
    if phase == "1" and stage != "early":
        return []
    if phase == "2" and stage not in ("comm_cont", "legacy_2019"):
        return []
    if phase == "3" and stage != "comm_exp":
        return []
    return sorted(
        (f for f in RAW_DIR.glob(pattern) if f.stat().st_size > 0),
        key=lambda p: p.name,
    )


def open_writer(filename: str, fieldnames: list):
    """Open a gzipped CSV writer in CLEAN_DIR; extra fields are dropped, missing fields default to ''."""
    fh = gzip.open(CLEAN_DIR / filename, "wt", encoding="utf-8", newline="")
    w  = csv.DictWriter(fh, fieldnames=fieldnames,
                        extrasaction="ignore", restval="")
    w.writeheader()
    return fh, w


def open_writer_append(filename: str, fieldnames: list):
    """Reopen a gzipped CSV writer in CLEAN_DIR in append mode, no header row.

    Appending to an existing .gz file creates a new gzip "member"; gzip's
    decompressor (and pandas) read concatenated multi-member gzip streams
    transparently as one continuous stream. Used by IDAHO_PHASE=2 to resume
    writing to contributions/expenditures/loans_debts after a phase-1 run.
    """
    fh = gzip.open(CLEAN_DIR / filename, "at", encoding="utf-8", newline="")
    w  = csv.DictWriter(fh, fieldnames=fieldnames,
                        extrasaction="ignore", restval="")
    return fh, w


def _fill(d: dict, key: str, val: str) -> None:
    """Set d[key] = val only if val is truthy and d[key] is currently empty."""
    if val and not d.get(key):
        d[key] = val


# ================= Legacy 2019 per-filer PDF reports (C-2 form) =================
# data/Idaho/raw/legacy_2019_pdfs/*.pdf — one C-2 "Campaign Financial
# Disclosure Report" PDF per filer (First Annual or Mid-Year report, some
# _amended/_terminated/_termination variants), covering 2019 activity not
# otherwise available (see docs/states/idaho.md). Text-extractable via
# `pdftotext -layout` (poppler-utils). Parses Schedule A (itemized
# contributions >$50), Schedule B (itemized expenditures >=$25), Schedule C
# (in-kind, paired contribution+expenditure), and Schedule D (loans).
# Schedules E (Credit Card/Debt) and F (Pledged Contributions) have not been
# observed populated in any sample and are not parsed — documented gap.
PDF_MONEY = r'\$[\d,]+\.\d{2}'
PDF_DATE  = r'\d{2}/\d{2}/\d{4}'

# Schedule A's header row is "Date  ElectionType  Contributor  Amount  YTDAmount" —
# ElectionType is often populated ("Primary"/"General") and otherwise sits
# between Date and Contributor as blank padding. Strip a leading ElectionType
# token from the captured contributor name when present.
_PDF_ELECTION_TYPE_RE = re.compile(
    r'^(Primary Election|General Election|Special Election|Runoff Election|'
    r'Primary|General|Special|Runoff|Consolidated)\s{2,}(.+)$'
)


def _pdf_section_header(line: str):
    """Return 'A'..'F' if `line` is a real "Schedule X" section header (not
    a "...Total of all Schedule X sheets..." reference on the summary
    page), else None. pdftotext -layout inserts a form-feed (\\x0c) before
    page-break section headers."""
    s = line.strip("\x0c").strip()
    m = re.match(r'^Schedule\s+([A-F])(\s*-\s*\S.*)?$', s)
    return m.group(1) if m else None


def _pdf_to_text(path: Path) -> str:
    import subprocess
    try:
        return subprocess.run(["pdftotext", "-layout", str(path), "-"],
                               capture_output=True, text=True, timeout=30).stdout
    except Exception:
        return ""


def _pdf_amt(s: str) -> str:
    return s.replace("$", "").replace(",", "")


def _pdf_addr(addr_lines: list[str]) -> tuple[str, str, str, str]:
    """Split trailing address lines into (street, city, state, zip)."""
    if not addr_lines:
        return "", "", "", ""
    last = addr_lines[-1]
    cm = re.match(r'^(.*?),\s*([A-Z]{2})\s+(\d{5}(?:-\d{4})?)$', last)
    if not cm:
        cm = re.match(r'^(.*?)\s+([A-Z]{2})\s+(\d{5}(?:-\d{4})?)$', last)
    if cm:
        city, state, zipc = cm.groups()
        return " ".join(addr_lines[:-1]), city, state, zipc
    return " ".join(addr_lines), "", "", ""


def _pdf_parse_header(text: str) -> tuple[str, str, str]:
    """Returns (filer_name, office, district) from the page-1 header line."""
    lines = text.split("\n")
    for i, line in enumerate(lines):
        if "Name of Candidate or Political Committee and Chairperson" in line:
            for j in range(i + 1, min(i + 4, len(lines))):
                data = lines[j].strip()
                if data:
                    parts = re.split(r'\s{2,}', data)
                    name = parts[0].strip()
                    office = parts[1].strip() if len(parts) > 1 else ""
                    district = parts[2].strip() if len(parts) > 2 else ""
                    return name, office, district
    return "", "", ""


def _pdf_schedule_a(text: str) -> list[dict]:
    """Schedule A — itemized contributions (>$50)."""
    rows, lines, in_a, i = [], text.split("\n"), False, 0
    while i < len(lines):
        sh = _pdf_section_header(lines[i])
        if sh == "A":
            in_a = True
            i += 1
            continue
        if in_a and sh is not None:
            break
        if not in_a:
            i += 1
            continue
        m = re.match(rf'^\s*({PDF_DATE})\s+(.+?)\s+({PDF_MONEY})(?:\s+({PDF_MONEY}))?\s*$', lines[i])
        if m:
            date_, name, amount, _ytd = m.groups()
            em = _PDF_ELECTION_TYPE_RE.match(name)
            if em:
                name = em.group(2)
            addr_lines, j = [], i + 1
            while j < len(lines):
                nxt = lines[j]
                if (re.match(rf'^\s*{PDF_DATE}\s+', nxt) or "Grand Total" in nxt
                        or _pdf_section_header(nxt) is not None):
                    break
                if nxt.strip():
                    addr_lines.append(nxt.strip())
                j += 1
            street, city, state, zipc = _pdf_addr(addr_lines)
            rows.append(dict(date=date_, contributor_name=name.strip(), amount=_pdf_amt(amount),
                              city=city, state=state, zipc=zipc, street=street))
            i = j
            continue
        i += 1
    return rows


def _pdf_schedule_b(text: str) -> list[dict]:
    """Schedule B — itemized expenditures (>=$25)."""
    rows, lines, in_b, i = [], text.split("\n"), False, 0
    while i < len(lines):
        sh = _pdf_section_header(lines[i])
        if sh == "B":
            in_b = True
            i += 1
            continue
        if in_b and sh is not None:
            break
        if not in_b:
            i += 1
            continue
        m = re.match(rf'^\s*({PDF_DATE})\s+(.+?)\s{{2,}}({PDF_MONEY})\s+(?:([A-Z])\s+)?(?:Explanation:\s*(.*))?$', lines[i])
        if m:
            date_, name, amount, code, expl = m.groups()
            addr_lines, j = [], i + 1
            expl_parts = [expl.strip()] if expl else []
            while j < len(lines):
                nxt = lines[j]
                if (re.match(rf'^\s*{PDF_DATE}\s+', nxt) or "Grand Total" in nxt
                        or _pdf_section_header(nxt) is not None):
                    break
                stripped = nxt.strip()
                if stripped:
                    # Address and (wrapped) Explanation/Support/Oppose text share
                    # this line in different columns; a 3+ space gap separates
                    # them. Support:/Oppose: tokens (with no values) are dropped.
                    parts = re.split(r'\s{3,}', stripped)
                    left = parts[0].strip()
                    if left and not re.match(r'^(Support|Oppose):?$', left):
                        addr_lines.append(left)
                    for extra in parts[1:]:
                        extra = extra.strip()
                        if extra and not re.match(r'^(Support|Oppose):?$', extra):
                            expl_parts.append(extra)
                j += 1
            street, city, state, zipc = _pdf_addr(addr_lines)
            rows.append(dict(date=date_, payee_name=name.strip(), amount=_pdf_amt(amount),
                              category=code or "", purpose=" ".join(expl_parts).strip(),
                              city=city, state=state, zipc=zipc, street=street))
            i = j
            continue
        i += 1
    return rows


def _pdf_schedule_c(text: str) -> list[dict]:
    """Schedule C — in-kind contributions (paired contributor -> recipient)."""
    rows, lines, in_c, i = [], text.split("\n"), False, 0
    while i < len(lines):
        sh = _pdf_section_header(lines[i])
        if sh == "C":
            in_c = True
            i += 1
            continue
        if in_c and sh is not None:
            break
        if not in_c:
            i += 1
            continue
        m = re.match(rf'^\s*({PDF_DATE})\s+(.+?)\s{{2,}}(.+?)\s{{2,}}({PDF_MONEY})\s*(?:([A-Z])\s*)?$', lines[i])
        if m:
            date_, contrib, recip, amount, code = m.groups()
            rows.append(dict(date=date_, contributor_name=contrib.strip(), payee_name=recip.strip(),
                              amount=_pdf_amt(amount), category=code or ""))
        i += 1
    return rows


def _pdf_schedule_d(text: str) -> list[dict]:
    """Schedule D — loans (new loan amount this period + repayment, per lender)."""
    rows, lines, in_d, i = [], text.split("\n"), False, 0
    while i < len(lines):
        sh = _pdf_section_header(lines[i])
        if sh == "D":
            in_d = True
            i += 1
            continue
        if in_d and sh is not None:
            break
        if not in_d:
            i += 1
            continue
        if "Grand Total" in lines[i]:
            i += 1
            continue
        m = re.match(rf'^(.+?)\s{{2,}}({PDF_MONEY})\s+({PDF_DATE})\s+({PDF_MONEY})\s*({PDF_MONEY})?\s*({PDF_DATE})?\s*({PDF_MONEY})?\s*({PDF_MONEY})?\s*$', lines[i])
        if m:
            lender, _prev_bal, ldate, new_amt, _interest, pmtdate, payment, _current = m.groups()
            rows.append(dict(lender=lender.strip(), date=ldate, amount=_pdf_amt(new_amt),
                              pmtdate=pmtdate or "", payment=_pdf_amt(payment) if payment else ""))
        i += 1
    return rows


def _open_portal_csv(path: Path):
    """Open a portal export, discarding the 'X Download as of ...' metadata
    line that precedes the real header on row 2."""
    f = open(path, encoding="utf-8-sig", errors="replace", newline="")
    f.readline()
    return f


# ===================== Archive (xls/xlsx) helpers ======================
def _read_rows(path: Path):
    """Yield (row_num, row_dict) for any .xls/.xlsx file, using row 1 as the header.

    row_dict also carries '__datemode__' (xlrd workbook datemode, or None
    for openpyxl) so date cells can be converted correctly downstream.
    """
    if path.suffix.lower() == ".xls":
        wb = xlrd.open_workbook(str(path))
        sh = wb.sheet_by_index(0)
        header = [str(h).strip() for h in sh.row_values(0)]
        for i in range(1, sh.nrows):
            row = dict(zip(header, sh.row_values(i)))
            row["__datemode__"] = wb.datemode
            yield i + 1, row
    else:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sh = wb[wb.sheetnames[0]]
        rows_iter = sh.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
        for i, vals in enumerate(rows_iter, start=2):
            if vals is None or all(v is None for v in vals):
                continue
            row = dict(zip(header, vals))
            row["__datemode__"] = None
            yield i, row
        wb.close()


def cell_str(val) -> str:
    """Coerce any xls/xlsx cell value to a clean string (ints come back without trailing '.0')."""
    if val is None:
        return ""
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val)
    return str(val).strip()


def cell_amount(val) -> str:
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        return str(val)
    return parse_amount(str(val))


def cell_date(val, datemode=None) -> str:
    """Convert an xls/xlsx date cell (Excel serial float, datetime, or string) -> 'YYYY-MM-DD'."""
    if val is None or val == "":
        return ""
    if isinstance(val, datetime):
        if val.year < EARLIEST_YEAR or val.year > MAX_VALID_YEAR:
            return ""
        return val.strftime("%Y-%m-%d")
    if isinstance(val, (int, float)):
        try:
            d = xlrd.xldate_as_datetime(val, datemode or 0)
        except Exception:
            return ""
        if d.year < EARLIEST_YEAR or d.year > MAX_VALID_YEAR:
            return ""
        return d.strftime("%Y-%m-%d")
    return parse_date(str(val))


def pick(row: dict, *names: str) -> "object":
    """Return the first non-empty value among the given column-name aliases."""
    for n in names:
        if n in row and row[n] not in (None, ""):
            return row[n]
    return ""


def _archive_year(path: Path) -> str:
    """Extract the 4-digit year from an 'archive_YYYY_*' filename."""
    m = re.search(r"archive_(\d{4})_", path.name)
    return m.group(1) if m else ""


# ---- Column-name alias groups, covering ~7 distinct archive header layouts (2000-2018) ----
NAME_FIRST = ("Cand First", "CandFirst", "FirstName")
NAME_MID   = ("Cand Mi", "Cand Mid", "CandMid", "CandMiddle", "MiddleName")
NAME_LAST  = ("Cand Last", "CandLast", "LastName")
NAME_SUF   = ("Cand Suf", "CandSuf", "CandSuffix")
PARTY      = ("Party", "CandParty", "PartyName")
OFFICE     = ("Office", "CandOffice", "OfficeName")
DISTRICT   = ("District", "CandDistrict")

CONTR_DATE   = ("Date", "ContrDate", "ContributorDate")
CONTR_AMOUNT = ("Amount", "ContrAmount", "ContributorAmount")
CONTR_TYPE   = ("Type", "ContrType", "Contribution Type")
CONTR_CP     = ("ContrCP",)

CONTRIBUTOR_FIRST = ("Contrib First", "ContrFirst", "ContributorFirstName", "ContributorFirst")
CONTRIBUTOR_MID   = ("ContrMid", "ContributorMiddleName", "ContributorMiddle")
CONTRIBUTOR_LAST  = ("Contributor Last", "Contrib Last", "ContrLast", "ContrName", "ContributorName", "ContributorLastName")
CONTRIBUTOR_SUF   = ("ContrSuf", "ContributorSuffix")
CONTRIBUTOR_CITY  = ("City", "ContrCity", "ContributorCity")
CONTRIBUTOR_STATE = ("State", "ContrSt", "ContrState", "ContributorState")
CONTRIBUTOR_ZIP   = ("Zip", "ContrZip", "ContributorZipcode")
# 2018 layout: when the contributor is a PAC/company (not a person), the
# name lives in one of these columns instead of Contributor*Name.
CONTRIBUTOR_ENTITY = ("Contributor Committee/Company", "Contributing Committee/Company Name")

EXP_DATE   = ("Date", "ExpDate", "ExpenDate", "RecipientDate")
EXP_AMOUNT = ("Amount", "ExpAmount", "ExpenAmount")
EXP_TYPE   = ("Type", "ExpType", "ExpenType", "Expenditure Type")
EXP_CP     = ("ExpenCP",)

RECIP_FIRST = ("Recip First", "RecipFirst", "RecipientFirst", "ExpenFirst", "RecipientFirstName")
RECIP_MID   = ("RecipientMiddle", "ExpenMid", "RecipientMiddleName")
RECIP_LAST  = ("Recipient Last", "RecipLast", "RecipientName", "ExpenName", "RecipientLastName")
RECIP_SUF   = ("RecipientSuffix", "ExpenSuf")
RECIP_CITY  = ("City", "RecipCity", "RecipientCity", "ExpenCity")
RECIP_STATE = ("State", "RecipSt", "RecipientState", "ExpenState")
RECIP_ZIP   = ("Zip", "RecipZip", "RecipientZipcode", "ExpenZip")
# 2018 layout: when the recipient is a PAC/company (not a person), the
# name lives in one of these columns instead of Recipient*Name.
RECIP_ENTITY = ("Recipient Committee/Company Name", "Recipient Company/Committee")

PURPOSE  = ("Expense Purpose", "ExpensePurpose", "Expend Explanation",
            "Purpose Code 1 Description", "Purpose Code 1 Explanation")
CATEGORY = ("PurposeCodes", "PurposeCodeOne", "Purpose Code 1")

# 2018 layout: comm_cont/comm_exp identify the filing committee/PAC by name
# in one of these columns instead of the older "Committee" column.
COMMITTEE_NAME_COL = ("Committee", "Committee Name", "Name")


# ================================ Main ================================
def run():
    log = get_logger("idaho", "parse")
    t0  = time.perf_counter()
    log.info("Starting Idaho parser")
    log._emit("parse_started")

    # See raw_files() docstring: temporary three-call checkpoint/resume split
    # for environments with a hard wall-clock cap per invocation.
    PHASE = os.environ.get("IDAHO_PHASE", "")
    CKPT1 = CLEAN_DIR / "_checkpoint_idaho_phase1.pkl"
    CKPT2 = CLEAN_DIR / "_checkpoint_idaho_phase2.pkl"

    candidates: dict[str, dict] = {}
    committees: dict[str, dict] = {}

    # fid -> office sought, populated from portal_candidates.csv; used to
    # enrich the "office" field on portal contribution/expenditure rows
    # (which carry Filing Entity ID but not office directly).
    fid_office: dict[str, str] = {}

    # normalized candidate_name -> normalized committee/campaign name.
    # Seeded from portal_candidates.csv (real "Campaign Name" values) and
    # extended as legacy/archive candidates are registered, so transaction
    # rows from any source resolve to a consistent committee_name.
    name_to_committee: dict[str, str] = {}

    total_contributions = 0
    total_expenditures  = 0
    total_loans         = 0

    file_handles = []

    # =================== Registry helpers ===================
    def _register_candidate_committee(prefix: str, cand_name: str, election_year: str,
                                       party: str = "", office: str = "", district: str = "",
                                       jurisdiction: str = "", candidate_first: str = "",
                                       candidate_last: str = "", raw_file: str = "",
                                       row_num="") -> str:
        """Find-or-create a candidates + committees row for a legacy/archive
        candidate (no source ID available). Keys are synthesized from the
        normalized name + election year, prefixed to avoid collisions with
        portal-derived keys. Returns the resolved committee_name."""
        if not cand_name:
            return ""

        cand_key = f"{prefix}_cand_{cand_name}_{election_year}"
        cand = candidates.get(cand_key)
        if cand is None:
            cand = {
                "state":           STATE,
                "candidate_name":  cand_name,
                "candidate_first": candidate_first,
                "candidate_last":  candidate_last,
                "office":          office,
                "district":        district,
                "jurisdiction":    jurisdiction,
                "party":           party,
                "election_year":   election_year,
                "incumbent":       "",
                "state_filer_id":  "",
                "raw_file":        raw_file,
                "row_num":         row_num,
            }
            candidates[cand_key] = cand
        else:
            _fill(cand, "party", party)
            _fill(cand, "office", office)
            _fill(cand, "district", district)
            _fill(cand, "jurisdiction", jurisdiction)
            _fill(cand, "candidate_first", candidate_first)
            _fill(cand, "candidate_last", candidate_last)

        cmte_name = name_to_committee.get(cand_name, cand_name)
        name_to_committee.setdefault(cand_name, cmte_name)

        cmte_key = f"{prefix}_cmte_{cand_name}_{election_year}"
        if cmte_key not in committees:
            committees[cmte_key] = {
                "state":           STATE,
                "committee_name":  cmte_name,
                "committee_type":  "Candidate Committee",
                "election_year":   election_year,
                "candidate_name":  cand_name,
                "treasurer_name":  "",
                "city":            "",
                "zip":             "",
                "active":          "",
                "state_filer_id":  "",
                "raw_file":        raw_file,
                "row_num":         row_num,
            }
        return cmte_name

    def _register_pac(committee_name: str, election_year: str,
                       raw_file: str = "", row_num="") -> None:
        """Find-or-create a committees row for an archive-era PAC (comm_cont/comm_exp)."""
        if not committee_name:
            return
        key = f"arc_pac_{committee_name}"
        if key not in committees:
            committees[key] = {
                "state":           STATE,
                "committee_name":  committee_name,
                "committee_type":  "PAC",
                "election_year":   election_year,
                "candidate_name":  "",
                "treasurer_name":  "",
                "city":            "",
                "zip":             "",
                "active":          "",
                "state_filer_id":  "",
                "raw_file":        raw_file,
                "row_num":         row_num,
            }
        else:
            _fill(committees[key], "election_year", election_year)

    try:
        if PHASE in ("2", "3"):
            ckpt_path = CKPT1 if PHASE == "2" else CKPT2
            with open(ckpt_path, "rb") as f:
                _ckpt = pickle.load(f)
            candidates.update(_ckpt["candidates"])
            committees.update(_ckpt["committees"])
            name_to_committee.update(_ckpt["name_to_committee"])
            fid_office.update(_ckpt["fid_office"])
            total_contributions = _ckpt["total_contributions"]
            total_expenditures  = _ckpt["total_expenditures"]
            total_loans         = _ckpt["total_loans"]
            cont_fh, cont_w = open_writer_append("contributions.csv.gz", C.CONTRIBUTIONS)
            expn_fh, expn_w = open_writer_append("expenditures.csv.gz",  C.EXPENDITURES)
            loan_fh, loan_w = open_writer_append("loans_debts.csv.gz",   C.LOANS_DEBTS)
            log.info(f"Resumed from checkpoint: {ckpt_path}")
        else:
            cont_fh, cont_w = open_writer("contributions.csv.gz", C.CONTRIBUTIONS)
            expn_fh, expn_w = open_writer("expenditures.csv.gz",  C.EXPENDITURES)
            loan_fh, loan_w = open_writer("loans_debts.csv.gz",   C.LOANS_DEBTS)
        file_handles = [cont_fh, expn_fh, loan_fh]

        # =================== portal_candidates.csv ===================
        path = RAW_DIR / "portal_candidates.csv"
        if path.exists() and path.stat().st_size > 0 and PHASE in ("", "1"):
            ft, count = time.perf_counter(), 0
            f = _open_portal_csv(path)
            try:
                for row_num, row in enumerate(csv.DictReader(f), start=3):
                    fid = clean(row.get("Filing Entity Id", ""))
                    if not fid:
                        continue
                    first = clean(row.get("Candidate First Name", ""))
                    mid   = clean(row.get("Candidate Middle Name", ""))
                    last  = clean(row.get("Candidate Last Name", ""))
                    cname = utils.clean_name(f"{first} {mid} {last}")
                    ey    = clean(row.get("Election Year", ""))
                    office       = clean(row.get("Office Sought", ""))
                    district     = clean(row.get("District", ""))
                    jurisdiction = clean(row.get("District Type", ""))
                    party        = clean(row.get("Party Affiliation", ""))

                    candidates[f"cand_{fid}_{ey}"] = {
                        "state":           STATE,
                        "candidate_name":  cname,
                        "candidate_first": first,
                        "candidate_last":  last,
                        "office":          office,
                        "district":        district,
                        "jurisdiction":    jurisdiction,
                        "party":           party,
                        "election_year":   ey,
                        "incumbent":       "",
                        "state_filer_id":  fid,
                        "raw_file":        path.name,
                        "row_num":         row_num,
                    }

                    campaign_name = utils.clean_name(row.get("Campaign Name", "")) or cname
                    treas = utils.clean_name(
                        f"{clean(row.get('Treasurer First Name', ''))} {clean(row.get('Treasurer Last Name', ''))}"
                    )
                    committees[f"cand_{fid}_{ey}"] = {
                        "state":           STATE,
                        "committee_name":  campaign_name,
                        "committee_type":  "Candidate Committee",
                        "election_year":   ey,
                        "candidate_name":  cname,
                        "treasurer_name":  treas,
                        "city":            utils.clean_name(row.get("Candidate City", "")),
                        "zip":             clean_zip_field(row.get("Candidate Zip Code", "")),
                        "active":          "1" if clean(row.get("Account Status", "")) == "Active" else "0",
                        "state_filer_id":  fid,
                        "raw_file":        path.name,
                        "row_num":         row_num,
                    }

                    if office:
                        fid_office[fid] = office
                    name_to_committee[cname] = campaign_name
                    count += 1
            finally:
                f.close()
            log.file_parsed(path.name, "candidates", count,
                            duration_s=round(time.perf_counter() - ft, 2), bytes=path.stat().st_size)

        # =================== portal_committees.csv ===================
        path = RAW_DIR / "portal_committees.csv"
        if path.exists() and path.stat().st_size > 0 and PHASE in ("", "1"):
            ft, count = time.perf_counter(), 0
            f = _open_portal_csv(path)
            try:
                for row_num, row in enumerate(csv.DictReader(f), start=3):
                    fid = clean(row.get("Filing Entity Id", ""))
                    if not fid:
                        continue
                    ey = clean(row.get("Filing Year", ""))
                    treas = utils.clean_name(
                        f"{clean(row.get('Treasurer First Name', ''))} {clean(row.get('Treasurer Last Name', ''))}"
                    )
                    committees[f"cmte_{fid}_{ey}"] = {
                        "state":           STATE,
                        "committee_name":  utils.clean_name(row.get("Committee Name", "")),
                        "committee_type":  clean(row.get("Committee Type", "")),
                        "election_year":   ey,
                        "candidate_name":  "",
                        "treasurer_name":  treas,
                        "city":            "",
                        "zip":             "",
                        "active":          "1" if clean(row.get("Account Status", "")) == "Active" else "0",
                        "state_filer_id":  fid,
                        "raw_file":        path.name,
                        "row_num":         row_num,
                    }
                    count += 1
            finally:
                f.close()
            log.file_parsed(path.name, "committees", count,
                            duration_s=round(time.perf_counter() - ft, 2), bytes=path.stat().st_size)

        # =================== portal_contributions_{YYYY}.csv ===================
        # NOTE: unlike portal_candidates.csv/portal_committees.csv, these
        # transaction exports have NO junk metadata line — the real header
        # is row 1, so do not use _open_portal_csv() here.
        for path in raw_files("portal_contributions_*.csv"):
            ft, n_cont, n_loan = time.perf_counter(), 0, 0
            f = open(path, encoding="utf-8-sig", errors="replace", newline="")
            try:
                for row_num, row in enumerate(csv.DictReader(f), start=2):
                    amount = parse_amount(row.get("Transaction Amount", ""))
                    if not amount:
                        continue
                    date_    = parse_date(row.get("Transaction Date", ""))
                    fid      = clean(row.get("Filing Entity ID", ""))
                    reg_type = clean(row.get("Registration Type", ""))
                    # NOTE: the portal export's "Election Type"/"Election Year"
                    # headers are swapped relative to their content — the
                    # column labeled "Election Type" actually holds the
                    # 4-digit election year (e.g. "2026"), while the column
                    # labeled "Election Year" holds the type ("Primary",
                    # "General", etc). Confirmed against raw data 2026-06-12.
                    ey       = clean(row.get("Election Type", ""))
                    txn_type = clean(row.get("Transaction Type", ""))
                    fen      = clean(row.get("Filing Entity Name", "")) or clean(row.get("Filing Entity Name ", ""))

                    if reg_type == "Candidate":
                        candidate_name = utils.clean_name(format_name(fen))
                        committee_name = utils.clean_name(row.get("Campaign Name", "")) or candidate_name
                        office = fid_office.get(fid, "")
                        name_to_committee.setdefault(candidate_name, committee_name)
                    else:
                        candidate_name = ""
                        committee_name = utils.clean_name(fen)
                        office = ""

                    contributor_type = clean(row.get("Contributor Type", ""))
                    c_first   = clean(row.get("Contributor First Name", ""))
                    c_last    = clean(row.get("Contributor Last Name", ""))
                    c_company = clean(row.get("Contributor Company Name", ""))
                    if c_first or c_last:
                        contributor_name = utils.clean_name(f"{c_first} {c_last}")
                    else:
                        contributor_name = utils.clean_name(c_company)
                    contributor_city  = utils.clean_name(row.get("Contributor Address City", ""))
                    contributor_state = clean(row.get("Contributor Address State", ""))
                    contributor_zip   = clean_zip_field(row.get("Contributor Address Zip Code", ""))
                    amended   = bool01(row.get("Amended", ""))
                    filing_id = clean(row.get("Transaction Id", ""))

                    if txn_type in LOAN_INCOME_TYPES:
                        loan_w.writerow({
                            "state":              STATE,
                            "committee_name":     committee_name,
                            "original_amount":    amount,
                            "date":               date_,
                            "record_type":        txn_type,
                            "counterparty_name":  contributor_name,
                            "counterparty_city":  contributor_city,
                            "counterparty_state": contributor_state,
                            "counterparty_zip":   contributor_zip,
                            "candidate_name":     candidate_name,
                            "election_year":      ey,
                            "amended":            amended,
                            "filing_id":          filing_id,
                            "raw_file":           path.name,
                            "row_num":            row_num,
                        })
                        n_loan += 1
                    else:
                        cont_w.writerow({
                            "state":             STATE,
                            "committee_name":    committee_name,
                            "amount":            amount,
                            "date":              date_,
                            "transaction_type":  txn_type or "Contribution",
                            "contributor_name":  contributor_name,
                            "contributor_type":  contributor_type,
                            "contributor_city":  contributor_city,
                            "contributor_state": contributor_state,
                            "contributor_zip":   contributor_zip,
                            "employer":          "",
                            "occupation":        "",
                            "candidate_name":    candidate_name,
                            "office":            office,
                            "election_year":     ey,
                            "amended":           amended,
                            "filing_id":         filing_id,
                            "raw_file":          path.name,
                            "row_num":           row_num,
                        })
                        n_cont += 1
            finally:
                f.close()
            log.file_parsed(path.name, "contributions", n_cont,
                            duration_s=round(time.perf_counter() - ft, 2), bytes=path.stat().st_size)
            if n_loan:
                log.file_parsed(path.name, "loans_debts", n_loan, bytes=path.stat().st_size)
            total_contributions += n_cont
            total_loans += n_loan

        # =================== portal_expenditures_{YYYY}.csv ===================
        # NOTE: same as contributions — no junk metadata line, header is row 1.
        for path in raw_files("portal_expenditures_*.csv"):
            ft, n_exp, n_loan = time.perf_counter(), 0, 0
            f = open(path, encoding="utf-8-sig", errors="replace", newline="")
            try:
                for row_num, row in enumerate(csv.DictReader(f), start=2):
                    amount = parse_amount(row.get("Transaction Amount", ""))
                    if not amount:
                        continue
                    date_    = parse_date(row.get("Transaction Date", ""))
                    fid      = clean(row.get("Filing Entity ID", ""))
                    reg_type = clean(row.get("Registration Type", ""))
                    # See note in the contributions loop above — "Election
                    # Type"/"Election Year" headers are swapped vs. content.
                    ey       = clean(row.get("Election Type", ""))
                    txn_type = clean(row.get("Transaction Type", ""))
                    fen      = clean(row.get("Filing Entity Name", "")) or clean(row.get("Filing Entity Name ", ""))

                    if reg_type == "Candidate":
                        candidate_name = utils.clean_name(format_name(fen))
                        committee_name = utils.clean_name(row.get("Campaign Name", "")) or candidate_name
                        office = fid_office.get(fid, "")
                        name_to_committee.setdefault(candidate_name, committee_name)
                    else:
                        candidate_name = ""
                        committee_name = utils.clean_name(fen)
                        office = ""

                    p_first   = clean(row.get("Payee First Name", ""))
                    p_last    = clean(row.get("Payee Last Name", ""))
                    p_company = clean(row.get("Payee Company Name", ""))
                    if p_first or p_last:
                        payee_name = utils.clean_name(f"{p_first} {p_last}")
                    else:
                        payee_name = utils.clean_name(p_company)
                    payee_city  = utils.clean_name(row.get("Payee Address City", ""))
                    payee_state = clean(row.get("Payee Address State", ""))
                    payee_zip   = clean_zip_field(row.get("Payee Address Zip Code", ""))
                    purpose   = clean(row.get("Purpose", ""))
                    category  = clean(row.get("Transaction Sub Type", ""))
                    amended   = bool01(row.get("Amended", ""))
                    filing_id = clean(row.get("Transaction Id", ""))

                    if txn_type in LOAN_EXPENSE_TYPES:
                        loan_w.writerow({
                            "state":              STATE,
                            "committee_name":     committee_name,
                            "original_amount":    amount,
                            "date":               date_,
                            "record_type":        txn_type,
                            "counterparty_name":  payee_name,
                            "counterparty_city":  payee_city,
                            "counterparty_state": payee_state,
                            "counterparty_zip":   payee_zip,
                            "candidate_name":     candidate_name,
                            "election_year":      ey,
                            "amended":            amended,
                            "filing_id":          filing_id,
                            "raw_file":           path.name,
                            "row_num":            row_num,
                        })
                        n_loan += 1
                    else:
                        expn_w.writerow({
                            "state":            STATE,
                            "committee_name":   committee_name,
                            "amount":           amount,
                            "date":             date_,
                            "transaction_type": txn_type or "Expenditure",
                            "payee_name":       payee_name,
                            "purpose":          purpose,
                            "category":         category,
                            "payee_city":       payee_city,
                            "payee_state":      payee_state,
                            "payee_zip":        payee_zip,
                            "candidate_name":   candidate_name,
                            "office":           office,
                            "election_year":    ey,
                            "amended":          amended,
                            "filing_id":        filing_id,
                            "raw_file":         path.name,
                            "row_num":          row_num,
                        })
                        n_exp += 1
            finally:
                f.close()
            log.file_parsed(path.name, "expenditures", n_exp,
                            duration_s=round(time.perf_counter() - ft, 2), bytes=path.stat().st_size)
            if n_loan:
                log.file_parsed(path.name, "loans_debts", n_loan, bytes=path.stat().st_size)
            total_expenditures += n_exp
            total_loans += n_loan

        # =================== electionstats_activity.csv (2020-2022 gap-fill) ===================
        # canvass.sos.idaho.gov "donate"/"spend" activity feed — full pull,
        # 382,124 rows spanning 2020-2024. Per Henry's "dedup is the parser's
        # job" decision: only years ES_YEAR_MIN..ES_YEAR_MAX are emitted here.
        # That window is exactly the gap the current portal (real data from
        # 2023) and the SOS archive (<=2018) don't cover; 2023+ rows are
        # skipped to avoid double-counting against the portal. "file"
        # activity_type rows (filing/registration metadata, no amount) are
        # skipped entirely — they carry no transaction.
        #
        # Row shape (verified against the full 382,124-row pull):
        #   donate: to_entity_type is always "Candidate" or "PAC" (the
        #     recipient committee); from_* describes the contributor.
        #   spend:  from_entity_type is always "Candidate" or "PAC" (the
        #     spending committee); to_* describes the payee.
        ES_YEAR_MIN, ES_YEAR_MAX = 2020, 2022

        ES_DONATE_TYPE_MAP = {
            "Normal":           "Contribution",
            "Unitemized":       "Unitemized Contribution",
            "In-Kind":          "In-Kind Contribution",
            "Return":           "Return Contribution",
            "Cash":             "Contribution – Cash",
            "Credit Card Item": "Contribution – Credit Card Item",
        }
        ES_SPEND_TYPE_MAP = {
            "Normal":                        "Expenditure",
            "Unitemized":                    "Unitemized Expenditure",
            "In-Kind":                       "In-Kind Expenditure",
            "Independent Expenditures":      "Independent Expenditure",
            "Electioneering Communications": "Electioneering Communication",
            "Credit Card Item":              "Expenditure",
        }
        # spend_type values routed to loans_debts.csv.gz instead of expenditures.csv.gz
        ES_LOAN_SPEND_TYPES = {"Loan Payment", "Loan Interest", "Credit Card Payment", "Credit Card Interest/Fee"}
        ES_ENTITY_TYPE_MAP = {"Individual": "Person", "Company": "Company"}

        path = RAW_DIR / "electionstats_activity.csv"
        if path.exists() and path.stat().st_size > 0 and PHASE in ("", "1"):
            ft, n_cont, n_exp, n_loan, n_skipped_year = time.perf_counter(), 0, 0, 0, 0
            with open(path, newline="", encoding="utf-8", errors="replace") as f:
                for row_num, row in enumerate(csv.DictReader(f), start=2):
                    activity = clean(row.get("activity_type", ""))
                    if activity not in ("donate", "spend"):
                        continue
                    amount = parse_amount(row.get("amount", ""))
                    if not amount:
                        continue
                    date_ = parse_date(row.get("date", ""))
                    if not date_:
                        continue
                    year = int(date_[:4])
                    if year < ES_YEAR_MIN or year > ES_YEAR_MAX:
                        n_skipped_year += 1
                        continue

                    elec_year = clean(row.get("elec_year", ""))
                    if elec_year in ("", "0"):
                        elec_year = date_[:4]
                    filing_id = clean(row.get("report_id", ""))

                    if activity == "donate":
                        # to_* = recipient committee/candidate; from_* = contributor
                        to_type = clean(row.get("to_entity_type", ""))
                        if to_type == "Candidate":
                            candidate_name = utils.clean_name(row.get("to_display_name", ""))
                            office = clean(row.get("to_office_name", ""))
                            committee_name = _register_candidate_committee(
                                "es", candidate_name, elec_year,
                                party=clean(row.get("to_party_code", "")),
                                office=office,
                                district=clean(row.get("to_district_name", "")),
                                raw_file=path.name, row_num=row_num,
                            )
                        else:
                            candidate_name, office = "", ""
                            committee_name = utils.clean_name(row.get("to_display_name", ""))
                            _register_pac(committee_name, elec_year, raw_file=path.name, row_num=row_num)

                        from_type         = clean(row.get("from_entity_type", ""))
                        contributor_type  = ES_ENTITY_TYPE_MAP.get(from_type, from_type)
                        contributor_name  = utils.clean_name(row.get("from_display_name", ""))
                        contributor_city  = utils.clean_name(row.get("from_city", ""))
                        contributor_state = clean(row.get("from_state", ""))
                        contributor_zip   = clean_zip_field(row.get("from_zip", ""))

                        donate_type = clean(row.get("donate_type", ""))
                        if donate_type == "Loan":
                            loan_w.writerow({
                                "state":              STATE,
                                "committee_name":     committee_name,
                                "original_amount":    amount,
                                "date":               date_,
                                "record_type":        "Loan Received",
                                "counterparty_name":  contributor_name,
                                "counterparty_city":  contributor_city,
                                "counterparty_state": contributor_state,
                                "counterparty_zip":   contributor_zip,
                                "candidate_name":     candidate_name,
                                "election_year":      elec_year,
                                "amended":            "",
                                "filing_id":          filing_id,
                                "raw_file":           path.name,
                                "row_num":            row_num,
                            })
                            n_loan += 1
                        else:
                            cont_w.writerow({
                                "state":             STATE,
                                "committee_name":    committee_name,
                                "amount":            amount,
                                "date":              date_,
                                "transaction_type":  ES_DONATE_TYPE_MAP.get(donate_type, "Contribution"),
                                "contributor_name":  contributor_name,
                                "contributor_type":  contributor_type,
                                "contributor_city":  contributor_city,
                                "contributor_state": contributor_state,
                                "contributor_zip":   contributor_zip,
                                "employer":          "",
                                "occupation":        "",
                                "candidate_name":    candidate_name,
                                "office":            office,
                                "election_year":     elec_year,
                                "amended":           "",
                                "filing_id":         filing_id,
                                "raw_file":          path.name,
                                "row_num":           row_num,
                            })
                            n_cont += 1

                    else:  # activity == "spend"
                        # from_* = spending committee/candidate; to_* = payee
                        from_type = clean(row.get("from_entity_type", ""))
                        if from_type == "Candidate":
                            candidate_name = utils.clean_name(row.get("from_display_name", ""))
                            office = clean(row.get("from_office_name", ""))
                            committee_name = _register_candidate_committee(
                                "es", candidate_name, elec_year,
                                party=clean(row.get("from_party_code", "")),
                                office=office,
                                district=clean(row.get("from_district_name", "")),
                                raw_file=path.name, row_num=row_num,
                            )
                        else:
                            candidate_name, office = "", ""
                            committee_name = utils.clean_name(row.get("from_display_name", ""))
                            _register_pac(committee_name, elec_year, raw_file=path.name, row_num=row_num)

                        payee_name  = utils.clean_name(row.get("to_display_name", ""))
                        payee_city  = utils.clean_name(row.get("to_city", ""))
                        payee_state = clean(row.get("to_state", ""))
                        payee_zip   = clean_zip_field(row.get("to_zip", ""))

                        spend_type = clean(row.get("spend_type", ""))
                        if spend_type in ES_LOAN_SPEND_TYPES:
                            loan_w.writerow({
                                "state":              STATE,
                                "committee_name":     committee_name,
                                "original_amount":    amount,
                                "date":               date_,
                                "record_type":        spend_type,
                                "counterparty_name":  payee_name,
                                "counterparty_city":  payee_city,
                                "counterparty_state": payee_state,
                                "counterparty_zip":   payee_zip,
                                "candidate_name":     candidate_name,
                                "election_year":      elec_year,
                                "amended":            "",
                                "filing_id":          filing_id,
                                "raw_file":           path.name,
                                "row_num":            row_num,
                            })
                            n_loan += 1
                        else:
                            expn_w.writerow({
                                "state":             STATE,
                                "committee_name":    committee_name,
                                "amount":            amount,
                                "date":              date_,
                                "transaction_type":  ES_SPEND_TYPE_MAP.get(spend_type, "Expenditure"),
                                "payee_name":        payee_name,
                                "purpose":           "",
                                "category":          spend_type,
                                "payee_city":        payee_city,
                                "payee_state":       payee_state,
                                "payee_zip":         payee_zip,
                                "candidate_name":    candidate_name,
                                "office":            office,
                                "election_year":     elec_year,
                                "amended":           "",
                                "filing_id":         filing_id,
                                "raw_file":          path.name,
                                "row_num":           row_num,
                            })
                            n_exp += 1
            log.file_parsed(path.name, "contributions", n_cont,
                            duration_s=round(time.perf_counter() - ft, 2), bytes=path.stat().st_size)
            if n_exp:
                log.file_parsed(path.name, "expenditures", n_exp, bytes=path.stat().st_size)
            if n_loan:
                log.file_parsed(path.name, "loans_debts", n_loan, bytes=path.stat().st_size)
            total_contributions += n_cont
            total_expenditures += n_exp
            total_loans += n_loan

        # =================== Archive: cand_cont (2000-2018) ===================
        for path in raw_files("archive_*_cand_cont.xls*"):
            ft, n_cont, n_loan = time.perf_counter(), 0, 0
            ey = _archive_year(path)
            for row_num, row in _read_rows(path):
                dm = row.get("__datemode__")
                amount = cell_amount(pick(row, *CONTR_AMOUNT))
                if not amount:
                    continue
                date_ = cell_date(pick(row, *CONTR_DATE), dm)

                cand_first = cell_str(pick(row, *NAME_FIRST))
                cand_mid   = cell_str(pick(row, *NAME_MID))
                cand_last  = cell_str(pick(row, *NAME_LAST))
                cand_suf   = cell_str(pick(row, *NAME_SUF))
                cand_name  = utils.clean_name(f"{cand_first} {cand_mid} {cand_last} {cand_suf}")
                party    = cell_str(pick(row, *PARTY))
                office   = cell_str(pick(row, *OFFICE))
                district = cell_str(pick(row, *DISTRICT))
                committee_name = _register_candidate_committee(
                    "arc", cand_name, ey, party=party, office=office, district=district,
                    candidate_first=cand_first, candidate_last=cand_last,
                    raw_file=path.name, row_num=row_num,
                )

                contr_first = cell_str(pick(row, *CONTRIBUTOR_FIRST))
                contr_mid   = cell_str(pick(row, *CONTRIBUTOR_MID))
                contr_last  = cell_str(pick(row, *CONTRIBUTOR_LAST))
                contr_suf   = cell_str(pick(row, *CONTRIBUTOR_SUF))
                contr_entity = cell_str(pick(row, *CONTRIBUTOR_ENTITY))
                if contr_first or contr_mid or contr_suf:
                    contributor_name = utils.clean_name(f"{contr_first} {contr_mid} {contr_last} {contr_suf}")
                elif contr_last:
                    contributor_name = utils.clean_name(contr_last)
                else:
                    contributor_name = utils.clean_name(contr_entity)
                cp = cell_str(pick(row, *CONTR_CP))
                if cp:
                    contributor_type = cp
                elif contr_first:
                    contributor_type = "Person"
                elif contr_last or contr_entity:
                    contributor_type = "Company"
                else:
                    contributor_type = ""
                contributor_city  = utils.clean_name(pick(row, *CONTRIBUTOR_CITY))
                contributor_state = cell_str(pick(row, *CONTRIBUTOR_STATE))
                contributor_zip   = utils.clean_zip(cell_str(pick(row, *CONTRIBUTOR_ZIP)))

                type_code = cell_str(pick(row, *CONTR_TYPE))
                if type_code in ("L", "Loan"):
                    loan_w.writerow({
                        "state":              STATE,
                        "committee_name":     committee_name,
                        "original_amount":    amount,
                        "date":               date_,
                        "record_type":        "Loan Received",
                        "counterparty_name":  contributor_name,
                        "counterparty_city":  contributor_city,
                        "counterparty_state": contributor_state,
                        "counterparty_zip":   contributor_zip,
                        "candidate_name":     cand_name,
                        "election_year":      ey,
                        "amended":            "",
                        "filing_id":          "",
                        "raw_file":           path.name,
                        "row_num":            row_num,
                    })
                    n_loan += 1
                else:
                    txn_type = CONT_TYPE_MAP.get(type_code, "Contribution")
                    cont_w.writerow({
                        "state":             STATE,
                        "committee_name":    committee_name,
                        "amount":            amount,
                        "date":              date_,
                        "transaction_type":  txn_type,
                        "contributor_name":  contributor_name,
                        "contributor_type":  contributor_type,
                        "contributor_city":  contributor_city,
                        "contributor_state": contributor_state,
                        "contributor_zip":   contributor_zip,
                        "employer":          "",
                        "occupation":        "",
                        "candidate_name":    cand_name,
                        "office":            office,
                        "election_year":     ey,
                        "amended":           "",
                        "filing_id":         "",
                        "raw_file":          path.name,
                        "row_num":           row_num,
                    })
                    n_cont += 1
            log.file_parsed(path.name, "contributions", n_cont,
                            duration_s=round(time.perf_counter() - ft, 2), bytes=path.stat().st_size)
            if n_loan:
                log.file_parsed(path.name, "loans_debts", n_loan, bytes=path.stat().st_size)
            total_contributions += n_cont
            total_loans += n_loan

        # =================== Archive: cand_exp (2000-2018) ===================
        for path in raw_files("archive_*_cand_exp.xls*"):
            ft, n_exp, n_loan = time.perf_counter(), 0, 0
            ey = _archive_year(path)
            for row_num, row in _read_rows(path):
                dm = row.get("__datemode__")
                amount = cell_amount(pick(row, *EXP_AMOUNT))
                if not amount:
                    continue
                date_ = cell_date(pick(row, *EXP_DATE), dm)

                cand_first = cell_str(pick(row, *NAME_FIRST))
                cand_mid   = cell_str(pick(row, *NAME_MID))
                cand_last  = cell_str(pick(row, *NAME_LAST))
                cand_suf   = cell_str(pick(row, *NAME_SUF))
                cand_name  = utils.clean_name(f"{cand_first} {cand_mid} {cand_last} {cand_suf}")
                party    = cell_str(pick(row, *PARTY))
                office   = cell_str(pick(row, *OFFICE))
                district = cell_str(pick(row, *DISTRICT))
                committee_name = _register_candidate_committee(
                    "arc", cand_name, ey, party=party, office=office, district=district,
                    candidate_first=cand_first, candidate_last=cand_last,
                    raw_file=path.name, row_num=row_num,
                )

                rec_first = cell_str(pick(row, *RECIP_FIRST))
                rec_mid   = cell_str(pick(row, *RECIP_MID))
                rec_last  = cell_str(pick(row, *RECIP_LAST))
                rec_suf   = cell_str(pick(row, *RECIP_SUF))
                rec_entity = cell_str(pick(row, *RECIP_ENTITY))
                if rec_first or rec_mid or rec_suf:
                    payee_name = utils.clean_name(f"{rec_first} {rec_mid} {rec_last} {rec_suf}")
                elif rec_last:
                    payee_name = utils.clean_name(rec_last)
                else:
                    payee_name = utils.clean_name(rec_entity)
                payee_city  = utils.clean_name(pick(row, *RECIP_CITY))
                payee_state = cell_str(pick(row, *RECIP_STATE))
                payee_zip   = utils.clean_zip(cell_str(pick(row, *RECIP_ZIP)))
                purpose  = cell_str(pick(row, *PURPOSE))
                category = cell_str(pick(row, *CATEGORY))

                type_code = cell_str(pick(row, *EXP_TYPE))
                if type_code == "Repayment":
                    loan_w.writerow({
                        "state":              STATE,
                        "committee_name":     committee_name,
                        "original_amount":    amount,
                        "date":               date_,
                        "record_type":        "Loan Payment",
                        "counterparty_name":  payee_name,
                        "counterparty_city":  payee_city,
                        "counterparty_state": payee_state,
                        "counterparty_zip":   payee_zip,
                        "candidate_name":     cand_name,
                        "election_year":      ey,
                        "amended":            "",
                        "filing_id":          "",
                        "raw_file":           path.name,
                        "row_num":            row_num,
                    })
                    n_loan += 1
                else:
                    txn_type = EXP_TYPE_MAP.get(type_code, "Expenditure")
                    expn_w.writerow({
                        "state":            STATE,
                        "committee_name":   committee_name,
                        "amount":           amount,
                        "date":             date_,
                        "transaction_type": txn_type,
                        "payee_name":       payee_name,
                        "purpose":          purpose,
                        "category":         category,
                        "payee_city":       payee_city,
                        "payee_state":      payee_state,
                        "payee_zip":        payee_zip,
                        "candidate_name":   cand_name,
                        "office":           office,
                        "election_year":    ey,
                        "amended":          "",
                        "filing_id":        "",
                        "raw_file":         path.name,
                        "row_num":          row_num,
                    })
                    n_exp += 1
            log.file_parsed(path.name, "expenditures", n_exp,
                            duration_s=round(time.perf_counter() - ft, 2), bytes=path.stat().st_size)
            if n_loan:
                log.file_parsed(path.name, "loans_debts", n_loan, bytes=path.stat().st_size)
            total_expenditures += n_exp
            total_loans += n_loan

        if PHASE == "1":
            for fh in file_handles:
                fh.close()
            file_handles = []
            with open(CKPT1, "wb") as f:
                pickle.dump({
                    "candidates": candidates,
                    "committees": committees,
                    "name_to_committee": name_to_committee,
                    "fid_office": fid_office,
                    "total_contributions": total_contributions,
                    "total_expenditures": total_expenditures,
                    "total_loans": total_loans,
                }, f)
            duration = round(time.perf_counter() - t0, 1)
            log.info(f"Phase 1 complete in {duration}s; checkpoint written to {CKPT1}")
            log._emit("parse_completed", status="phase1_complete", duration_s=duration,
                      contributions=total_contributions, expenditures=total_expenditures,
                      loans_debts=total_loans, committees=len(committees), candidates=len(candidates))
            return

        # =================== Archive: comm_cont (2000-2018) ===================
        for path in raw_files("archive_*_comm_cont.xls*"):
            ft, n_cont, n_loan = time.perf_counter(), 0, 0
            ey = _archive_year(path)
            for row_num, row in _read_rows(path):
                dm = row.get("__datemode__")
                amount = cell_amount(pick(row, *CONTR_AMOUNT))
                if not amount:
                    continue
                date_ = cell_date(pick(row, *CONTR_DATE), dm)

                committee_name = utils.clean_name(cell_str(pick(row, *COMMITTEE_NAME_COL)))
                _register_pac(committee_name, ey, raw_file=path.name, row_num=row_num)

                contr_first = cell_str(pick(row, *CONTRIBUTOR_FIRST))
                contr_mid   = cell_str(pick(row, *CONTRIBUTOR_MID))
                contr_last  = cell_str(pick(row, *CONTRIBUTOR_LAST))
                contr_suf   = cell_str(pick(row, *CONTRIBUTOR_SUF))
                contr_entity = cell_str(pick(row, *CONTRIBUTOR_ENTITY))
                if contr_first or contr_mid or contr_suf:
                    contributor_name = utils.clean_name(f"{contr_first} {contr_mid} {contr_last} {contr_suf}")
                elif contr_last:
                    contributor_name = utils.clean_name(contr_last)
                else:
                    contributor_name = utils.clean_name(contr_entity)
                cp = cell_str(pick(row, *CONTR_CP))
                if cp:
                    contributor_type = cp
                elif contr_first:
                    contributor_type = "Person"
                elif contr_last or contr_entity:
                    contributor_type = "Company"
                else:
                    contributor_type = ""
                contributor_city  = utils.clean_name(pick(row, *CONTRIBUTOR_CITY))
                contributor_state = cell_str(pick(row, *CONTRIBUTOR_STATE))
                contributor_zip   = utils.clean_zip(cell_str(pick(row, *CONTRIBUTOR_ZIP)))

                type_code = cell_str(pick(row, *CONTR_TYPE))
                if type_code in ("L", "Loan"):
                    loan_w.writerow({
                        "state":              STATE,
                        "committee_name":     committee_name,
                        "original_amount":    amount,
                        "date":               date_,
                        "record_type":        "Loan Received",
                        "counterparty_name":  contributor_name,
                        "counterparty_city":  contributor_city,
                        "counterparty_state": contributor_state,
                        "counterparty_zip":   contributor_zip,
                        "candidate_name":     "",
                        "election_year":      ey,
                        "amended":            "",
                        "filing_id":          "",
                        "raw_file":           path.name,
                        "row_num":            row_num,
                    })
                    n_loan += 1
                else:
                    txn_type = CONT_TYPE_MAP.get(type_code, "Contribution")
                    cont_w.writerow({
                        "state":             STATE,
                        "committee_name":    committee_name,
                        "amount":            amount,
                        "date":              date_,
                        "transaction_type":  txn_type,
                        "contributor_name":  contributor_name,
                        "contributor_type":  contributor_type,
                        "contributor_city":  contributor_city,
                        "contributor_state": contributor_state,
                        "contributor_zip":   contributor_zip,
                        "employer":          "",
                        "occupation":        "",
                        "candidate_name":    "",
                        "office":            "",
                        "election_year":     ey,
                        "amended":           "",
                        "filing_id":         "",
                        "raw_file":          path.name,
                        "row_num":           row_num,
                    })
                    n_cont += 1
            log.file_parsed(path.name, "contributions", n_cont,
                            duration_s=round(time.perf_counter() - ft, 2), bytes=path.stat().st_size)
            if n_loan:
                log.file_parsed(path.name, "loans_debts", n_loan, bytes=path.stat().st_size)
            total_contributions += n_cont
            total_loans += n_loan

        # =================== Legacy 2019 per-filer PDF reports ===================
        pdf_paths = raw_files("legacy_2019_pdfs/*.pdf")
        if pdf_paths:
            ft = time.perf_counter()
            n_cont = n_exp = n_loan = n_files = n_errors = 0
            for path in pdf_paths:
                text = _pdf_to_text(path)
                if not text.strip():
                    n_errors += 1
                    continue
                filer_name, office, district = _pdf_parse_header(text)
                filer_name = utils.clean_name(filer_name)
                if not filer_name:
                    n_errors += 1
                    continue
                n_files += 1
                office_c   = utils.clean_name(office)
                district_c = utils.clean_name(district)

                if office_c:
                    committee_name = _register_candidate_committee(
                        "leg19", filer_name, "2019", office=office_c, district=district_c,
                        raw_file=path.name, row_num=1,
                    )
                    candidate_name = filer_name
                else:
                    committee_name = filer_name
                    candidate_name = ""
                    _register_pac(committee_name, "2019", raw_file=path.name, row_num=1)

                # Per-table running counters — row_num must be a plain
                # integer (BIGINT column; non-numeric values get the whole
                # row silently dropped by tabulate's read_csv_auto). Schedule
                # letter/idx info is not preserved in row_num; raw_file alone
                # already disambiguates the source PDF.
                cont_rn = exp_rn = loan_rn = 0

                for idx, r in enumerate(_pdf_schedule_a(text), start=1):
                    amount = parse_amount(r["amount"])
                    if not amount:
                        continue
                    cont_rn += 1
                    cont_w.writerow({
                        "state": STATE, "committee_name": committee_name, "amount": amount,
                        "date": parse_date(r["date"]), "transaction_type": "Contribution",
                        "contributor_name": utils.clean_name(r["contributor_name"]),
                        "contributor_type": "",
                        "contributor_city": utils.clean_name(r["city"]),
                        "contributor_state": r["state"], "contributor_zip": utils.clean_zip(r["zipc"]),
                        "employer": "", "occupation": "",
                        "candidate_name": candidate_name, "office": office_c,
                        "election_year": "2019", "amended": "", "filing_id": "",
                        "raw_file": path.name, "row_num": cont_rn,
                    })
                    n_cont += 1

                for idx, r in enumerate(_pdf_schedule_b(text), start=1):
                    amount = parse_amount(r["amount"])
                    if not amount:
                        continue
                    exp_rn += 1
                    expn_w.writerow({
                        "state": STATE, "committee_name": committee_name, "amount": amount,
                        "date": parse_date(r["date"]), "transaction_type": "Expenditure",
                        "payee_name": utils.clean_name(r["payee_name"]),
                        "purpose": r["purpose"], "category": r["category"],
                        "payee_city": utils.clean_name(r["city"]),
                        "payee_state": r["state"], "payee_zip": utils.clean_zip(r["zipc"]),
                        "candidate_name": candidate_name, "office": office_c,
                        "election_year": "2019", "amended": "", "filing_id": "",
                        "raw_file": path.name, "row_num": exp_rn,
                    })
                    n_exp += 1

                for idx, r in enumerate(_pdf_schedule_c(text), start=1):
                    amount = parse_amount(r["amount"])
                    if not amount:
                        continue
                    date_ = parse_date(r["date"])
                    cont_rn += 1
                    cont_w.writerow({
                        "state": STATE, "committee_name": committee_name, "amount": amount,
                        "date": date_, "transaction_type": "In-Kind",
                        "contributor_name": utils.clean_name(r["contributor_name"]),
                        "contributor_type": "", "contributor_city": "", "contributor_state": "",
                        "contributor_zip": "", "employer": "", "occupation": "",
                        "candidate_name": candidate_name, "office": office_c,
                        "election_year": "2019", "amended": "", "filing_id": "",
                        "raw_file": path.name, "row_num": cont_rn,
                    })
                    exp_rn += 1
                    expn_w.writerow({
                        "state": STATE, "committee_name": committee_name, "amount": amount,
                        "date": date_, "transaction_type": "In-Kind",
                        "payee_name": utils.clean_name(r["payee_name"]),
                        "purpose": "", "category": r["category"],
                        "payee_city": "", "payee_state": "", "payee_zip": "",
                        "candidate_name": candidate_name, "office": office_c,
                        "election_year": "2019", "amended": "", "filing_id": "",
                        "raw_file": path.name, "row_num": exp_rn,
                    })
                    n_cont += 1
                    n_exp += 1

                for idx, r in enumerate(_pdf_schedule_d(text), start=1):
                    new_amt = parse_amount(r["amount"])
                    if new_amt and float(new_amt) > 0:
                        loan_rn += 1
                        loan_w.writerow({
                            "state": STATE, "committee_name": committee_name,
                            "original_amount": new_amt, "date": parse_date(r["date"]),
                            "record_type": "Loan Received",
                            "counterparty_name": utils.clean_name(r["lender"]),
                            "counterparty_city": "", "counterparty_state": "", "counterparty_zip": "",
                            "candidate_name": candidate_name, "election_year": "2019",
                            "amended": "", "filing_id": "",
                            "raw_file": path.name, "row_num": loan_rn,
                        })
                        n_loan += 1
                    payment = parse_amount(r["payment"])
                    if payment and float(payment) > 0:
                        loan_rn += 1
                        loan_w.writerow({
                            "state": STATE, "committee_name": committee_name,
                            "original_amount": payment,
                            "date": parse_date(r["pmtdate"]) or parse_date(r["date"]),
                            "record_type": "Loan Payment",
                            "counterparty_name": utils.clean_name(r["lender"]),
                            "counterparty_city": "", "counterparty_state": "", "counterparty_zip": "",
                            "candidate_name": candidate_name, "election_year": "2019",
                            "amended": "", "filing_id": "",
                            "raw_file": path.name, "row_num": loan_rn,
                        })
                        n_loan += 1

            total_bytes = sum(p.stat().st_size for p in pdf_paths)
            log.file_parsed("legacy_2019_pdfs/", "contributions", n_cont,
                            duration_s=round(time.perf_counter() - ft, 2), bytes=total_bytes)
            if n_exp:
                log.file_parsed("legacy_2019_pdfs/", "expenditures", n_exp)
            if n_loan:
                log.file_parsed("legacy_2019_pdfs/", "loans_debts", n_loan)
            log.info(f"    legacy_2019_pdfs: {n_files}/{len(pdf_paths)} files parsed, {n_errors} skipped (empty/no header)")
            total_contributions += n_cont
            total_expenditures += n_exp
            total_loans += n_loan

        if PHASE == "2":
            for fh in file_handles:
                fh.close()
            file_handles = []
            with open(CKPT2, "wb") as f:
                pickle.dump({
                    "candidates": candidates,
                    "committees": committees,
                    "name_to_committee": name_to_committee,
                    "fid_office": fid_office,
                    "total_contributions": total_contributions,
                    "total_expenditures": total_expenditures,
                    "total_loans": total_loans,
                }, f)
            duration = round(time.perf_counter() - t0, 1)
            log.info(f"Phase 2 complete in {duration}s; checkpoint written to {CKPT2}")
            log._emit("parse_completed", status="phase2_complete", duration_s=duration,
                      contributions=total_contributions, expenditures=total_expenditures,
                      loans_debts=total_loans, committees=len(committees), candidates=len(candidates))
            return

        # =================== Archive: comm_exp (2000-2018) ===================
        for path in raw_files("archive_*_comm_exp.xls*"):
            ft, n_exp, n_loan = time.perf_counter(), 0, 0
            ey = _archive_year(path)
            for row_num, row in _read_rows(path):
                dm = row.get("__datemode__")
                amount = cell_amount(pick(row, *EXP_AMOUNT))
                if not amount:
                    continue
                date_ = cell_date(pick(row, *EXP_DATE), dm)

                committee_name = utils.clean_name(cell_str(pick(row, *COMMITTEE_NAME_COL)))
                _register_pac(committee_name, ey, raw_file=path.name, row_num=row_num)

                rec_first = cell_str(pick(row, *RECIP_FIRST))
                rec_mid   = cell_str(pick(row, *RECIP_MID))
                rec_last  = cell_str(pick(row, *RECIP_LAST))
                rec_suf   = cell_str(pick(row, *RECIP_SUF))
                rec_entity = cell_str(pick(row, *RECIP_ENTITY))
                if rec_first or rec_mid or rec_suf:
                    payee_name = utils.clean_name(f"{rec_first} {rec_mid} {rec_last} {rec_suf}")
                elif rec_last:
                    payee_name = utils.clean_name(rec_last)
                else:
                    payee_name = utils.clean_name(rec_entity)
                payee_city  = utils.clean_name(pick(row, *RECIP_CITY))
                payee_state = cell_str(pick(row, *RECIP_STATE))
                payee_zip   = utils.clean_zip(cell_str(pick(row, *RECIP_ZIP)))
                purpose  = cell_str(pick(row, *PURPOSE))
                category = cell_str(pick(row, *CATEGORY))

                type_code = cell_str(pick(row, *EXP_TYPE))
                if type_code == "Repayment":
                    loan_w.writerow({
                        "state":              STATE,
                        "committee_name":     committee_name,
                        "original_amount":    amount,
                        "date":               date_,
                        "record_type":        "Loan Payment",
                        "counterparty_name":  payee_name,
                        "counterparty_city":  payee_city,
                        "counterparty_state": payee_state,
                        "counterparty_zip":   payee_zip,
                        "candidate_name":     "",
                        "election_year":      ey,
                        "amended":            "",
                        "filing_id":          "",
                        "raw_file":           path.name,
                        "row_num":            row_num,
                    })
                    n_loan += 1
                else:
                    txn_type = EXP_TYPE_MAP.get(type_code, "Expenditure")
                    expn_w.writerow({
                        "state":            STATE,
                        "committee_name":   committee_name,
                        "amount":           amount,
                        "date":             date_,
                        "transaction_type": txn_type,
                        "payee_name":       payee_name,
                        "purpose":          purpose,
                        "category":         category,
                        "payee_city":       payee_city,
                        "payee_state":      payee_state,
                        "payee_zip":        payee_zip,
                        "candidate_name":   "",
                        "office":           "",
                        "election_year":    ey,
                        "amended":          "",
                        "filing_id":        "",
                        "raw_file":         path.name,
                        "row_num":          row_num,
                    })
                    n_exp += 1
            log.file_parsed(path.name, "expenditures", n_exp,
                            duration_s=round(time.perf_counter() - ft, 2), bytes=path.stat().st_size)
            if n_loan:
                log.file_parsed(path.name, "loans_debts", n_loan, bytes=path.stat().st_size)
            total_expenditures += n_exp
            total_loans += n_loan

        # =================== Flush candidates + committees ===================
        cand_fh, cand_w = open_writer("candidates.csv.gz", C.CANDIDATES)
        cmte_fh, cmte_w = open_writer("committees.csv.gz", C.COMMITTEES)
        file_handles += [cand_fh, cmte_fh]

        for row in candidates.values():
            row["candidate_name"]  = utils.clean_name(row.get("candidate_name", ""))
            row["candidate_first"] = utils.clean_name(row.get("candidate_first", ""))
            row["candidate_last"]  = utils.clean_name(row.get("candidate_last", ""))
            row["office"]          = utils.clean_name(row.get("office", ""))
            row["district"]        = utils.clean_name(row.get("district", ""))
            row["jurisdiction"]    = utils.clean_name(row.get("jurisdiction", ""))
            row["party"]           = utils.clean_name(row.get("party", ""))
            cand_w.writerow(row)

        for row in committees.values():
            row["committee_name"] = utils.clean_name(row.get("committee_name", ""))
            row["candidate_name"] = utils.clean_name(row.get("candidate_name", ""))
            cmte_w.writerow(row)

        # Close handles before person-ID assignment
        for fh in file_handles:
            fh.close()
        file_handles = []

        utils.assign_person_ids(CLEAN_DIR / "candidates.csv.gz", id_model="name_hash")
        utils.assign_committee_person_ids(CLEAN_DIR / "committees.csv.gz",
                                          CLEAN_DIR / "candidates.csv.gz")

        if PHASE == "3":
            for p in (CKPT1, CKPT2):
                try:
                    if p.exists():
                        p.unlink()
                except OSError:
                    pass  # sandbox quirk: unlink sometimes EPERMs; harmless leftover scratch file

        def _out_bytes(name):
            p = CLEAN_DIR / name
            return p.stat().st_size if p.exists() else 0

        log.file_parsed("contributions.csv.gz", "contributions", total_contributions, role="output",
                        bytes=_out_bytes("contributions.csv.gz"))
        log.file_parsed("expenditures.csv.gz",  "expenditures",  total_expenditures,  role="output",
                        bytes=_out_bytes("expenditures.csv.gz"))
        log.file_parsed("loans_debts.csv.gz",   "loans_debts",   total_loans,         role="output",
                        bytes=_out_bytes("loans_debts.csv.gz"))
        log.file_parsed("committees.csv.gz",    "committees",    len(committees),     role="output",
                        bytes=_out_bytes("committees.csv.gz"))
        log.file_parsed("candidates.csv.gz",    "candidates",    len(candidates),     role="output",
                        bytes=_out_bytes("candidates.csv.gz"))

        duration = round(time.perf_counter() - t0, 1)
        log.info(f"Done in {duration}s")
        log._emit("parse_completed", status="completed", duration_s=duration,
                  contributions=total_contributions, expenditures=total_expenditures,
                  loans_debts=total_loans, committees=len(committees), candidates=len(candidates))

    except KeyboardInterrupt:
        log._emit("parse_completed", status="interrupted",
                  duration_s=round(time.perf_counter() - t0, 1),
                  contributions=total_contributions, expenditures=total_expenditures,
                  loans_debts=total_loans, committees=len(committees), candidates=len(candidates))
        raise

    except Exception as e:
        log._emit("parse_completed", status="error",
                  duration_s=round(time.perf_counter() - t0, 1),
                  contributions=total_contributions, expenditures=total_expenditures,
                  loans_debts=total_loans, committees=len(committees), candidates=len(candidates),
                  error_type=type(e).__name__, error=str(e))
        raise

    finally:
        for fh in file_handles:
            try:
                fh.close()
            except Exception:
                pass


# ====== CLI ==================================
if __name__ == "__main__":
    try:
        run()
    except KeyboardInterrupt:
        sys.exit(130)
    except Exception:
        sys.exit(1)
